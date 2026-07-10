"""Sidnavigator för Investeringskalkyl.

Genererar PNG-knappar (monogram-tile + label) och applicerar på alla flikar
i en workbook. Varje knapp placeras i column A med hyperlänk på cell-laget
(robust, fungerar även när Excel inte respekterar img.hyperlink).

Designprinciper:
- Mörk INK-sidopanel ger website-känsla utan att kollidera med innehåll i B+
- Aktiv flik = ljus PAPER-fyllning + 3 px ACCENT vänsterbalk + INK-text
- Inaktiv flik = INK bakgrund + dämpad ljus text
- Monogram-tile (första bokstaven) ersätter ikon — institutionellt, inte emoji
"""
from __future__ import annotations
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from tools.theme import INK, ACCENT, PAPER, MUTED, SURFACE

# ── Sidnav-konfiguration ──────────────────────────────────────────────────────
NAV_ITEMS = [
    ("Försättsblad",       "F"),
    ("Översikt",           "Ö"),
    ("Indata",             "I"),
    ("Kassaflöde",         "K"),
    ("Finansiering",       "$"),
    ("Resultat",           "R"),
    ("Lönsamhetskontroll", "L"),
    ("Beräkningslogik",    "Σ"),
    ("Dokumentation",      "D"),
]

# PNG-dimensioner (2x för crisp render — Excel skalar ner)
SCALE = 2
BTN_W = 168 * SCALE
BTN_H = 38 * SCALE
TILE_W = 38 * SCALE
ACTIVE_BAR_W = 3 * SCALE
PAD_L = 12 * SCALE

# Cell-layout: 1 nav-rad per Excel-rad, börjar på rad 2
ROW_HEIGHT = 30  # points (≈ 40 px) — ger luft kring 38px-knappen
COL_A_WIDTH = 24  # chars ≈ 168 px
START_ROW = 2

ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets" / "nav"


def _hex_to_rgb(h: str) -> tuple[int, int, int]:
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))


def _font(size_px: int, weight: str = "regular") -> ImageFont.FreeTypeFont:
    """Hämta Segoe UI i önskad vikt — fallback DejaVu om Windows-font saknas."""
    candidates = {
        "regular":  ["segoeui.ttf", "DejaVuSans.ttf"],
        "semibold": ["seguisb.ttf", "segoeuib.ttf", "DejaVuSans-Bold.ttf"],
        "bold":     ["segoeuib.ttf", "DejaVuSans-Bold.ttf"],
    }
    for name in candidates[weight]:
        try:
            return ImageFont.truetype(name, size_px)
        except (OSError, IOError):
            continue
    return ImageFont.load_default()


def _render_button(label: str, monogram: str, active: bool) -> Image.Image:
    """Rendera en nav-knapp som PIL-bild."""
    ink_rgb = _hex_to_rgb(INK)
    accent_rgb = _hex_to_rgb(ACCENT)
    paper_rgb = _hex_to_rgb(PAPER)
    muted_rgb = (200, 210, 215)  # ljus dämpad för text på INK-bg

    if active:
        # Aktiv: ljus PAPER-fyllning + accent vänsterbalk + INK-text
        bg = paper_rgb
        text_color = ink_rgb
        tile_bg = ink_rgb
        tile_text = paper_rgb
    else:
        # Inaktiv: INK-bg med dämpad text
        bg = ink_rgb
        text_color = muted_rgb
        tile_bg = (38, 78, 96)  # tydligt ljusare än INK så tile syns konsekvent
        tile_text = (235, 240, 243)  # ljusare för bättre läsbarhet i tile

    img = Image.new("RGB", (BTN_W, BTN_H), bg)
    draw = ImageDraw.Draw(img)

    # Accent-balk till vänster (aktiv)
    if active:
        draw.rectangle([0, 0, ACTIVE_BAR_W, BTN_H], fill=accent_rgb)

    # Monogram-tile
    tile_x0 = ACTIVE_BAR_W if active else 0
    tile_x1 = tile_x0 + TILE_W - (ACTIVE_BAR_W if active else 0)
    draw.rectangle([tile_x0, 0, tile_x1, BTN_H], fill=tile_bg)

    # Monogram-bokstav i tile
    tile_font = _font(int(18 * SCALE), "semibold")
    bbox = draw.textbbox((0, 0), monogram, font=tile_font)
    mg_w = bbox[2] - bbox[0]
    mg_h = bbox[3] - bbox[1]
    mg_x = tile_x0 + (TILE_W - (ACTIVE_BAR_W if active else 0) - mg_w) // 2
    mg_y = (BTN_H - mg_h) // 2 - bbox[1]
    draw.text((mg_x, mg_y), monogram, font=tile_font, fill=tile_text)

    # Label
    weight = "semibold" if active else "regular"
    label_font = _font(int(11 * SCALE), weight)
    lbl_x = tile_x1 + PAD_L
    bbox = draw.textbbox((0, 0), label, font=label_font)
    lbl_h = bbox[3] - bbox[1]
    lbl_y = (BTN_H - lbl_h) // 2 - bbox[1]
    draw.text((lbl_x, lbl_y), label, font=label_font, fill=text_color)

    # Skala ner till slutstorlek för crisp render i Excel
    final = img.resize((BTN_W // SCALE, BTN_H // SCALE), Image.LANCZOS)
    return final


def generate_assets() -> dict[str, Path]:
    """Generera alla PNG-knappar. Returnerar dict {flik|aktiv: path}."""
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}
    for sheet_name, monogram in NAV_ITEMS:
        for active in (True, False):
            state = "active" if active else "inactive"
            safe = (sheet_name.lower()
                    .replace("ä", "a").replace("å", "a").replace("ö", "o"))
            fname = f"{safe}_{state}.png"
            fpath = ASSETS_DIR / fname
            img = _render_button(sheet_name, monogram, active)
            img.save(fpath, "PNG")
            paths[f"{sheet_name}|{state}"] = fpath
    return paths


def apply_to_sheet(ws, current_sheet: str, asset_paths: dict[str, Path]) -> None:
    """Applicera sidnav på en flik. ws = openpyxl Worksheet, current_sheet = dess namn."""
    # Sätt kolumn A bred + bakgrundsfyllning så hela gutter är mörk
    ws.column_dimensions["A"].width = COL_A_WIDTH

    ink_fill = PatternFill("solid", fgColor=INK)

    # Måla kolumn A med INK-bg hela vägen ner till sista innehållsraden —
    # nav-pelaren ska vara obruten över hela fliken (även när content är lång).
    last_nav_row = START_ROW + len(NAV_ITEMS) - 1
    # Min 80 rader: pelaren ska inte sluta mitt på skärmen på korta flikar
    # (kolumn A skrivs aldrig ut — print_area börjar på B överallt)
    bg_end_row = max(last_nav_row + 2, ws.max_row + 5, 80)
    for r in range(1, bg_end_row + 1):
        cell = ws.cell(r, 1)
        cell.fill = ink_fill

    # Header-cell ovan nav: "MENY" eller liknande — håll subtilt
    # (skippa just nu för att inte stjäla fokus)

    # Sätt radhöjd på nav-rader
    for i in range(len(NAV_ITEMS)):
        ws.row_dimensions[START_ROW + i].height = ROW_HEIGHT

    # Lägg in PNG + hyperlänk per nav-item
    for i, (sheet_name, _mono) in enumerate(NAV_ITEMS):
        row = START_ROW + i
        cell = ws.cell(row, 1)
        state = "active" if sheet_name == current_sheet else "inactive"
        path = asset_paths[f"{sheet_name}|{state}"]

        img = XLImage(str(path))
        # Excel-anchor i cellen
        img.anchor = f"A{row}"
        ws.add_image(img)

        # Hyperlänk på cellen bakom bilden — fungerar när användaren klickar i cellområdet
        if sheet_name != current_sheet:
            cell.hyperlink = Hyperlink(
                ref=cell.coordinate,
                location=f"'{sheet_name}'!A1",
                display=sheet_name,
            )
        # Texten i cellen bakom: tom (bilden täcker), men säkerställ ingen synlig text
        if cell.value is None:
            cell.value = ""

    # Frys så kolumn A följer med vid horisontell scroll
    # (vertikal scroll förlorar nav — accepterat trade-off i ren xlsx)
    if ws.freeze_panes is None or ws.freeze_panes == "A1":
        ws.freeze_panes = "B1"
