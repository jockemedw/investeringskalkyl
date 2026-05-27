"""Design system för Investeringskalkyl — Lejonfastigheter Premium-tema.

Palett extraherad från lejonfastigheter.se:
  PRIMARY    #10313E — mörk petrol (banners, sektion-headers)
  SECONDARY  #00657D — turkos (sub-headers, accentband)
  ACCENT     #FAB600 — senapsgul (nyckelresultat: kravhyra, IRR)
  POSITIVE   #00937C — mörkgrön (✓ Uppfyllt)
  NEUTRAL    #F4F6F8 — ljusgrå (zebra, datablock-bg)
  RULE       #DEE2E6 — subtila avgränsare
  TEXT_PRIM  #212529 — brödtext
  TEXT_MUTED #6C757D — etiketter, hjälptext

Typografi: Segoe UI genomgående (Windows-default, humanistisk sans).

Finansmodell-textkodning behålls:
  BLÅ  text — hårdkodade indata
  SVART text — formler
  GRÖN text — cross-sheet-referenser
"""
from __future__ import annotations
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

# ── LF-palett ──────────────────────────────────────────────────────────────
PRIMARY   = "10313E"   # mörk petrol — banners
SECONDARY = "00657D"   # turkos — sub-headers
ACCENT    = "FAB600"   # senapsgul — nyckelresultat
ACCENT_BG = "FFF6D9"   # mjuk gul bg för accentruta
POSITIVE  = "00937C"   # mörkgrön — ✓
POSITIVE_BG = "E0F2EE" # mjuk grön
WHITE     = "FFFFFF"
LIGHT     = "F4F6F8"   # neutral ljusgrå
INPUT_BG  = "EAF3F6"   # mjuk turkos-vit för input-fält
STATUS_BG = "E0F2EE"   # ljusgrön — ✓-celler
RULE_CLR  = "DEE2E6"   # subtil rule
NAVY_TEXT = "10313E"   # rubrik-text mot ljus bg

# ── Textfärger (finansmodell-standard) ────────────────────────────────────
BLUE  = "0B5F7A"   # mörk petrol-blå istället för #0000FF (mer harmoniskt)
GREEN = "00937C"   # cross-sheet — LF-grön
BLACK = "212529"   # brödtext
MUTED = "6C757D"   # sekundärtext

# ── Typografi ──────────────────────────────────────────────────────────────
FONT_FAM  = "Segoe UI"
FONT_FAM_BOLD = "Segoe UI Semibold"

# ── Talformat ──────────────────────────────────────────────────────────────
FMT_KR     = '#,##0 "kr";-#,##0 "kr";"-"'
FMT_KR_AR  = '#,##0 "kr/år";-#,##0 "kr/år";"-"'
FMT_KR_KVM = '#,##0 "kr/kvm";-#,##0 "kr/kvm";"-"'
FMT_PCT    = "0.0%"
FMT_PCT2   = "0.00%"
FMT_INT    = "#,##0"
FMT_AR     = '0 "år"'


# ── Hjälpfunktioner ────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _border_bottom(color: str = RULE_CLR) -> Border:
    side = Side(style="thin", color=color)
    return Border(bottom=side)

def _font(bold=False, size=10, color=BLACK, italic=False) -> Font:
    name = FONT_FAM_BOLD if bold else FONT_FAM
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Stilapplikatorer ───────────────────────────────────────────────────────

def section_header(cell, level: int = 1) -> None:
    """Sektion-rubrik: LF-petrol bg, vit semibold text."""
    if level == 1:
        cell.font = _font(bold=True, size=11, color=WHITE)
        cell.fill = _fill(PRIMARY)
    else:
        cell.font = _font(bold=True, size=10, color=SECONDARY)
        cell.fill = _fill(LIGHT)
    cell.alignment = _align(h="left", v="center", wrap=False)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def title_style(cell) -> None:
    cell.font = Font(name=FONT_FAM_BOLD, bold=True, size=16, color=NAVY_TEXT)
    cell.alignment = _align(h="left", v="center")

def subtitle_style(cell) -> None:
    cell.font = _font(bold=False, size=11, color=MUTED, italic=True)
    cell.alignment = _align(h="left", v="center")

def label_style(cell) -> None:
    cell.font = _font(bold=False, size=10, color=BLACK)
    cell.alignment = _align(h="left", v="center")

def input_style(cell) -> None:
    """Petrol text + mjuk turkos-vit bg — användarens fält."""
    cell.font = _font(bold=False, size=10, color=BLUE)
    cell.fill = _fill(INPUT_BG)
    cell.alignment = _align(h="left", v="center")

def formula_style(cell) -> None:
    cell.font = _font(bold=False, size=10, color=BLACK)
    cell.alignment = _align(h="right", v="center")

def crossref_style(cell) -> None:
    cell.font = _font(bold=False, size=10, color=SECONDARY)
    cell.alignment = _align(h="right", v="center")

def crossref_bold(cell) -> None:
    cell.font = _font(bold=True, size=11, color=NAVY_TEXT)
    cell.alignment = _align(h="right", v="center")

def accent_value(cell) -> None:
    """Senapsgul box för nyckelresultat (bindande kravhyra, IRR)."""
    cell.font = Font(name=FONT_FAM_BOLD, bold=True, size=14, color=NAVY_TEXT)
    cell.fill = _fill(ACCENT_BG)
    cell.alignment = _align(h="right", v="center")

def status_ok_style(cell) -> None:
    cell.font = _font(bold=True, size=10, color=POSITIVE)
    cell.fill = _fill(POSITIVE_BG)
    cell.alignment = _align(h="center", v="center")

def toc_link_style(cell) -> None:
    cell.font = Font(name=FONT_FAM_BOLD, bold=True, size=11,
                     color=SECONDARY, underline="single")
    cell.alignment = _align(h="left", v="center")

def toc_tag_style(cell) -> None:
    cell.font = _font(bold=False, size=10, color=MUTED, italic=True)
    cell.alignment = _align(h="left", v="center")

def toc_desc_style(cell) -> None:
    cell.font = _font(bold=False, size=9, color=MUTED)
    cell.alignment = _align(h="left", v="center", wrap=True)


# ── Oversikt ───────────────────────────────────────────────────────────────

def style_oversikt(ws: Worksheet) -> None:
    # Kolumnbredder — C bredare för senapsgul accent-värde
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 2
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 20

    # Radhöjder
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 16
    for r in [5, 9, 14, 19, 24, 30]:
        ws.row_dimensions[r].height = 20
    for r in [6, 7, 10, 11, 12, 15, 16, 17, 20, 21, 22, 25, 26, 27, 28]:
        ws.row_dimensions[r].height = 18

    # Titel
    title_style(ws["B2"])
    subtitle_style(ws["B3"])

    # A. PROJEKTINFORMATION
    section_header(ws["B5"])
    for r in [6, 7]:
        label_style(ws[f"B{r}"])
        crossref_style(ws[f"C{r}"])
        label_style(ws[f"E{r}"])
    input_style(ws["C6"])   # Fastighet — Fyll i
    input_style(ws["F6"])   # Utförd av — Fyll i
    crossref_style(ws["C7"])
    ws["C7"].number_format = "0"
    crossref_style(ws["F7"])
    ws["F7"].number_format = FMT_AR

    # B. KALKYLANTAGANDEN
    section_header(ws["B9"])
    for r in [10, 11, 12]:
        label_style(ws[f"B{r}"])
        crossref_style(ws[f"C{r}"])
        ws[f"C{r}"].number_format = FMT_PCT
    label_style(ws["E10"])
    label_style(ws["E11"])
    crossref_style(ws["F10"])
    crossref_style(ws["F11"])
    ws["F10"].number_format = FMT_PCT
    ws["F11"].number_format = FMT_PCT

    # C. BINDANDE KRAVHYRA
    section_header(ws["B14"])
    for r in [15, 16, 17]:
        label_style(ws[f"B{r}"])
        crossref_bold(ws[f"C{r}"])
        label_style(ws[f"E{r}"])
        crossref_style(ws[f"F{r}"])
    ws["C15"].number_format = FMT_KR_AR
    ws["C16"].number_format = FMT_KR_KVM
    ws["C17"].number_format = FMT_KR
    ws["F17"].number_format = FMT_KR_KVM

    # D. HYRESSPANN
    section_header(ws["B19"])
    for r in [20, 21, 22]:
        label_style(ws[f"B{r}"])
        crossref_style(ws[f"C{r}"])
        crossref_style(ws[f"E{r}"])
        ws[f"C{r}"].number_format = FMT_KR_AR
        ws[f"E{r}"].number_format = FMT_KR

    # E. LÖNSAMHETSKRAV
    section_header(ws["B24"])
    for r in [25, 26, 27]:
        label_style(ws[f"B{r}"])
        crossref_style(ws[f"C{r}"])
        ws[f"C{r}"].number_format = FMT_KR
        status_ok_style(ws[f"E{r}"])
    ws["C25"].number_format = FMT_KR   # NPV
    ws["C26"].number_format = FMT_KR   # NPV EK
    ws["C27"].number_format = FMT_KR   # MV diff
    label_style(ws["B28"])
    crossref_bold(ws["C28"])
    ws["C28"].number_format = FMT_PCT2
    label_style(ws["E28"])
    crossref_style(ws["F28"])
    ws["F28"].number_format = FMT_PCT2

    # F. INNEHÅLLSFÖRTECKNING
    section_header(ws["B30"])
    for r in [32, 34, 36, 38, 40, 42, 44]:
        toc_link_style(ws[f"B{r}"])
        toc_tag_style(ws[f"C{r}"])
        toc_desc_style(ws[f"E{r}"])
        ws.row_dimensions[r].height = 18


# ── Indata (sektion-rubriker och input-celler) ────────────────────────────

def style_indata(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 16

    # Kända sektion-rubriker i iter 8 Indata
    HEADERS = [3, 15, 24, 33, 51, 60, 67]
    for r in HEADERS:
        cell = ws[f"B{r}"]
        if cell.value:
            section_header(cell, level=2)

    # Input-celler (C-kolumnen, de flesta rader med siffervärden)
    for r in range(4, 70):
        c = ws[f"C{r}"]
        if c.value is not None and not (isinstance(c.value, str) and c.value.startswith("=")):
            input_style(c)


# ── Resultat (sektion-rubriker) ───────────────────────────────────────────

def style_resultat(ws: Worksheet) -> None:
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    for r in range(1, ws.max_row + 1):
        cell = ws[f"B{r}"]
        if isinstance(cell.value, str) and cell.value.isupper() and len(cell.value) > 4:
            section_header(cell, level=2)
