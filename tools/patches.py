"""Utility-funktioner för kirurgisk redigering av iter-xlsx (TECH "kirurgiskt").

Designprincip: arbete sker som delta från senaste iter-fil. Varje patch-funktion
modifierar en lastad workbook utan att röra orelaterad struktur. Spara bara via
save_iter() som verifierar att xlsx fortfarande är giltig.
"""
from __future__ import annotations
from pathlib import Path
from typing import Iterable
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.hyperlink import Hyperlink

ROOT = Path(__file__).resolve().parent.parent
ITER8 = ROOT / "Investeringskalkyl_iter8.xlsx"


def load_iter(path: Path | str = ITER8) -> Workbook:
    """Ladda iter-xlsx för redigering. data_only=False bevarar formler."""
    return load_workbook(Path(path), data_only=False)


def save_iter(wb: Workbook, out_path: Path | str) -> Path:
    """Spara med strukturvalidering."""
    out = Path(out_path)
    wb.save(out)
    # Round-trip test: kan filen öppnas igen?
    test = load_workbook(out, data_only=False)
    assert len(test.sheetnames) == 8, f"Förväntade 8 flikar, fick {len(test.sheetnames)}"
    return out


def set_value(ws: Worksheet, coord: str, value) -> None:
    """Sätt värde utan att röra cellens stil/format."""
    ws[coord] = value


def rename_label(ws: Worksheet, coord: str, new_text: str) -> None:
    """Byt etikett-text i en cell. Stilen bevaras automatiskt av openpyxl."""
    ws[coord] = new_text


def replace_in_cells(
    ws: Worksheet,
    coords: Iterable[str],
    old: str,
    new: str,
) -> int:
    """Sök/ersätt textfragment i utvalda celler. Returnerar antalet ändringar."""
    n = 0
    for coord in coords:
        v = ws[coord].value
        if isinstance(v, str) and old in v:
            ws[coord] = v.replace(old, new)
            n += 1
    return n


def add_internal_hyperlink(
    ws: Worksheet,
    coord: str,
    target_sheet: str,
    target_cell: str = "A1",
    display: str | None = None,
) -> None:
    """Lägg till intern hyperlänk (TECH §3 — Hyperlink-objekt, INTE sträng)."""
    cell = ws[coord]
    location = f"'{target_sheet}'!{target_cell}"
    cell.hyperlink = Hyperlink(
        ref=cell.coordinate,
        location=location,
        display=display or str(cell.value or ""),
    )
    cell.style = "Hyperlink"


def list_formula_cells(ws: Worksheet) -> list[tuple[str, str]]:
    """Returnerar [(coord, formula), ...] för alla formelceller — användbart för diff."""
    out = []
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                out.append((c.coordinate, c.value))
    return out


def diff_workbooks(path_a: Path | str, path_b: Path | str) -> dict:
    """Bit-för-bit-diff på cellnivå. Returnerar {sheet: [(coord, a_val, b_val), ...]}."""
    wa = load_workbook(Path(path_a), data_only=False)
    wb = load_workbook(Path(path_b), data_only=False)
    diffs: dict = {}
    sheets_a, sheets_b = set(wa.sheetnames), set(wb.sheetnames)
    if sheets_a != sheets_b:
        diffs["__sheets__"] = (sorted(sheets_a), sorted(sheets_b))
    for name in sheets_a & sheets_b:
        ws_a, ws_b = wa[name], wb[name]
        rows_max = max(ws_a.max_row, ws_b.max_row)
        cols_max = max(ws_a.max_column, ws_b.max_column)
        sheet_diffs = []
        for r in range(1, rows_max + 1):
            for c in range(1, cols_max + 1):
                va = ws_a.cell(r, c).value
                vb = ws_b.cell(r, c).value
                if va != vb:
                    sheet_diffs.append((ws_a.cell(r, c).coordinate, va, vb))
        if sheet_diffs:
            diffs[name] = sheet_diffs
    return diffs
