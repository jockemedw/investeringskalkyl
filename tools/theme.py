"""Design-arkitektur för Investeringskalkyl — Lejonfastigheter.

Tight system: 1 familj, 4 storlekar, 6 färger, 6 cell-roller.
All cellstyling går genom apply(cell, role).

Roller:
  title    — D / Semibold / WHITE on Ink    (banner)
  section  — H / Semibold / WHITE on Ink    (sektion-headers)
  label    — B / Regular / Muted            (etiketter)
  value    — B / Regular / Ink              (datavärden)
  hero     — D / Semibold / Ink on Accent   (nyckelresultat)
  status   — B / Semibold / Positive on tint (✓ Uppfyllt)
  caption  — C / Regular / Muted            (hjälptext, fotnoter)
"""
from __future__ import annotations
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell import Cell

# ── Färgtokens (LF-palett) ─────────────────────────────────────────────────
INK      = "10313E"   # primärtext + brand
MUTED    = "6C757D"   # sekundärtext
PAPER    = "FFFFFF"   # bakgrund
SURFACE  = "F4F6F8"   # zebra / alt-bg
ACCENT   = "FAB600"   # hero
POSITIVE = "00937C"   # ✓
RULE     = "DEE2E6"   # bottom-rule

# Härledda tints (för status-bg)
POSITIVE_TINT = "E0F2EE"

# ── Typ-skala ──────────────────────────────────────────────────────────────
FAMILY = "Segoe UI"

SIZE_D = 16   # Display — banner
SIZE_HERO = 13  # Hero — nyckelresultat (en aning mindre än D)
SIZE_H = 11   # Heading — sektion
SIZE_B = 10   # Body — data
SIZE_C = 9    # Caption — fotnot/sidfot

# ── Roller (font + fill + alignment) ──────────────────────────────────────
# Varje roll = (size, bold, font_color, fill_color_or_None, h_align, indent)
ROLES = {
    "title":   (SIZE_D, True,  PAPER, INK,           "left",   1),
    "section": (SIZE_H, True,  PAPER, INK,           "left",   1),
    "label":   (SIZE_B, False, MUTED, None,          "left",   1),
    "value":   (SIZE_B, False, INK,   None,          "right",  0),
    "hero":    (SIZE_HERO, True, INK,  ACCENT,        "right",  1),
    "hero_label": (SIZE_B, True, INK, None,          "left",   1),
    "status":  (SIZE_B, True,  POSITIVE, POSITIVE_TINT, "center", 0),
    "caption": (SIZE_C, False, MUTED, None,          "left",   1),
    "caption_right": (SIZE_C, False, MUTED, None,    "right",  0),
}

# ── Radhöjder ──────────────────────────────────────────────────────────────
ROW_BANNER  = 36
ROW_SECTION = 24
ROW_DATA    = 20
ROW_HERO    = 32
ROW_GAP     = 12

# ── Talformat ──────────────────────────────────────────────────────────────
FMT_KR     = '#,##0 "kr";-#,##0 "kr";"–"'
FMT_KR_AR  = '#,##0 "kr/år";-#,##0 "kr/år";"–"'
FMT_KR_KVM = '#,##0 "kr/kvm";-#,##0 "kr/kvm";"–"'
FMT_INT    = '#,##0;-#,##0;"–"'
FMT_PCT1   = "0.0%"
FMT_PCT2   = "0.00%"
FMT_AR     = '0 "år"'
FMT_M2     = '#,##0 "m²"'


# ── Apply ──────────────────────────────────────────────────────────────────

def apply(cell, role: str) -> None:
    """Sätt all styling för en cell enligt rollens specifikation.

    Roller är de enda godkända stilarna — använd inga inline-Font()-anrop.
    """
    if role not in ROLES:
        raise ValueError(f"Okänd roll: {role}. Tillgängliga: {list(ROLES)}")
    size, bold, fcolor, bgcolor, halign, indent = ROLES[role]
    cell.font = Font(name=FAMILY, size=size, bold=bold, color=fcolor)
    if bgcolor:
        cell.fill = PatternFill("solid", fgColor=bgcolor)
    else:
        cell.fill = PatternFill(fill_type=None)
    cell.alignment = Alignment(horizontal=halign, vertical="center", indent=indent)


def bottom_rule(cell) -> None:
    """Tunn ljusgrå avgränsare under sista raden i en sektion."""
    cell.border = Border(bottom=Side(style="thin", color=RULE))


def clear_format(cell) -> None:
    """Nollställ allt — användbar innan apply() på celler med arv från iter8."""
    cell.font = Font(name=FAMILY, size=SIZE_B, color=INK)
    cell.fill = PatternFill(fill_type=None)
    cell.border = Border()
    cell.alignment = Alignment(horizontal="left", vertical="center")
