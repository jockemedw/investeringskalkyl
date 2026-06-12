"""Regressionsgate för Investeringskalkyl_v2.xlsx.

Mappning mot tests/regression.py: IDENTITET — v2 replayar motorflikarna med
samma celladresser (Resultat!D14, Lönsamhetskontroll!C45, Indata!C65),
så baslinjetestet återanvänds oförändrat. Därutöver: nollfelsscan
(inga #REF!/#DIV/0!/#VALUE!/#NAME?/#N/A i cached values).
"""
from __future__ import annotations
import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))
V2 = ROOT / "build" / "oneshot" / "Investeringskalkyl_v2.xlsx"

from openpyxl import load_workbook
from tests.regression import check_baseline, EXPECTED

ERROR_VALUES = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")


def scan_errors(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value in ERROR_VALUES:
                    hits.append(f"{ws.title}!{c.coordinate} = {c.value}")
    return hits


def main() -> int:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else V2
    print(f"Regression v2 mot: {target.name}")

    res = check_baseline(target)
    if res["rent"] is not None:
        print(f"  Bindande kravhyra:  {res['rent']:>15,.2f} kr/år   (väntade {EXPECTED['binding_rent_kr_per_year']:,})")
    if res["irr"] is not None:
        print(f"  Faktisk IRR EK:     {res['irr']:>15.4%}              (väntade {EXPECTED['actual_equity_irr']:.2%})")
    if res["margin"] is not None:
        print(f"  Marginal mot krav:  {res['margin']*100:>+14.2f} pp           (väntade +{EXPECTED['irr_margin_pp']*100:.2f} pp)")

    errors = scan_errors(target)
    if errors:
        print(f"\n✗ FORMELFEL ({len(errors)}):")
        for e in errors[:30]:
            print(f"  - {e}")

    if res["ok"] and not errors:
        print("\n✓ Regression v2 OK (baslinje + nollfelsscan)")
        return 0
    print("\n✗ REGRESSION v2 FAILED:")
    for f in res["fails"]:
        print(f"  - {f}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
