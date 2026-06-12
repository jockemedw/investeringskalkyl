"""Extrahera komplett spec från build/iter9.xlsx → build/oneshot/spec_iter9.json.

Fångar per flik: celler (formel/värde/format/stil-id), merges, kolumnbredder,
radhöjder (+outline), page setup, print_area, freeze panes, bilder, hyperlänkar.
Spec:en är byggunderlaget för build_v2.py — iter9 är formelspec (ONESHOT).
"""
from __future__ import annotations
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent
SRC = ROOT / "build" / "iter9.xlsx"
OUT = ROOT / "build" / "oneshot" / "spec_iter9.json"


def cell_style(c) -> dict:
    f = c.font
    al = c.alignment
    fill = c.fill
    b = c.border
    d = {}
    if f and (f.name or f.sz or f.b or f.color):
        d["font"] = {
            "name": f.name, "size": f.sz, "bold": bool(f.b), "italic": bool(f.i),
            "color": f.color.rgb if (f.color and f.color.type == "rgb") else None,
        }
    if fill and fill.patternType == "solid":
        fg = fill.fgColor
        d["fill"] = fg.rgb if fg.type == "rgb" else str(fg.value)
    if al and (al.horizontal or al.vertical or al.wrapText or al.indent):
        d["align"] = {
            "h": al.horizontal, "v": al.vertical,
            "wrap": bool(al.wrapText), "indent": al.indent or 0,
        }
    sides = {}
    for side_name in ("top", "bottom", "left", "right"):
        s = getattr(b, side_name, None)
        if s and s.style:
            sides[side_name] = {
                "style": s.style,
                "color": s.color.rgb if (s.color and s.color.type == "rgb") else None,
            }
    if sides:
        d["border"] = sides
    if c.number_format and c.number_format != "General":
        d["numfmt"] = c.number_format
    return d


def main() -> int:
    wb = openpyxl.load_workbook(SRC, data_only=False)
    wbv = openpyxl.load_workbook(SRC, data_only=True)
    spec: dict = {"source": str(SRC.name), "sheets": []}

    for ws in wb.worksheets:
        wsv = wbv[ws.title]
        ps = ws.page_setup
        pm = ws.page_margins
        sheet: dict = {
            "name": ws.title,
            "state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "freeze": ws.freeze_panes,
            "print_area": ws.print_area,
            "page_setup": {
                "orientation": ps.orientation,
                "fitToWidth": ps.fitToWidth,
                "fitToHeight": ps.fitToHeight,
                "fitToPage": ws.sheet_properties.pageSetUpPr.fitToPage
                if ws.sheet_properties.pageSetUpPr else None,
                "paperSize": ps.paperSize,
                "scale": ps.scale,
            },
            "margins": {k: getattr(pm, k) for k in
                        ("left", "right", "top", "bottom", "header", "footer")},
            "show_gridlines": ws.sheet_view.showGridLines,
            "merges": [str(m) for m in ws.merged_cells.ranges],
            "col_widths": {
                k: {"width": v.width, "hidden": bool(v.hidden)}
                for k, v in ws.column_dimensions.items()
                if v.width is not None or v.hidden
            },
            "row_heights": {
                str(k): {"height": v.height, "outline": v.outlineLevel,
                         "hidden": bool(v.hidden), "collapsed": bool(v.collapsed)}
                for k, v in ws.row_dimensions.items()
                if v.height is not None or v.outlineLevel or v.hidden or v.collapsed
            },
            "tab_color": ws.sheet_properties.tabColor.rgb
            if (ws.sheet_properties.tabColor and
                ws.sheet_properties.tabColor.type == "rgb") else None,
            "validations": [
                {"type": d.type, "formula1": d.formula1, "allow_blank": bool(d.allowBlank),
                 "sqref": str(d.sqref)}
                for d in ws.data_validations.dataValidation
            ],
            "comments": [
                {"ref": c.coordinate, "text": c.comment.text, "author": c.comment.author}
                for row in ws.iter_rows() for c in row if c.comment
            ],
            "images": [
                {"anchor": getattr(im.anchor, "_from", None) and {
                    "col": im.anchor._from.col, "row": im.anchor._from.row},
                 "anchor_type": type(im.anchor).__name__,
                 "path": getattr(im, "path", None),
                 "width": im.width, "height": im.height}
                for im in getattr(ws, "_images", [])
            ],
            "cells": [],
        }
        for row in ws.iter_rows():
            for c in row:
                if c.value is None and not cell_style(c):
                    continue
                vc = wsv[c.coordinate]
                entry: dict = {"ref": c.coordinate}
                if c.value is not None:
                    if isinstance(c.value, str) and c.value.startswith("="):
                        entry["f"] = c.value
                        entry["cached"] = vc.value if not isinstance(vc.value, bytes) else None
                    else:
                        entry["v"] = c.value
                st = cell_style(c)
                if st:
                    entry["style"] = st
                if c.hyperlink is not None:
                    entry["link"] = {"location": c.hyperlink.location,
                                     "display": c.hyperlink.display}
                if entry.keys() - {"ref"}:
                    sheet["cells"].append(entry)
        spec["sheets"].append(sheet)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(spec, ensure_ascii=False, indent=1, default=str),
                   encoding="utf-8")
    n_cells = sum(len(s["cells"]) for s in spec["sheets"])
    n_formulas = sum(1 for s in spec["sheets"] for c in s["cells"] if "f" in c)
    print(f"OK: {len(spec['sheets'])} flikar, {n_cells} celler, {n_formulas} formler -> {OUT}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
