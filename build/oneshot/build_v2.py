"""ONESHOT build_v2 — bygger Investeringskalkyl_v2.xlsx från noll.

Pipeline:
  1. Replay av spec_iter9.json (motorflikar 1:1 — identiska celladresser,
     så tests/regression.py fungerar oförändrat mot v2-filen)
  2. v2-moduler (Översikt-redesign, tkr-fix, Grafer, locale-fix, sidnav)
  3. Spara → XML-patch (summaryBelow) → recalc (Excel COM) → regression

Kör:  python build/oneshot/build_v2.py [--no-recalc]
"""
from __future__ import annotations
import io
import json
import re
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
import zipfile
import shutil
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.comments import Comment

SPEC_PATH = ROOT / "build" / "oneshot" / "spec_iter9.json"
OUT = ROOT / "build" / "oneshot" / "Investeringskalkyl_v2.xlsx"

_HEX = re.compile(r"^[0-9A-Fa-f]{8}$")
_font_cache: dict[str, Font] = {}
_fill_cache: dict[str, PatternFill] = {}
_border_cache: dict[str, Border] = {}
_align_cache: dict[str, Alignment] = {}


def _mkfont(d: dict) -> Font:
    key = json.dumps(d, sort_keys=True)
    if key not in _font_cache:
        color = d.get("color")
        _font_cache[key] = Font(
            name=d.get("name"), size=d.get("size"), bold=d.get("bold"),
            italic=d.get("italic"),
            color=color if (color and _HEX.match(color)) else None,
        )
    return _font_cache[key]


def _mkfill(rgb: str) -> PatternFill | None:
    if not (rgb and _HEX.match(rgb)) or rgb == "00000000":
        return None
    if rgb not in _fill_cache:
        _fill_cache[rgb] = PatternFill("solid", fgColor=rgb)
    return _fill_cache[rgb]


def _mkborder(d: dict) -> Border:
    key = json.dumps(d, sort_keys=True)
    if key not in _border_cache:
        sides = {}
        for name, sd in d.items():
            color = sd.get("color")
            sides[name] = Side(
                style=sd["style"],
                color=color if (color and _HEX.match(color)) else None,
            )
        _border_cache[key] = Border(**sides)
    return _border_cache[key]


def _mkalign(d: dict) -> Alignment:
    key = json.dumps(d, sort_keys=True)
    if key not in _align_cache:
        _align_cache[key] = Alignment(
            horizontal=d.get("h"), vertical=d.get("v"),
            wrap_text=d.get("wrap") or None, indent=d.get("indent") or 0,
        )
    return _align_cache[key]


def replay_sheet(ws, s: dict) -> None:
    for c in s["cells"]:
        cell = ws[c["ref"]]
        if "f" in c:
            cell.value = c["f"]
        elif "v" in c:
            cell.value = c["v"]
        st = c.get("style")
        if st:
            if "font" in st:
                cell.font = _mkfont(st["font"])
            if "fill" in st:
                fill = _mkfill(st["fill"])
                if fill:
                    cell.fill = fill
            if "align" in st:
                cell.alignment = _mkalign(st["align"])
            if "border" in st:
                cell.border = _mkborder(st["border"])
            if "numfmt" in st:
                cell.number_format = st["numfmt"]
        if "link" in c:
            cell.hyperlink = Hyperlink(
                ref=c["ref"], location=c["link"]["location"],
                display=c["link"].get("display"),
            )

    for m in s["merges"]:
        ws.merge_cells(m)

    for col, d in s["col_widths"].items():
        cd = ws.column_dimensions[col]
        if d["width"] is not None:
            cd.width = d["width"]
        cd.hidden = d["hidden"]

    for row, d in s["row_heights"].items():
        rd = ws.row_dimensions[int(row)]
        if d["height"] is not None:
            rd.height = d["height"]
        if d["outline"]:
            rd.outlineLevel = d["outline"]
        rd.hidden = d["hidden"]
        if d.get("collapsed"):
            rd.collapsed = True

    if s["freeze"]:
        ws.freeze_panes = s["freeze"]

    if s["print_area"]:
        ws.print_area = s["print_area"].split("!")[-1].replace("$", "")

    ps = s["page_setup"]
    ws.page_setup.orientation = ps["orientation"]
    if ps["fitToWidth"] is not None:
        ws.page_setup.fitToWidth = ps["fitToWidth"]
    if ps["fitToHeight"] is not None:
        ws.page_setup.fitToHeight = ps["fitToHeight"]
    if ps["paperSize"] is not None:
        ws.page_setup.paperSize = ps["paperSize"]
    if ps["fitToPage"]:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    pm = s["margins"]
    for k, v in pm.items():
        setattr(ws.page_margins, k, v)

    ws.sheet_view.showGridLines = s["show_gridlines"]

    for dv in s.get("validations", []):
        d = DataValidation(type=dv["type"], formula1=dv["formula1"],
                           allow_blank=dv["allow_blank"])
        for part in dv["sqref"].split():
            d.add(part)
        ws.add_data_validation(d)

    for cm in s.get("comments", []):
        ws[cm["ref"]].comment = Comment(cm["text"], cm["author"] or "Lejonfastigheter")

    if s.get("tab_color"):
        ws.sheet_properties.tabColor = s["tab_color"]


def patch_summary_below(xlsx_path: Path, sheet_indices: list[int]) -> None:
    """outlinePr summaryBelow='0' via direkt XML-patch (TECH_NOTES §1)."""
    names = [f"sheet{i + 1}.xml" for i in sheet_indices]
    tmp = xlsx_path.with_suffix(".xlsx.tmp")
    with zipfile.ZipFile(xlsx_path, "r") as zin, \
         zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            data = zin.read(item)
            if item.split("/")[-1] in names:
                text = data.decode("utf-8")
                if "<outlinePr" in text:
                    text = re.sub(r'<outlinePr([^/>]*?)\s*summaryBelow="[01]"',
                                  r"<outlinePr\1", text)
                    text = text.replace("<outlinePr", '<outlinePr summaryBelow="0"', 1)
                elif "<sheetPr" in text:
                    if "<sheetPr/>" in text:
                        text = text.replace(
                            "<sheetPr/>",
                            '<sheetPr><outlinePr summaryBelow="0"/></sheetPr>', 1)
                    elif "<sheetPr>" in text:
                        text = text.replace(
                            "<sheetPr>",
                            '<sheetPr><outlinePr summaryBelow="0"/>', 1)
                    else:  # <sheetPr attr=..>
                        text = re.sub(r"<sheetPr([^/>]*)>",
                                      r'<sheetPr\1><outlinePr summaryBelow="0"/>',
                                      text, count=1)
                else:
                    text = text.replace(
                        "<sheetFormatPr",
                        '<sheetPr><outlinePr summaryBelow="0"/></sheetPr><sheetFormatPr',
                        1)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp, xlsx_path)


def build(no_recalc: bool = False) -> int:
    spec = json.loads(SPEC_PATH.read_text(encoding="utf-8"))
    wb = Workbook()
    wb.remove(wb.active)

    for s in spec["sheets"]:
        ws = wb.create_sheet(s["name"])
        replay_sheet(ws, s)
        print(f"  replay: {s['name']} ({len(s['cells'])} celler)")

    # ── v2-moduler (aktiveras i milstolpe 3) ────────────────────────────────
    import v2_rounds
    v2_rounds.apply_all(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)

    # Index i wb-ordning (sheetN.xml följer skapandeordningen — Grafer skiftar den)
    outline_sheets = [i for i, ws in enumerate(wb.worksheets)
                      if any(rd.outlineLevel for rd in ws.row_dimensions.values())]
    if outline_sheets:
        patch_summary_below(OUT, outline_sheets)
        print(f"  XML-patch summaryBelow: sheets {outline_sheets}")

    v2_rounds.post_save(OUT)

    if not no_recalc:
        from tools.recalc import recalc_excel
        print("  recalc (Excel COM)...")
        recalc_excel(OUT, timeout=180)
        import subprocess
        r = subprocess.run([sys.executable, str(ROOT / "build" / "oneshot" / "regression_v2.py")],
                           capture_output=True, text=True, encoding="utf-8")
        print(r.stdout)
        if r.returncode != 0:
            print(r.stderr)
            return 1
    print(f"OK -> {OUT}")
    return 0


if __name__ == "__main__":
    sys.path.insert(0, str(Path(__file__).parent))
    raise SystemExit(build(no_recalc="--no-recalc" in sys.argv))
