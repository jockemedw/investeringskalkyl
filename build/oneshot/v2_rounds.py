"""v2-förbättringsmoduler ovanpå spec-replayen. En funktion per V2-beslut (ONESHOT.md).

Ordning i apply_all spelar roll: Översikt återskapas, Grafer skapas,
sidnav appliceras sist (målar kolumn A på alla flikar).
"""
from __future__ import annotations
import json
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent
import sys
sys.path.insert(0, str(ROOT))

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.properties import PageSetupProperties

from tools.theme import (
    INK, MUTED, PAPER, SURFACE, ACCENT, POSITIVE, RULE, POSITIVE_TINT,
    FAMILY, apply, bottom_rule, FMT_KR, FMT_KR_KVM, FMT_INT, FMT_PCT1,
    FMT_PCT2, FMT_AR, FMT_M2,
)

SPEC = json.loads((Path(__file__).parent / "spec_iter9.json").read_text(encoding="utf-8"))
ACCENT_TINT = "FCF3CF"   # ljus gul (iter9:s spann-celler)


# ── F-3 ──────────────────────────────────────────────────────────────────────

def fix_iter9_name_errors(wb) -> None:
    """F-3: två pedagogiska anteckningar i iter9 är inskrivna som formler
    ('=-npv_0 / b', '=max av de tre') → #NAME? i produktionsfilen.
    v2 lagrar dem som text — samma innehåll, inget formelfel."""
    ws = wb["Beräkningslogik"]
    # OBS: får inte börja med "=" — openpyxl skriver det som formel igen.
    ws["D18"].value = "dvs −NPV₀ / b"
    ws["D30"].value = "dvs max av de tre"


# ── V2-04: locale-säkra TEXT()-ersättningar ─────────────────────────────────

def round_locale_fix(wb) -> None:
    """F-1: TEXT(x,"#,##0")-mönster ger skräp i svensk Excel (formatsträngen
    tolkas per UI-locale). Översiktens förekomster försvinner med redesignen;
    här fixas Försättsblad + Resultat."""
    fs = wb["Försättsblad"]
    fs["E63"].value = '="Krav "&ROUND(Indata!C65*100,1)&" %"'
    fs["C73"].value = (
        '=IF(Indata!D80=0,"Neutral (extern yield rakt av)",'
        'IF(Indata!D80>0,"+","")&ROUND(Indata!D80*100,2)&" pp · justerad yield "'
        '&ROUND(Indata!D81*100,2)&" %")'
    )
    rs = wb["Resultat"]
    rs["C19"].value = (
        "Avtalet bör ange hyresspannet ovan (lägsta–högsta utfall). "
        "Slutlig hyra fastställs när investeringsutfallet är känt vid projektavslut."
    )


def round_forsattsblad_polish(wb) -> None:
    """Försättsblad: kolumn D/E saknade bredd (8,43) → ######## i
    PROJEKTBESKRIVNING-tabellens beloppskolumner. Kalkylstart hade
    årsformat ('2026 år'). Etikett G60 höggs vid kolumnkanten."""
    ws = wb["Försättsblad"]
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws["H8"].number_format = "0"
    ws["G60"] = "Avkastningskrav EK"


# ── V2-02/V2-03: Översikt — redesign till 1-sidigt beslutsdokument ──────────

def _kpi_value(ws, ref, fmt, color=INK):
    c = ws[ref]
    c.font = Font(name="Segoe UI Light", size=26, color=color)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.number_format = fmt
    return c


def round_indata_polish(wb) -> None:
    """Indata: Kalkylstart visade '2 026 år' (tusentalsformat på årtal),
    ######## i Investering-kolumnen (R utan bredd), dubbla procenttecken
    (värdecellen %-formaterad OCH separat enhetskolumn '%')."""
    ws = wb["Indata"]
    ws["C5"].number_format = "0"
    ws.column_dimensions["R"].width = 14
    ws.column_dimensions["Q"].width = 12
    ws.column_dimensions["M"].width = 13   # 'Hyra efter avtal'-rubriken kläms
    for ref in ("D6", "D7", "D8", "D9", "D10", "D11", "D12", "D13",
                "D18", "D54", "D61", "D62", "D63", "D64", "D65", "D68"):
        ws[ref].value = None
    # Sektion 4: rubrikerna 'DoU-höjning från år 11' / 'Ytterligare höjning…'
    # överlappade — wrappa i rubrikraden i stället
    for ref in ("E35", "F35"):
        c = ws[ref]
        c.alignment = Alignment(horizontal="center", vertical="bottom",
                                wrap_text=True)
    if ws.row_dimensions[35].height is None or ws.row_dimensions[35].height < 42:
        ws.row_dimensions[35].height = 42


def round_oversikt_v2(wb) -> None:
    """Bygger om Översikt från noll: ingen flytande hero-bild (rotorsak till
    18-sidersbuggen), print_area över hela layouten, locale-säkra formler,
    nytt LÖNSAMHETSKRAV-statusblock. 1 sida liggande."""
    idx = wb.sheetnames.index("Översikt")
    wb.remove(wb["Översikt"])
    ws = wb.create_sheet("Översikt", idx)

    ws.sheet_view.showGridLines = False
    widths = {"B": 24, "C": 16, "D": 3, "E": 24, "F": 16, "G": 3,
              "H": 22, "I": 8, "J": 16, "K": 3, "L": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    heights = {1: 8, 2: 14, 3: 46, 4: 18, 5: 10, 6: 10, 7: 14, 8: 36, 9: 16,
               10: 8, 11: 10, 12: 20, 13: 18, 14: 18, 15: 18, 16: 18, 17: 18,
               18: 18, 19: 18, 20: 8, 21: 10, 22: 20, 23: 6, 24: 30, 25: 16,
               26: 18, 27: 8, 28: 16}
    for r, h in heights.items():
        ws.row_dimensions[r].height = h

    apply(ws["B2"], "eyebrow")
    ws["B2"] = "LEJONFASTIGHETER · INVESTERINGSANALYS"

    ws["B3"] = "Investeringskalkyl"
    ws["B3"].font = Font(name="Segoe UI Light", size=36, color=INK)
    ws["B3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("B3:J3")

    apply(ws["B4"], "h1_subtitle")
    ws["B4"] = '=Indata!B26&" · Kalkylstart "&Indata!C5'
    ws.merge_cells("B4:J4")

    for col in "BCDEFGHIJK":
        bottom_rule(ws[f"{col}5"])

    # KPI-band
    for ref, txt in (("B7", "BINDANDE KRAVHYRA"), ("E7", "PER M² OCH ÅR"),
                     ("H7", "FAKTISK IRR EK")):
        apply(ws[ref], "h2_section")
        ws[ref] = txt
        ws[ref].font = Font(name="Segoe UI Semibold", size=9, bold=True, color=MUTED)

    _kpi_value(ws, "B8", '#,##0')
    ws["B8"] = "=Resultat!D14"
    ws.merge_cells("B8:C8")
    _kpi_value(ws, "E8", '#,##0')
    ws["E8"] = "=Resultat!D16"
    ws.merge_cells("E8:F8")
    _kpi_value(ws, "H8", "0.00%", color=POSITIVE)
    ws["H8"] = "=Lönsamhetskontroll!C45"
    ws.merge_cells("H8:I8")

    apply(ws["B9"], "caption")
    ws["B9"] = "kr/år · högsta av de tre lönsamhetskraven"
    apply(ws["E9"], "caption")
    ws["E9"] = "kr/m²/år (totalarea)"
    ws["H9"] = "=Lönsamhetskontroll!C47*100"
    ws["H9"].font = Font(name=FAMILY, size=9, color=MUTED)
    ws["H9"].alignment = Alignment(horizontal="left", vertical="center")
    ws["H9"].number_format = '"Marginal +"0.00" pp";"Marginal −"0.00" pp"'
    ws["J9"] = "=Indata!C65"
    ws["J9"].font = Font(name=FAMILY, size=9, color=MUTED)
    ws["J9"].alignment = Alignment(horizontal="left", vertical="center")
    ws["J9"].number_format = '"Krav "0.0%'

    for col in "BCDEFGHIJK":
        bottom_rule(ws[f"{col}10"])

    # Tre paneler
    for ref, txt in (("B12", "FÖRUTSÄTTNINGAR"), ("E12", "DRIFTNETTO ÅR FÖR ÅR"),
                     ("H12", "LÖNSAMHETSKRAV")):
        apply(ws[ref], "h2_section")
        ws[ref] = txt

    facts = [
        ("Total investering", "=Indata!R31", FMT_KR),
        ("Investering per m²", "=IFERROR(Indata!R31/Indata!F31,0)", FMT_KR_KVM),
        ("Verksamhetsyta", "=Indata!F31", FMT_M2),
        ("Kalkylperiod", "=Indata!C16", FMT_AR),
        ("Belåningsgrad", "=Indata!C64", FMT_PCT1),
        ("Avkastningskrav EK", "=Indata!C65", FMT_PCT1),
        ("Kalkylränta driftnetto", "=Indata!C7", FMT_PCT1),
    ]
    for i, (label, formula, fmt) in enumerate(facts):
        r = 13 + i
        apply(ws[f"B{r}"], "label")
        ws[f"B{r}"] = label
        apply(ws[f"C{r}"], "value")
        ws[f"C{r}"] = formula
        ws[f"C{r}"].number_format = fmt

    apply(ws["E13"], "label")
    ws["E13"] = '="År 1 ("&Indata!C5&")"'
    apply(ws["F13"], "value")
    ws["F13"] = "=Kassaflöde!D23"
    ws["F13"].number_format = FMT_KR
    apply(ws["E14"], "label")
    ws["E14"] = '="År "&Indata!C16&" ("&(Indata!C5+Indata!C16-1)&")"'
    apply(ws["F14"], "value")
    ws["F14"] = "=INDEX(Kassaflöde!D23:AB23,MATCH(Indata!C16,Kassaflöde!D4:AB4,0))"
    ws["F14"].number_format = FMT_KR
    apply(ws["E15"], "label")
    ws["E15"] = "Tillväxt över perioden"
    apply(ws["F15"], "value")
    ws["F15"] = "=IFERROR(F14/F13-1,0)"
    ws["F15"].number_format = '+0.0%;−0.0%'
    apply(ws["E17"], "caption")
    ws["E17"] = ("Driftnetto = nettohyra minus drift och underhåll. "
                 "Hela projektionen finns på fliken Kassaflöde.")
    ws["E17"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells("E17:F19")

    krav = [
        ("1.  NPV ≥ 0", "=Lönsamhetskontroll!D9"),
        ("2.  IRR EK ≥ avkastningskrav", "=Lönsamhetskontroll!D45"),
        ("3.  Marknadsvärde ≥ bokfört värde", "=Lönsamhetskontroll!D30"),
    ]
    for i, (label, formula) in enumerate(krav):
        r = 13 + i
        apply(ws[f"H{r}"], "label")
        ws[f"H{r}"] = label
        ws[f"J{r}"] = formula
        ws[f"J{r}"].font = Font(name=FAMILY, size=10, bold=True, color=POSITIVE)
        ws[f"J{r}"].fill = PatternFill("solid", fgColor=POSITIVE_TINT)
        ws[f"J{r}"].alignment = Alignment(horizontal="center", vertical="center")
    # Villkorsstyrd färg: rött vid "✗" så status aldrig ljuger
    red_dxf = DifferentialStyle(font=Font(color="C0392B", bold=True),
                                fill=PatternFill("solid", bgColor="FDEDEC"))
    rule = Rule(type="containsText", operator="containsText", text="✗", dxf=red_dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("✗",J13)))']
    ws.conditional_formatting.add("J13:J15", rule)

    apply(ws["H17"], "caption")
    ws["H17"] = ("Utvärderas vid bindande kravhyra och mål-utfall. "
                 "Detaljer på fliken Lönsamhetskontroll.")
    ws["H17"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells("H17:J19")

    for col in "BCDEFGHIJK":
        bottom_rule(ws[f"{col}20"])

    # Hyresspann
    apply(ws["B22"], "h2_section")
    ws["B22"] = "HYRESSPANN VID INVESTERINGSUTFALL"
    ws["H22"] = '="Investeringsintervall ±"&Indata!C68*100&" %"'
    apply(ws["H22"], "caption")

    spann = [("B24", "=Resultat!C14", ACCENT_TINT, 14, '#,##0'),
             ("E24", "=Resultat!D14", ACCENT, 16, '#,##0" kr/år"'),
             ("H24", "=Resultat!E14", ACCENT_TINT, 14, '#,##0')]
    for ref, formula, fill, size, fmt in spann:
        c = ws[ref]
        c.value = formula
        c.font = Font(name="Segoe UI Semibold", size=size, bold=True, color=INK)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = fmt
    ws.merge_cells("B24:C24")
    ws.merge_cells("E24:F24")
    ws.merge_cells("H24:J24")

    apply(ws["B25"], "caption")
    ws["B25"] = '="Lägsta utfall (−"&Indata!C68*100&" %)"'
    ws["B25"].alignment = Alignment(horizontal="center")
    apply(ws["E25"], "caption")
    ws["E25"] = "Mål-utfall (bindande kravhyra)"
    ws["E25"].alignment = Alignment(horizontal="center")
    apply(ws["H25"], "caption")
    ws["H25"] = '="Högsta utfall (+"&Indata!C68*100&" %)"'
    ws["H25"].alignment = Alignment(horizontal="center")

    apply(ws["B26"], "caption")
    ws["B26"] = ("Avtalet bör ange hyresspannet. Slutlig hyra fastställs när "
                 "investeringsutfallet är känt vid projektavslut.")
    ws.merge_cells("B26:J26")

    apply(ws["B28"], "caption")
    ws["B28"] = "Lejonfastigheter AB · Investeringskalkyl · Beslutsunderlag"

    ws.freeze_panes = "B1"
    ws.print_area = "B1:L29"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


# ── V2-05: Beräkningslogik — tekniska blocket i tkr ─────────────────────────

def round_berakningslogik_tkr(wb) -> None:
    """Löser ########: tekniska rådatablockets penningceller (rad 48+) visas i
    tkr ('#,##0,'). Beslut per rad utifrån spec-cachade magnituder: penningrad
    om max|värde| ≥ 100 000 (areor/år/kr-per-kvm-rader lämnas orörda).
    Endast visningsformat ändras — inga värden."""
    ws = wb["Beräkningslogik"]
    sheet_spec = next(s for s in SPEC["sheets"] if s["name"] == "Beräkningslogik")

    by_row: dict[int, list[tuple[str, float | None, str]]] = {}
    for c in sheet_spec["cells"]:
        fmt = (c.get("style") or {}).get("numfmt", "")
        if not fmt.startswith("#,##0") or "%" in fmt or "#,##0." in fmt:
            continue
        m = re.match(r"([A-Z]+)(\d+)", c["ref"])
        col, row = m.group(1), int(m.group(2))
        if row < 48 or not ("C" <= col <= "Z" or col in ("AA", "AB", "AC", "AD")):
            continue
        cached = c.get("cached")
        val = cached if isinstance(cached, (int, float)) else None
        by_row.setdefault(row, []).append((c["ref"], val, fmt))

    TKR = '#,##0,;-#,##0,;"–"'
    converted = 0
    for row, cells in by_row.items():
        vals = [abs(v) for _, v, _ in cells if v is not None]
        row_max = max(vals) if vals else 0
        if row_max >= 100_000 or row_max == 0:
            for ref, _, _ in cells:
                ws[ref].number_format = TKR
                converted += 1

    ws["B46"] = "TEKNISKA BERÄKNINGSBLOCK — BELOPP I TKR"
    for ref in ("B59", "B121", "B164", "B179", "B197"):
        cur = ws[ref].value
        if cur and "tkr" not in str(cur):
            ws[ref] = f"{cur} · tkr"
    print(f"    tkr-konvertering: {converted} celler")


def round_berakningslogik_polish(wb) -> None:
    """F-4: stale radreferenser i pedagogiska annoteringarna (pekar på
    'Kassaflöde rad 10-70' — blocken ligger på Beräkningslogik rad 59-204
    sedan blocken flyttades). Plus: huggen headertext (merge B46:E46 för smal),
    klippt radhöjd B35, rena sidbrytningar vid blockgränser, trimmad print_area."""
    from openpyxl.worksheet.pagebreak import Break
    ws = wb["Beräkningslogik"]

    ws["D7"] = "(rad 119 — NPV vid 0 hyra)"
    ws["D8"] = "(rad 139 — NPV vid 1 Mkr)"
    ws["D12"] = "(rad 140 — b_NPV)"
    ws["D25"] = "Block nedan på denna flik"
    ws["D26"] = "Block A (rad 59–119) + Block B (rad 121–140)"
    ws["D27"] = "Block D (rad 164–177) + Block E (rad 179–195)"
    ws["D28"] = "Block F (rad 197–204)"

    # Header-merge för smal (text höggs vid kolumn E) — bredda till K som B47
    from openpyxl.utils import range_boundaries  # noqa: F401  (dokumenterande)
    if "B46:E46" in [str(m) for m in ws.merged_cells.ranges]:
        ws.unmerge_cells("B46:E46")
    ws.merge_cells("B46:K46")
    ws["B47"] = ("Rådata bakom kravhyrorna. Block A/B ger NPV-kravet, "
                 "D/E ger IRR-kravet, F ger MV-kravet. Belopp i tkr.")
    ws.row_dimensions[35].height = 60  # wrappad text klipptes vid 43.5

    # Rena sidbrytningar: pedagogik | tekniska block | restvärdesresonemang
    ws.row_breaks.append(Break(id=45))
    ws.row_breaks.append(Break(id=207))
    ws.print_area = "B1:AB230"  # AC:AD tomma — gav smalare kolumner per sida


# ── F-5/F-6: Kassaflöde — infasningsbugg + trasig utskrift ──────────────────

def round_kassaflode_print(wb) -> None:
    """F-5: INFASNING-formlerna multiplicerar med $E$7 (= 'Area Bef år 2' = 0)
    → hela tabellen visar 0 trots att Skola tillträder år 1. v2 binder raden
    till objektets andel av faktisk bruttohyra det året (rad 16) med
    tillträdes-flagga, och 'Andel av fullt flöde' divideras per år (rad 16) i
    stället för MAX över år 1-10 — då blir raden 100 % när alla tillträtt.

    F-6: print: årskolumner E..AB hade defaultbredd 8,43 → ####### för alla
    8-siffriga belopp (år 2+), och print_area B1:AB156 täckte 120 tomma rader
    → 6 av 10 sidor blanka. Innehållet slutar på rad 36."""
    from openpyxl.utils import get_column_letter
    ws = wb["Kassaflöde"]

    # Tabellen täcker uppstartsåren 1-10 (årshuvuden D..M, som i iter9)
    for i, row in enumerate(range(30, 35)):
        krow = 26 + i  # Indata-rad för objekt 1-5
        for col in range(4, 14):  # D..M
            yc = get_column_letter(col)
            ws.cell(row=row, column=col).value = (
                f'=IF(OR(Indata!$K${krow}="",{yc}$29<Indata!$K${krow}),0,'
                f"(Indata!$R${krow}/Indata!$R$31)*{yc}$16)"
            )
        for col in range(14, 29):  # N..AB: rensa ströceller utan årshuvud
            ws.cell(row=row, column=col).value = None
    for col in range(4, 14):
        yc = get_column_letter(col)
        ws.cell(row=36, column=col).value = f"=IFERROR({yc}35/{yc}16,0)"
    for col in range(14, 29):
        ws.cell(row=35, column=col).value = None
        ws.cell(row=36, column=col).value = None

    for col in range(5, 29):  # E..AB som D (13)
        ws.column_dimensions[get_column_letter(col)].width = 13

    ws.print_area = "B1:W36"  # år 1-20; år 21-25 (X:AB) syns på skärm, skrivs ej ut

    # Finansiering: samma defaultbredd-sjuka (####### i år 2+)
    fin = wb["Finansiering"]
    for col in range(5, 29):
        fin.column_dimensions[get_column_letter(col)].width = 13
    fin.print_area = "B1:W18"


# ── Lönsamhetskontroll: ####### i känslighetstabellerna ─────────────────────

def round_lonsamhetskontroll_print(wb) -> None:
    """Kolumn D/E saknade bredd → ####### i 'Diff mot BV' och i
    restvärdestabellens Bedömt/Pessimistisk-kolumner. Formel-captions (E27/E29)
    höggs vid printkanten. fitToWidth=1 så den breddade tabellen skalas."""
    ws = wb["Lönsamhetskontroll"]
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws["E27"] = "Formel: byggnadsvärde × (1 − N × avskr.)"
    ws["E29"] = "Formel: driftnetto år 21 / direktavk."
    ws.merge_cells("B51:E51")  # caption överflödade printkanten — wrap i merge
    ws["B51"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[51].height = 30
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


# ── V2-06: Grafer ────────────────────────────────────────────────────────────

def round_grafer(wb) -> None:
    """Ny flik med tre diagram à la LM 371:s Grafer-flik: driftnetto per år,
    ackumulerat driftnetto vs investering (återbetalning), hyresspann.
    Diagramdata-underlag synligt och refererat — inga hårdkodade tal."""
    idx = wb.sheetnames.index("Resultat") + 1
    ws = wb.create_sheet("Grafer", idx)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 28

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 14
    apply(ws["B2"], "eyebrow")
    ws["B2"] = "LEJONFASTIGHETER · GRAFER"
    ws.row_dimensions[3].height = 30
    ws["B3"] = "Grafer"
    ws["B3"].font = Font(name="Segoe UI Light", size=22, color=INK)
    apply(ws["B4"], "h1_subtitle")
    ws["B4"] = "Kalkylens utveckling i bilder — vid bindande kravhyra och mål-utfall"

    # Diagramdata (synligt underlag, under diagrammen)
    DATA0 = 44  # första datarad
    apply(ws[f"B{DATA0 - 3}"], "h2_section")
    ws[f"B{DATA0 - 3}"] = "DIAGRAMDATA — UNDERLAG"
    apply(ws[f"B{DATA0 - 2}"], "caption")
    ws[f"B{DATA0 - 2}"] = "Refererar Kassaflöde, Indata och Resultat. Diagrammen ovan läser dessa rader."

    MKR = '#,##0.0,,'
    apply(ws["B44"], "label")
    ws["B44"] = "Kalenderår"
    apply(ws["B45"], "label")
    ws["B45"] = "Driftnetto (Mkr)"
    apply(ws["B46"], "label")
    ws["B46"] = "Ackumulerat driftnetto (Mkr)"
    apply(ws["B47"], "label")
    ws["B47"] = "Total investering (Mkr)"
    for i in range(20):  # år 1-20 i kolumn C..V
        ws.cell(row=44, column=3 + i).value = f"=Kassaflöde!{chr(ord('D') + i)}3"
        ws.cell(row=44, column=3 + i).number_format = "0"
        ws.cell(row=45, column=3 + i).value = f"=Kassaflöde!{chr(ord('D') + i)}23"
        ws.cell(row=45, column=3 + i).number_format = MKR
        prev = "" if i == 0 else f"+{ws.cell(row=46, column=2 + i).coordinate}"
        ws.cell(row=46, column=3 + i).value = (
            f"={ws.cell(row=45, column=3 + i).coordinate}{prev}"
        )
        ws.cell(row=46, column=3 + i).number_format = MKR
        ws.cell(row=47, column=3 + i).value = "=Indata!$R$31"
        ws.cell(row=47, column=3 + i).number_format = MKR
        for r in (44, 45, 46, 47):
            ws.cell(row=r, column=3 + i).font = Font(name=FAMILY, size=9, color=MUTED)

    apply(ws["B49"], "label")
    ws["B49"] = "Hyresspann (kr/år)"
    for j, (label_f, val_f) in enumerate([
            ('="Lägsta (−"&Indata!C68*100&" %)"', "=Resultat!C14"),
            ("Mål (bindande)", "=Resultat!D14"),
            ('="Högsta (+"&Indata!C68*100&" %)"', "=Resultat!E14")]):
        lc = ws.cell(row=50, column=3 + j)
        lc.value = label_f
        lc.font = Font(name=FAMILY, size=9, color=MUTED)
        vc = ws.cell(row=51, column=3 + j)
        vc.value = val_f
        vc.number_format = FMT_INT
        vc.font = Font(name=FAMILY, size=9, color=MUTED)

    kf = wb["Kassaflöde"]

    def _axes(ch, y_fmt='#,##0,,'):
        # openpyxl-fälla: utan explicit delete=False döljer Excel axeletiketterna
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        ch.y_axis.number_format = y_fmt   # miljoner
        ch.y_axis.majorTickMark = "out"
        ch.x_axis.majorTickMark = "out"

    # 1) Driftnetto per år
    ch1 = BarChart()
    ch1.type = "col"
    ch1.title = "Driftnetto per år (Mkr)"
    ch1.height = 8.5
    ch1.width = 22
    ch1.legend = None
    ch1.gapWidth = 60
    data = Reference(kf, min_col=4, max_col=23, min_row=23, max_row=23)
    cats = Reference(kf, min_col=4, max_col=23, min_row=3, max_row=3)
    ch1.add_data(data, from_rows=True, titles_from_data=False)
    ch1.set_categories(cats)
    ch1.series[0].graphicalProperties.solidFill = INK
    _axes(ch1)
    ws.add_chart(ch1, "B6")

    # 2) Ackumulerat driftnetto vs investering
    ch2 = LineChart()
    ch2.title = "Ackumulerat driftnetto vs investering (Mkr)"
    ch2.height = 8.5
    ch2.width = 13.5
    s_ack = Series(Reference(ws, min_col=3, max_col=22, min_row=46, max_row=46),
                   title="Ackumulerat driftnetto")
    s_inv = Series(Reference(ws, min_col=3, max_col=22, min_row=47, max_row=47),
                   title="Total investering")
    ch2.append(s_ack)
    ch2.append(s_inv)
    ch2.set_categories(Reference(ws, min_col=3, max_col=22, min_row=44, max_row=44))
    s_ack.graphicalProperties.line = LineProperties(solidFill=ACCENT, w=28000)
    s_ack.marker = Marker(symbol="none")
    s_inv.graphicalProperties.line = LineProperties(solidFill=INK, w=18000,
                                                    prstDash="dash")
    s_inv.marker = Marker(symbol="none")
    ch2.legend.position = "b"
    ch2.legend.overlay = False  # annars renderas legenden över x-axeln
    ch2.x_axis.tickLblSkip = 4
    ch2.x_axis.tickMarkSkip = 2
    _axes(ch2)
    ws.add_chart(ch2, "B24")

    # 3) Hyresspann
    ch3 = BarChart()
    ch3.type = "col"
    ch3.title = "Kravhyra vid investeringsutfall (Mkr/år)"
    ch3.height = 8.5
    ch3.width = 8.5
    ch3.legend = None
    ch3.gapWidth = 40
    ch3.add_data(Reference(ws, min_col=3, max_col=5, min_row=51, max_row=51),
                 from_rows=True, titles_from_data=False)
    ch3.set_categories(Reference(ws, min_col=3, max_col=5, min_row=50, max_row=50))
    ch3.series[0].graphicalProperties.solidFill = ACCENT
    _axes(ch3, y_fmt='#,##0.0,,')
    ws.add_chart(ch3, "J24")

    ws.freeze_panes = "B1"
    ws.print_area = "B1:N53"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


# ── Dokumentation: klippt text + stale referenser ───────────────────────────

DOC_REF_FIXES = [
    ("Rad på Kassaflöde: C70 (NPV vid 0 hyra), C91 (känslighet)",
     "Rad på Beräkningslogik: C119 (NPV vid 0 hyra), C140 (b_NPV)"),
    ("Rad på Kassaflöde: C128 (NPV EK vid 0 hyra), C144 (b_IRR)",
     "Rad på Beräkningslogik: C177 (NPV EK vid 0 hyra), C193 (b_IRR)"),
    ("Rad på Kassaflöde: C150-C156",
     "Rad på Beräkningslogik: Block F (rad 197–204)"),
    ("(Kassaflöde rad 33)", "(Kassaflöde rad 22)"),
    ("(Kassaflöde C123)", "(Beräkningslogik rad 172)"),
]


def round_dokumentation_polish(wb) -> None:
    """Dokumentation: (1) ~20 stycken klipptes vertikalt — radhöjderna var
    satta för lågt för den wrappade texten (render_local-renderaren visade
    annan radbrytning än Excel). Autofit: höjd från textlängd vid B-bredd 100.
    (2) Samma stale 'Rad på Kassaflöde'-referenser som F-4 — blocken ligger
    på Beräkningslogik."""
    import math
    ws = wb["Dokumentation"]
    CPL = 92          # tecken per rad vid bredd 100, Segoe UI 10-11pt (konservativt)
    PT_PER_LINE = 14.5

    for row in ws.iter_rows(min_col=2, max_col=2):
        cell = row[0]
        if not isinstance(cell.value, str):
            continue
        for old, new in DOC_REF_FIXES:
            if old in cell.value:
                cell.value = cell.value.replace(old, new)
        wrapped = bool(cell.alignment and cell.alignment.wrapText)
        if not wrapped and len(cell.value) > 110:
            # långa o-wrappade rader (t.ex. sektion 14:s markerade) klipps
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal if cell.alignment else "left",
                vertical="top", wrap_text=True)
            wrapped = True
        if wrapped:
            lines = max(1, math.ceil(len(cell.value) / CPL))
            needed = lines * PT_PER_LINE + 4
            rd = ws.row_dimensions[cell.row]
            if rd.height is None or rd.height < needed:
                rd.height = needed

    # Merges B:E sträcker sig utanför print_area B1:B169 → svansen klipptes.
    ws.print_area = "B1:E169"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


# ── V2-07: sidnav med 10 poster ─────────────────────────────────────────────

def round_sidenav(wb) -> None:
    import tools.sidenav as sidenav
    sidenav.NAV_ITEMS = [
        ("Försättsblad", "F"),
        ("Översikt", "Ö"),
        ("Indata", "I"),
        ("Kassaflöde", "K"),
        ("Finansiering", "$"),
        ("Resultat", "R"),
        ("Grafer", "G"),
        ("Lönsamhetskontroll", "L"),
        ("Beräkningslogik", "Σ"),
        ("Dokumentation", "D"),
    ]
    paths = sidenav.generate_assets()
    for ws in wb.worksheets:
        for r in range(1, 13):  # rensa replayade nav-länkar/texter (9-postersnav)
            cell = ws.cell(r, 1)
            cell.hyperlink = None
            if cell.value is not None:
                cell.value = None
        sidenav.apply_to_sheet(ws, ws.title, paths)


# ── FINAL m1: printkomprimering ─────────────────────────────────────────────

def round_print_compact(wb) -> None:
    """FINAL m1: Beräkningslogik 16 sidor → pedagogik-only i print.
    Tekniska rådatablocket (rad 48–206) outline-kollapsas som default — all
    data kvar i filen, expanderas via +-symbolen (dolda rader skrivs inte ut).
    Print smalnas till B1:K230 med fitToWidth=1; restvärdesresonemangets
    B:R-merges om-mergas till B:K så inget klipps vid nya printkanten.
    Dokumentation: tomma spacer-rader 15 → 8 pt (prosan orörd)."""
    import math
    from openpyxl.worksheet.pagebreak import RowBreak

    ws = wb["Beräkningslogik"]
    for r in range(48, 207):
        rd = ws.row_dimensions[r]
        rd.outlineLevel = 1
        rd.hidden = True
    # summaryBelow=0 (XML-patch i build_v2) → +-symbolen ligger på raden ovanför
    ws.row_dimensions[47].collapsed = True

    ws["B47"] = ("Rådata bakom kravhyrorna. Block A/B ger NPV-kravet, D/E ger "
                 "IRR-kravet, F ger MV-kravet. Belopp i tkr. Blocket är kollapsat — "
                 "expandera med +-symbolen i vänstermarginalen. Kollapsade rader "
                 "skrivs inte ut.")
    ws["B47"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[47].height = 34

    CPL = 145  # tecken per rad vid B:K-bredd (38+22+8×13 ≈ 164 enheter)
    for m in [str(x) for x in list(ws.merged_cells.ranges)]:
        match = re.fullmatch(r"B(\d+):R(\d+)", m)
        if match and match.group(1) == match.group(2) and int(match.group(1)) >= 208:
            r = int(match.group(1))
            ws.unmerge_cells(m)
            ws.merge_cells(f"B{r}:K{r}")
            text = ws.cell(row=r, column=2).value
            if isinstance(text, str):
                lines = max(1, math.ceil(len(text) / CPL))
                ws.row_dimensions[r].height = lines * 14.5 + 6

    ws.row_breaks = RowBreak()  # sektionbrytningarna vid 45/207 behövs inte längre
    ws.print_area = "B1:K230"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Dokumentation: prosan wrappade i B (100 enheter) men printbredden var
    # B:E (139) — 28 % av sidbredden stod tom. Bredda B till hela printbredden,
    # släpp B:E-merges och autofit om med tightare leading. Spacer-rader → 6 pt.
    dok = wb["Dokumentation"]
    for m in [str(x) for x in list(dok.merged_cells.ranges)]:
        if re.fullmatch(r"B\d+:E\d+", m):
            dok.unmerge_cells(m)
    W_DOK = 152
    dok.column_dimensions["B"].width = W_DOK
    for r in range(4, 170):
        cell = dok.cell(row=r, column=2)
        v = cell.value
        rd = dok.row_dimensions[r]
        if not isinstance(v, str) or not v.strip():
            rd.height = 2.5
            continue  # spacer
        wrapped = bool(cell.alignment and cell.alignment.wrapText)
        size = cell.font.size or 10
        if cell.font.name == "Segoe UI" and size == 10 and not cell.font.bold:
            cell.font = Font(name="Segoe UI", size=9, color=cell.font.color)
            size = 9
        if wrapped and size <= 11:
            # ≈0,88 tecken/breddenhet×10/size (konservativt mot Excelwrap);
            # 11pt Semibold behöver ~15,5pt/rad — annars klipps highlight-boxarna
            cpl = int(W_DOK * 8.8 / size)
            lines = max(1, math.ceil(len(v) / cpl))
            leading = 1.45 if size > 10 else 1.37
            rd.height = lines * (leading * size) + 2
        else:
            cur = rd.height or 15
            if cur > 1.5 * size:
                rd.height = 1.5 * size  # trimma luftiga enrads-rader (rubriker + bullets)
    dok.print_area = "B1:B169"
    dok.page_margins.top = 0.4
    dok.page_margins.bottom = 0.4


def round_print_polish(wb) -> None:
    """FINAL m1b: fynd från trogen PDF-granskning (sida-för-sida, 22-sidorspasset).
    - H1-nedstaplar (g/y) klipptes: Översikt 36pt / Försättsblad 42pt i 30pt-rad
    - Resultat: ####### i kolumn E (Högsta-utfall) + rad kluven av sidbrytning → 1 sida
    - Kassaflöde/Finansiering: årsremsa 2034–2045 utan radetiketter + föräldralösa
      svanssidor → fitToHeight=1 + upprepade etikettkolumner B:C
    - Grafer: print_area slutade vid N → diagramdata-kolumnerna O..V klipptes
    - Lönsamhetskontroll: B51-mergen (t.o.m. H) låg utanför print_area (G) → klippt caption
    - Försättsblad: sidenav-bandet slutar abrupt mitt på sid 2 i print → print utan kolumn A
      (nav är skärmnavigation; övriga flikar printar redan från B)"""
    from openpyxl.utils import get_column_letter

    wb["Översikt"].row_dimensions[3].height = 48
    fs = wb["Försättsblad"]
    fs.row_dimensions[4].height = 56
    fs.print_area = "B1:I92"  # högerkolumnens värde-merges (H:I) ska med
    fs.page_setup.fitToWidth = 1
    fs.page_setup.fitToHeight = 0
    fs.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    from openpyxl.worksheet.pagebreak import Break as _Break, RowBreak as _RowBreak
    fs.row_breaks = _RowBreak()
    fs.row_breaks.append(_Break(id=45))  # sid 1 = memo/cover, sid 2 = antaganden→underskrifter

    r = wb["Resultat"]
    r.column_dimensions["E"].width = 18
    r.row_dimensions[19].height = 48  # Tolkning (C19:F19, 134 tecken ≈ 3 rader) klipptes vid 27,75
    r.page_setup.fitToWidth = 1
    r.page_setup.fitToHeight = 1
    r.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    for name, caprows in (("Kassaflöde", (11, 27)), ("Finansiering", (8,))):
        ws = wb[name]
        ws.page_setup.fitToHeight = 1
        ws.print_title_cols = "B:C"
        # Captions är spec-mergade B:M (in i årskolumnerna) → klipps vid remsgränsen
        # när bara B:C upprepas på sid 2. Krymp mergen till B:C och wrappa.
        for caprow in caprows:
            for m in [str(x) for x in list(ws.merged_cells.ranges)]:
                if m.startswith(f"B{caprow}:"):
                    ws.unmerge_cells(m)
            ws.merge_cells(start_row=caprow, start_column=2, end_row=caprow, end_column=3)
            cap = ws.cell(row=caprow, column=2)
            cap.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            lines = max(1, -(-len(str(cap.value)) // 50))
            ws.row_dimensions[caprow].height = lines * 14 + 4

    g = wb["Grafer"]
    for col in ("C", "D", "E"):
        g.column_dimensions[col].width = 15
    g.print_area = "B1:V53"
    g.page_setup.orientation = "landscape"
    g.page_setup.fitToWidth = 1
    g.page_setup.fitToHeight = 1
    g.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    lk = wb["Lönsamhetskontroll"]
    lk.print_area = "B1:H87"
    lk.row_dimensions[51].height = 34


# ── FINAL m2: Faktisk IRR EK per yield-scenario ─────────────────────────────

def round_irr_yield_scenarios(wb) -> None:
    """FINAL m2: Faktisk IRR EK per scenario i restvärdestabellen (rad 73+).
    EK-cashflow rekonstrueras per scenario i ett dolt outline-block (rad 89–91),
    samma mönster som MV-exit-tabellens rad 64–70: kassaflödet är linjärt i
    hyran (Beräkningslogik block D/E: rad 173 vid 0 kr, rad 188 vid 1 Mkr),
    plus exit-delta i år 20-kolumnen (Y) för scenarioyielden. Bedömt-kolumnen
    (D78 = Indata!D81, D82 = MAX(D79:D81) = Resultat!D14) är per konstruktion
    identisk med rad 43 → IRR exakt = C45.

    Fixar också en timing-bugg i MV-exit-tabellen (D-20): ΔMV adderades i
    kolumn AD (år 25) men exiten ligger i Y (år 20) — scenariots IRR fick
    kassaflödesjusteringen 5 år för sent. Bas-scenariot opåverkat (Δ=0)."""
    from copy import copy
    from openpyxl.utils import get_column_letter
    ws = wb["Lönsamhetskontroll"]

    # (a) MV-exit: flytta ΔMV från AD (år 25) till Y (år 20)
    for i, r in enumerate(range(64, 71)):
        ws.cell(row=r, column=30).value = "=AD43"
        ws.cell(row=r, column=25).value = f"=Y43+$C{54 + i}-$C$58"

    # (b) Synlig IRR-rad i yield-tabellen (rad 86, inom print_area B1:H87)
    b = ws["B86"]
    b.value = "Faktisk IRR EK"
    b.font = copy(ws["B84"].font)
    for col, cfrow in (("C", 89), ("D", 90), ("E", 91)):
        c = ws[f"{col}86"]
        c.value = f'=IFERROR(IRR(F{cfrow}:AD{cfrow}),"")'
        c.font = copy(ws[f"{col}84"].font)
        c.fill = copy(ws[f"{col}84"].fill)
        c.alignment = copy(ws[f"{col}84"].alignment)
        c.number_format = "0.00%"
    ws.row_dimensions[86].height = 18
    ws["B87"] = ("IRR EK räknas vid respektive scenarios bindande kravhyra och yield "
                 "(EK-cashflöde per scenario i expanderbart block nedan). Bedömt = "
                 "Faktisk IRR EK (C45). Om krav 2 (IRR) är bindande är IRR EK exakt "
                 "= avkastningskravet; annars strikt högre än kravet.")

    # (c) Dolt EK-cashflow-block, rad 89–91 (utanför print_area; dolda rader skrivs ej ut)
    cap = ws["B88"]
    cap.value = "EK-cashflöde per yield-scenario — klicka [+] för att expandera"
    cap.font = copy(ws["B63"].font)
    for r, col in ((89, "C"), (90, "D"), (91, "E")):
        for cc in range(6, 31):  # F..AD ↔ Beräkningslogik D..AB (år 1..25)
            bl = get_column_letter(cc - 2)
            f = (f"=Beräkningslogik!{bl}173"
                 f"+(Beräkningslogik!{bl}188-Beräkningslogik!{bl}173)*{col}$82/1000000")
            if cc == 25:  # Y = år 20: exit-delta (MV vid scenarioyield − MV vid basyield)
                f += (f"+(Beräkningslogik!$C$199"
                      f"+(Beräkningslogik!$C$200-Beräkningslogik!$C$199)*{col}$82/1000000)"
                      f"*(1/{col}$78-1/Indata!$D$81)")
            cell = ws.cell(row=r, column=cc)
            cell.value = f
            cell.number_format = "#,##0"
        rd = ws.row_dimensions[r]
        rd.outlineLevel = 1
        rd.hidden = True
    ws.row_dimensions[88].collapsed = True  # summaryBelow=0 → + på raden ovanför


# ── FINAL m3: design-excellens ──────────────────────────────────────────────

def round_design_final(wb) -> None:
    """FINAL m3: designfynd från trogen PDF-granskning. Indata sektion 5
    (re-investering): tomma mallrader renderade som ett kantlöst blått block
    med sex lösryckta nollor i Total-kolumnen — hårlinjer mellan raderna och
    tomt i stället för 0 tills något matas in."""
    ws = wb["Indata"]
    hair = Side(style="hair", color=RULE)
    for r in range(45, 51):
        for c in range(2, 7):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=hair)
        ws.cell(row=r, column=6).number_format = '#,##0;-#,##0;""'


# ── Pipeline ────────────────────────────────────────────────────────────────

def apply_all(wb) -> None:
    """Körs på workbook-objektet innan save."""
    fix_iter9_name_errors(wb)
    round_locale_fix(wb)
    round_forsattsblad_polish(wb)
    round_indata_polish(wb)
    round_oversikt_v2(wb)
    round_berakningslogik_tkr(wb)
    round_berakningslogik_polish(wb)
    round_kassaflode_print(wb)
    round_lonsamhetskontroll_print(wb)
    round_dokumentation_polish(wb)
    round_grafer(wb)
    round_irr_yield_scenarios(wb)
    round_design_final(wb)
    round_print_compact(wb)
    round_sidenav(wb)
    # OBS: polish sist — sidenav sätter radhöjd 30 på rad 2–11 på ALLA flikar,
    # vilket klipper H1-nedstaplar (Översikt 36pt, Försättsblad 42pt)
    round_print_polish(wb)


def post_save(out_path: Path) -> None:
    """Körs på den sparade filen (XML-nivå) efter save + summaryBelow-patch."""
    pass
