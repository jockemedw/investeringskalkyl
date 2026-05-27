"""Comp-workbook: två designvarianter av Översikt + Resultat parallellt.

Variant A — "Annual report"  : editorial, monokromt, hairline-rules, ingen bg
Variant B — "Dashboard hero"  : KPI-shapes, accentfärg, mer levande

Bygger build/comps.xlsx med 4 flikar för side-by-side-jämförelse.
"""
from __future__ import annotations
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palett ─────────────────────────────────────────────────────────────────
INK      = "10313E"
MUTED    = "6C757D"
PAPER    = "FFFFFF"
SURFACE  = "F4F6F8"
ACCENT   = "FAB600"
POSITIVE = "00937C"
HAIRLINE = "DEE2E6"

FAMILY = "Segoe UI"

# Demo-data (samma som iter9 baseline)
DATA = {
    "project": "Skola — Nybyggnad",
    "fastighet": "Linköping kommun, kv. Tornet 5",
    "kalkylstart": 2026,
    "period": 20,
    "kalkyl_drift": 0.040,
    "irr_krav": 0.063,
    "dir_avk": 0.050,
    "belaning": 0.370,
    "inflation": 0.020,
    "area": 5000,
    "investering": 200_000_000,
    "kr_kvm": 40_000,
    "kravhyra": 11_631_222,
    "kravhyra_per_kvm": 2_326,
    "lagsta": 10_726_244,
    "hogsta": 12_536_199,
    "irr_faktisk": 0.0765,
    "marginal": 0.0135,
    "mv_ar20": 235_869_728,
    "bv_ar20": 80_000_000,
}


# ── Hjälpare ───────────────────────────────────────────────────────────────

def _font(size=10, bold=False, color=INK, italic=False, light=False):
    weight = False
    name = FAMILY
    if light:
        name = "Segoe UI Light"
    if bold:
        name = "Segoe UI Semibold"
        weight = True
    return Font(name=name, size=size, bold=weight, color=color, italic=italic)

def _fill(c): return PatternFill("solid", fgColor=c)
def _align(h="left", v="center", indent=0, wrap=False):
    return Alignment(horizontal=h, vertical=v, indent=indent, wrap_text=wrap)
def _rule_bottom(weight="thin", color=HAIRLINE):
    return Border(bottom=Side(style=weight, color=color))
def _rule_top(weight="thin", color=HAIRLINE):
    return Border(top=Side(style=weight, color=color))
def _rule_top_bottom(weight="thin", color=HAIRLINE):
    return Border(top=Side(style=weight, color=color),
                  bottom=Side(style=weight, color=color))


# ── VARIANT A: ÅRSREDOVISNING ──────────────────────────────────────────────
# Monokromt, typografisk hierarki, hairline rules, generös whitespace.

def build_oversikt_a(ws):
    """Översikt i editorial finansiell rapport-stil."""
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    # Kolumnbredder: A=margin, B=label, C=value, D=spacer, E=label, F=value, G=margin
    widths = {"A": 3, "B": 28, "C": 16, "D": 4, "E": 28, "F": 16, "G": 3}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Övre marginal: 3 tomma rader
    for r in range(1, 4):
        ws.row_dimensions[r].height = 14

    # Banner-rad: två-nivå titel
    ws["B4"] = "Investeringskalkyl"
    ws["B4"].font = Font(name="Segoe UI Light", size=28, color=INK)
    ws["B4"].alignment = _align(h="left", v="bottom")
    ws.row_dimensions[4].height = 44

    ws["B5"] = "LEJONFASTIGHETER AB"
    ws["B5"].font = _font(size=9, bold=True, color=MUTED)
    ws["B5"].alignment = _align(h="left", v="center")
    ws["B5"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[5].height = 18

    # Tunn rule under titel — spänner hela bredden
    ws.row_dimensions[6].height = 4
    for col in "BCDEF":
        ws[f"{col}6"].border = _rule_bottom("thin", HAIRLINE)

    ws.row_dimensions[7].height = 24  # luft

    # Projekt-info — minimalistisk rad
    ws["B8"] = "Projekt"
    ws["B8"].font = _font(size=9, color=MUTED)
    ws["C8"] = DATA["project"]
    ws["C8"].font = _font(size=11, color=INK)
    ws["C8"].alignment = _align(h="left")

    ws["E8"] = "Kalkylstart"
    ws["E8"].font = _font(size=9, color=MUTED)
    ws["F8"] = f"{DATA['kalkylstart']} · {DATA['period']} år"
    ws["F8"].font = _font(size=11, color=INK)
    ws["F8"].alignment = _align(h="left")

    ws["B9"] = "Fastighet"
    ws["B9"].font = _font(size=9, color=MUTED)
    ws["C9"] = DATA["fastighet"]
    ws["C9"].font = _font(size=11, color=INK)
    ws["C9"].alignment = _align(h="left")

    ws["E9"] = "Total investering"
    ws["E9"].font = _font(size=9, color=MUTED)
    ws["F9"] = DATA["investering"]
    ws["F9"].font = _font(size=11, color=INK)
    ws["F9"].number_format = '#,##0 "kr"'
    ws["F9"].alignment = _align(h="left")

    for r in [8, 9]:
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[10].height = 32  # luft

    # ── KRAVHYRA — central nyckelinformation ─────────────────────────────
    ws["B11"] = "BINDANDE KRAVHYRA"
    ws["B11"].font = _font(size=9, bold=True, color=MUTED)
    # spärrad textspacing via tomma cellrad ovanför
    ws.row_dimensions[11].height = 18

    # Stor siffra — Light vikt, mycket stor. Merge B:C för plats.
    ws.merge_cells("B12:C12")
    ws["B12"] = DATA["kravhyra"]
    ws["B12"].font = Font(name="Segoe UI Light", size=36, color=INK)
    ws["B12"].number_format = '#,##0'
    ws["B12"].alignment = _align(h="left", v="center")
    ws.row_dimensions[12].height = 52

    ws["B13"] = "kr/år"
    ws["B13"].font = _font(size=10, color=MUTED)
    ws.row_dimensions[13].height = 16

    # Höger sida av samma rad: per kvm + spann
    ws["E11"] = "PER KVM/ÅR"
    ws["E11"].font = _font(size=9, bold=True, color=MUTED)

    ws.merge_cells("E12:F12")
    ws["E12"] = DATA["kravhyra_per_kvm"]
    ws["E12"].font = Font(name="Segoe UI Light", size=36, color=INK)
    ws["E12"].number_format = '#,##0'
    ws["E12"].alignment = _align(h="left", v="center")

    ws["E13"] = "kr/kvm"
    ws["E13"].font = _font(size=10, color=MUTED)

    ws.row_dimensions[14].height = 32  # luft

    # ── Hyresspann — som balansräkning, två kolumner ─────────────────────
    ws["B15"] = "HYRESSPANN"
    ws["B15"].font = _font(size=9, bold=True, color=MUTED)
    for col in "BC":
        ws[f"{col}15"].border = _rule_bottom("thin", HAIRLINE)
    ws.row_dimensions[15].height = 22

    rows_left = [
        ("Lägsta utfall (−10%)", DATA["lagsta"]),
        ("Mål-utfall (bindande)", DATA["kravhyra"]),
        ("Högsta utfall (+10%)", DATA["hogsta"]),
    ]
    for i, (label, val) in enumerate(rows_left):
        r = 16 + i
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = _font(size=10, color=INK)
        ws[f"C{r}"] = val
        ws[f"C{r}"].font = _font(size=10, color=INK,
                                  bold=(label.startswith("Mål")))
        ws[f"C{r}"].number_format = '#,##0'
        ws[f"C{r}"].alignment = _align(h="right")
        ws.row_dimensions[r].height = 20

    # Top-rule på mål-utfall (rad 17)
    for col in "BC":
        ws[f"{col}17"].border = _rule_top("thin", HAIRLINE)
        ws[f"{col}17"].border = _rule_top_bottom("thin", HAIRLINE)

    # Höger kolumn: Lönsamhetskrav
    ws["E15"] = "LÖNSAMHETSKRAV"
    ws["E15"].font = _font(size=9, bold=True, color=MUTED)
    for col in "EF":
        ws[f"{col}15"].border = _rule_bottom("thin", HAIRLINE)

    rows_right = [
        ("NPV ≥ 0", "Uppfyllt"),
        ("IRR EK ≥ avkastningskrav", "Uppfyllt"),
        ("Marknadsvärde ≥ Bokfört värde", "Uppfyllt"),
    ]
    for i, (label, status) in enumerate(rows_right):
        r = 16 + i
        ws[f"E{r}"] = label
        ws[f"E{r}"].font = _font(size=10, color=INK)
        ws[f"F{r}"] = status
        ws[f"F{r}"].font = _font(size=10, color=POSITIVE, bold=True)
        ws[f"F{r}"].alignment = _align(h="right")
        ws.row_dimensions[r].height = 20

    # Faktisk IRR + marginal — egen rad under lönsamhetskrav
    ws["E20"] = "Faktisk IRR EK"
    ws["E20"].font = _font(size=10, color=INK, bold=True)
    ws["F20"] = DATA["irr_faktisk"]
    ws["F20"].font = _font(size=10, color=INK, bold=True)
    ws["F20"].number_format = '0.00%'
    ws["F20"].alignment = _align(h="right")
    for col in "EF":
        ws[f"{col}20"].border = _rule_top("thin", HAIRLINE)
    ws.row_dimensions[20].height = 22

    # ── Footer ───────────────────────────────────────────────────────────
    for r in range(21, 24):
        ws.row_dimensions[r].height = 14
    ws["B24"] = "Lejonfastigheter AB · Investeringskalkyl · 2026"
    ws["B24"].font = _font(size=8, color=MUTED, italic=True)
    ws.row_dimensions[24].height = 16


# ── VARIANT B: DASHBOARD HERO ──────────────────────────────────────────────
# KPI-shapes, accentfärg, levande typografi.

def build_oversikt_b(ws):
    """Översikt i dashboard hero-stil."""
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    # Bredder: KPI-grid med plats för stora tal (merga B:C, E:F, H:I, K:L per kort)
    widths = {"A": 2, "B": 22, "C": 10, "D": 2, "E": 22, "F": 10,
              "G": 2, "H": 18, "I": 10, "J": 2, "K": 20, "L": 10, "M": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Topp-banner: petrol bg ───────────────────────────────────────────
    for r in [1, 2, 3]:
        for col in "ABCDEFGHIJKLM":
            ws[f"{col}{r}"].fill = _fill(INK)
        ws.row_dimensions[r].height = 18

    ws["B2"] = "INVESTERINGSKALKYL"
    ws["B2"].font = _font(size=10, bold=True, color=PAPER)
    ws["B2"].alignment = _align(h="left", v="center")
    ws.row_dimensions[2].height = 36

    ws["B3"] = f"{DATA['project']} · Lejonfastigheter AB"
    ws["B3"].font = _font(size=18, color=PAPER, light=False)
    ws["B3"].font = Font(name="Segoe UI Light", size=20, color=PAPER)
    ws["B3"].alignment = _align(h="left", v="top")
    ws.row_dimensions[3].height = 38

    ws.row_dimensions[4].height = 20  # luft

    # ── 4 KPI-kort i rad ─────────────────────────────────────────────────
    # Rad 5-8 = en KPI-kort-grupp
    def kpi_card(start_col, end_col, label, value, fmt, footer, color=INK):
        """KPI-kort spänner start_col:end_col, rad 5-7."""
        label_r, val_r, foot_r = 5, 6, 7
        ws.merge_cells(f"{start_col}{label_r}:{end_col}{label_r}")
        ws.merge_cells(f"{start_col}{val_r}:{end_col}{val_r}")
        ws.merge_cells(f"{start_col}{foot_r}:{end_col}{foot_r}")
        ws[f"{start_col}{label_r}"] = label
        ws[f"{start_col}{label_r}"].font = _font(size=8, bold=True, color=MUTED)
        ws[f"{start_col}{label_r}"].alignment = _align(h="left", v="center")
        ws[f"{start_col}{val_r}"] = value
        ws[f"{start_col}{val_r}"].font = Font(name="Segoe UI Light", size=28, color=color)
        ws[f"{start_col}{val_r}"].number_format = fmt
        ws[f"{start_col}{val_r}"].alignment = _align(h="left", v="center")
        ws[f"{start_col}{foot_r}"] = footer
        ws[f"{start_col}{foot_r}"].font = _font(size=9, color=MUTED)
        ws[f"{start_col}{foot_r}"].alignment = _align(h="left", v="top")

    kpi_card("B", "C", "BINDANDE KRAVHYRA", DATA["kravhyra"], "#,##0", "kr/år")
    kpi_card("E", "F", "PER KVM/ÅR", DATA["kravhyra_per_kvm"], "#,##0", "kr/kvm")
    kpi_card("H", "I", "FAKTISK IRR EK", DATA["irr_faktisk"], "0.00%", "Krav 6,30%", color=POSITIVE)
    kpi_card("K", "L", "MARGINAL MOT KRAV", DATA["marginal"], "+0.00%", "procentenheter", color=POSITIVE)

    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 40
    ws.row_dimensions[7].height = 18

    # Subtil hairline-rule under kort-raden
    for col in "BCDEFGHIJKL":
        ws[f"{col}8"].border = _rule_top("thin", HAIRLINE)
    ws.row_dimensions[8].height = 6

    ws.row_dimensions[9].height = 20  # luft

    # ── Investeringsruta (vänster) + hyresspann (höger) ──────────────────
    # Sektion-band: accent-bar + label
    ws["B10"] = "INVESTERING"
    ws["B10"].font = _font(size=9, bold=True, color=INK)
    # Tunn accent-bar — använd top-border som hairline + ackumulera
    ws["B10"].border = Border(left=Side(style="medium", color=ACCENT))
    ws["B10"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[10].height = 22

    rows_left = [
        ("Total investering", DATA["investering"], "#,##0 \"kr\""),
        ("Investering/kvm", DATA["kr_kvm"], "#,##0 \"kr/kvm\""),
        ("Area", DATA["area"], "#,##0 \"m²\""),
        ("Kalkylperiod", DATA["period"], "0 \"år\""),
        ("Belåningsgrad", DATA["belaning"], "0.0%"),
    ]
    for i, (label, val, fmt) in enumerate(rows_left):
        r = 11 + i
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = _font(size=10, color=MUTED)
        ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws[f"C{r}"] = val
        ws[f"C{r}"].font = _font(size=10, color=INK)
        ws[f"C{r}"].number_format = fmt
        ws[f"C{r}"].alignment = _align(h="right")
        ws.row_dimensions[r].height = 22

    # Höger sektion
    ws["E10"] = "HYRESSPANN"
    ws["E10"].font = _font(size=9, bold=True, color=INK)
    ws["E10"].border = Border(left=Side(style="medium", color=ACCENT))
    ws["E10"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    rows_right = [
        ("Lägsta utfall (−10%)", DATA["lagsta"]),
        ("Mål-utfall (bindande)", DATA["kravhyra"]),
        ("Högsta utfall (+10%)", DATA["hogsta"]),
    ]
    for i, (label, val) in enumerate(rows_right):
        r = 11 + i
        ws[f"E{r}"] = label
        is_target = label.startswith("Mål")
        ws[f"E{r}"].font = _font(size=10, color=INK if is_target else MUTED,
                                 bold=is_target)
        ws[f"E{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws[f"F{r}"] = val
        ws[f"F{r}"].font = _font(size=10, color=INK, bold=is_target)
        ws[f"F{r}"].number_format = '#,##0'
        ws[f"F{r}"].alignment = _align(h="right")
        ws.row_dimensions[r].height = 22

    # Lönsamhetskrav (höger nedre)
    ws["H10"] = "LÖNSAMHETSKRAV"
    ws["H10"].font = _font(size=9, bold=True, color=INK)
    ws["H10"].border = Border(left=Side(style="medium", color=POSITIVE))
    ws["H10"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    krav = [
        ("NPV ≥ 0", "✓"),
        ("IRR EK ≥ avkastningskrav", "✓"),
        ("MV år 20 ≥ Bokfört värde", "✓"),
    ]
    for i, (label, status) in enumerate(krav):
        r = 11 + i
        ws[f"H{r}"] = label
        ws[f"H{r}"].font = _font(size=10, color=MUTED)
        ws[f"H{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws[f"I{r}"] = status
        ws[f"I{r}"].font = _font(size=14, bold=True, color=POSITIVE)
        ws[f"I{r}"].alignment = _align(h="center")

    # ── Footer ───────────────────────────────────────────────────────────
    for r in range(16, 19):
        ws.row_dimensions[r].height = 14
    ws["B19"] = "Lejonfastigheter AB · Investeringskalkyl · 2026"
    ws["B19"].font = _font(size=8, color=MUTED, italic=True)
    ws.row_dimensions[19].height = 16


# ── Resultat-varianter ─────────────────────────────────────────────────────

def build_resultat_a(ws):
    """Resultat i editorial-stil — som en balansräkning."""
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    widths = {"A": 3, "B": 32, "C": 16, "D": 16, "E": 16, "F": 3}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for r in range(1, 4):
        ws.row_dimensions[r].height = 14

    ws["B4"] = "Resultat"
    ws["B4"].font = Font(name="Segoe UI Light", size=24, color=INK)
    ws.row_dimensions[4].height = 38

    ws["B5"] = f"Kravhyra och hyresspann · {DATA['project']}"
    ws["B5"].font = _font(size=10, color=MUTED, italic=True)
    ws.row_dimensions[5].height = 18

    # Rule under titel
    for col in "BCDEF":
        ws[f"{col}6"].border = _rule_bottom("thin", HAIRLINE)
    ws.row_dimensions[6].height = 4
    ws.row_dimensions[7].height = 24

    # ── Tabell: KRAVHYRA — PROJEKTETS ÅRSHYRA ─────────────────────────────
    # Tabell-header
    ws["B8"] = "Belopp i kr"
    ws["B8"].font = _font(size=9, bold=True, color=MUTED)
    ws["C8"] = "Lägsta (−10%)"
    ws["D8"] = "Mål"
    ws["E8"] = "Högsta (+10%)"
    for col in "CDE":
        ws[f"{col}8"].font = _font(size=9, bold=True, color=MUTED)
        ws[f"{col}8"].alignment = _align(h="right")
    # Rule under header
    for col in "BCDE":
        ws[f"{col}8"].border = _rule_bottom("thin", HAIRLINE)
    ws.row_dimensions[8].height = 22

    # Datarader
    rows = [
        ("Investering",            180_000_000, 200_000_000, 220_000_000, "#,##0"),
        ("Investering per kvm",    36_000, 40_000, 44_000, "#,##0"),
        ("",                       None, None, None, None),  # luft
        ("Kravhyra för NPV ≥ 0",   10_726_244, 11_631_222, 12_536_199, "#,##0"),
        ("Kravhyra för IRR ≥ avkastningskrav", 10_507_485, 11_129_728, 11_751_970, "#,##0"),
        ("Kravhyra för MV ≥ BV år 20", 5_426_681, 5_729_582, 6_032_483, "#,##0"),
    ]
    for i, row in enumerate(rows):
        label, *vals = row
        r = 9 + i
        if label:
            ws[f"B{r}"] = label
            ws[f"B{r}"].font = _font(size=10, color=INK)
            ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for j, col in enumerate("CDE"):
                if vals[j] is not None:
                    ws[f"{col}{r}"] = vals[j]
                    ws[f"{col}{r}"].font = _font(size=10, color=INK)
                    ws[f"{col}{r}"].number_format = vals[3] if isinstance(vals[3], str) else "#,##0"
                    ws[f"{col}{r}"].alignment = _align(h="right")
        ws.row_dimensions[r].height = 20

    # ── BINDANDE KRAVHYRA — summa-rad ────────────────────────────────────
    r_summa = 16
    ws[f"B{r_summa}"] = "Bindande kravhyra"
    ws[f"B{r_summa}"].font = _font(size=10, bold=True, color=INK)
    ws[f"B{r_summa}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for j, (col, val) in enumerate(zip("CDE", [DATA["lagsta"], DATA["kravhyra"], DATA["hogsta"]])):
        ws[f"{col}{r_summa}"] = val
        ws[f"{col}{r_summa}"].font = _font(size=12, bold=True, color=INK)
        ws[f"{col}{r_summa}"].number_format = '#,##0'
        ws[f"{col}{r_summa}"].alignment = _align(h="right")
    for col in "BCDE":
        ws[f"{col}{r_summa}"].border = _rule_top_bottom("thin", HAIRLINE)
    ws.row_dimensions[r_summa].height = 28

    # Genomsnitt kr/kvm/år
    r_avg = 17
    ws[f"B{r_avg}"] = "Genomsnitt kr/m²/år"
    ws[f"B{r_avg}"].font = _font(size=9, color=MUTED, italic=True)
    ws[f"B{r_avg}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col, val in zip("CDE", [2_145, 2_326, 2_507]):
        ws[f"{col}{r_avg}"] = val
        ws[f"{col}{r_avg}"].font = _font(size=9, color=MUTED, italic=True)
        ws[f"{col}{r_avg}"].number_format = '#,##0'
        ws[f"{col}{r_avg}"].alignment = _align(h="right")
    ws.row_dimensions[r_avg].height = 20

    ws.row_dimensions[18].height = 28  # luft

    # ── Tolkning ─────────────────────────────────────────────────────────
    ws["B19"] = "TOLKNING"
    ws["B19"].font = _font(size=9, bold=True, color=MUTED)
    ws.row_dimensions[19].height = 18

    ws["B20"] = (f"Bindande kravhyra är {DATA['kravhyra']:,} kr/år vid mål-utfall "
                 .replace(",", " ")
                 + f"({DATA['lagsta']:,}–{DATA['hogsta']:,} kr/år vid ±10% investering)."
                 .replace(",", " "))
    ws["B20"].font = _font(size=10, color=INK)
    ws["B20"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=0)
    ws.merge_cells("B20:E22")
    ws.row_dimensions[20].height = 18
    ws.row_dimensions[21].height = 18
    ws.row_dimensions[22].height = 18

    # Footer
    ws.row_dimensions[24].height = 16
    ws["B25"] = "Lejonfastigheter AB · Investeringskalkyl · 2026"
    ws["B25"].font = _font(size=8, color=MUTED, italic=True)


def build_resultat_b(ws):
    """Resultat i dashboard-stil — KPI-band ovan + tabell under."""
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    widths = {"A": 2, "B": 22, "C": 16, "D": 4, "E": 18, "F": 18, "G": 18, "H": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Top-banner
    for r in [1, 2, 3]:
        for col in "ABCDEFGH":
            ws[f"{col}{r}"].fill = _fill(INK)
        ws.row_dimensions[r].height = 14

    ws["B2"] = "RESULTAT"
    ws["B2"].font = _font(size=10, bold=True, color=PAPER)
    ws.row_dimensions[2].height = 36
    ws["B3"] = f"Kravhyra och hyresspann · {DATA['project']}"
    ws["B3"].font = Font(name="Segoe UI Light", size=18, color=PAPER)
    ws.row_dimensions[3].height = 36

    ws.row_dimensions[4].height = 20

    # ── Hero-rad: en stor KPI ────────────────────────────────────────────
    ws["B5"] = "BINDANDE KRAVHYRA"
    ws["B5"].font = _font(size=9, bold=True, color=MUTED)
    ws.row_dimensions[5].height = 18

    ws["B6"] = DATA["kravhyra"]
    ws["B6"].font = Font(name="Segoe UI Light", size=42, color=INK)
    ws["B6"].number_format = "#,##0"
    ws.row_dimensions[6].height = 60

    ws["B7"] = "kr/år"
    ws["B7"].font = _font(size=11, color=MUTED)
    ws.row_dimensions[7].height = 18

    # Höger om: spann-mini
    ws["E5"] = "INTERVALL"
    ws["E5"].font = _font(size=9, bold=True, color=MUTED)
    ws["E6"] = f"{DATA['lagsta']:,}".replace(",", " ")
    ws["E6"].font = Font(name="Segoe UI Light", size=18, color=MUTED)
    ws["E7"] = "Lägsta"
    ws["E7"].font = _font(size=8, color=MUTED)

    ws["F6"] = f"{DATA['kravhyra']:,}".replace(",", " ")
    ws["F6"].font = Font(name="Segoe UI Light", size=22, color=INK)
    ws["F6"].fill = _fill("FFF6D9")
    ws["F6"].alignment = _align(h="center")
    ws["F7"] = "Mål"
    ws["F7"].font = _font(size=8, color=INK, bold=True)
    ws["F7"].alignment = _align(h="center")

    ws["G6"] = f"{DATA['hogsta']:,}".replace(",", " ")
    ws["G6"].font = Font(name="Segoe UI Light", size=18, color=MUTED)
    ws["G7"] = "Högsta"
    ws["G7"].font = _font(size=8, color=MUTED)

    ws.row_dimensions[8].height = 28

    # Hairline-separator
    for col in "BCDEFG":
        ws[f"{col}9"].border = _rule_top("thin", HAIRLINE)
    ws.row_dimensions[9].height = 4
    ws.row_dimensions[10].height = 18

    # ── Detaljtabell ─────────────────────────────────────────────────────
    ws["B11"] = "DELKRAVHYROR PER LÖNSAMHETSKRAV"
    ws["B11"].font = _font(size=9, bold=True, color=INK)
    ws["B11"].border = Border(left=Side(style="medium", color=ACCENT))
    ws["B11"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[11].height = 24

    # Tabell-header
    ws["B13"] = "Krav"
    ws["E13"] = "Lägsta"
    ws["F13"] = "Mål"
    ws["G13"] = "Högsta"
    for col in "BEFG":
        ws[f"{col}13"].font = _font(size=9, bold=True, color=MUTED)
        if col != "B":
            ws[f"{col}13"].alignment = _align(h="right")
    for col in "BCDEFG":
        ws[f"{col}13"].border = _rule_bottom("thin", HAIRLINE)
    ws.row_dimensions[13].height = 22

    rows = [
        ("NPV ≥ 0",                10_726_244, 11_631_222, 12_536_199),
        ("IRR ≥ avkastningskrav",  10_507_485, 11_129_728, 11_751_970),
        ("MV ≥ Bokfört värde",     5_426_681,  5_729_582,  6_032_483),
    ]
    for i, (label, *vals) in enumerate(rows):
        r = 14 + i
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = _font(size=10, color=INK)
        ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for col, val in zip("EFG", vals):
            ws[f"{col}{r}"] = val
            ws[f"{col}{r}"].font = _font(size=10, color=INK,
                                          bold=(col == "F"))
            ws[f"{col}{r}"].number_format = "#,##0"
            ws[f"{col}{r}"].alignment = _align(h="right")
        ws.row_dimensions[r].height = 22

    # Footer
    ws.row_dimensions[19].height = 16
    ws["B20"] = "Lejonfastigheter AB · Investeringskalkyl · 2026"
    ws["B20"].font = _font(size=8, color=MUTED, italic=True)


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    out = ROOT / "build" / "comps.xlsx"
    wb = Workbook()
    # Default sheet
    wb.remove(wb.active)

    ws_a1 = wb.create_sheet("A · Översikt")
    build_oversikt_a(ws_a1)

    ws_a2 = wb.create_sheet("A · Resultat")
    build_resultat_a(ws_a2)

    ws_b1 = wb.create_sheet("B · Översikt")
    build_oversikt_b(ws_b1)

    ws_b2 = wb.create_sheet("B · Resultat")
    build_resultat_b(ws_b2)

    # Page setup på alla
    for ws in wb.worksheets:
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

    wb.save(out)
    print(f"Sparat: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
