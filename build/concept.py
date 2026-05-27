"""Concept — Investeringspresentation med foto, charts och asymmetri.

En enda flik (Översikt) byggd som en mini-årsredovisning för Lejonfastigheter.
Bild hero överst, KPI-band, driftnetto-chart, känslighets-band.
"""
from __future__ import annotations
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.fill import SolidColorFillProperties, ColorChoice as FillColor
from openpyxl.chart.marker import DataPoint

# ── Palett ─────────────────────────────────────────────────────────────────
INK      = "10313E"
MUTED    = "6C757D"
PAPER    = "FFFFFF"
SURFACE  = "F4F6F8"
ACCENT   = "FAB600"
POSITIVE = "00937C"
HAIRLINE = "DEE2E6"

FAMILY = "Segoe UI"


# Demo-data
DATA = {
    "project": "Skola",
    "fastighet": "Linköping kommun, kv. Tornet 5",
    "story": ("Nybyggnad av skola för 5 000 m² verksamhetsyta. "
              "Investering 200 Mkr över 20 år. "
              "Lejonfastigheter projektledare och fastighetsägare."),
    "kalkylstart": 2026,
    "period": 20,
    "investering": 200_000_000,
    "area": 5000,
    "kravhyra": 11_631_222,
    "kravhyra_per_kvm": 2_326,
    "lagsta": 10_726_244,
    "hogsta": 12_536_199,
    "irr_faktisk": 0.0765,
    "irr_krav": 0.063,
    "marginal": 0.0135,
    # 20-årig driftnetto-projektion (växer med inflation 2%)
    "driftnetto": [
        9_231_222, 9_546_059, 9_462_216, 9_579_705, 9_690_539, 9_818_732,
        9_940_295, 10_063_343, 10_187_507, 10_313_341, 10_440_519, 10_569_133,
        10_699_196, 10_830_722, 10_963_724, 11_098_216, 11_234_210, 11_371_721,
        11_510_762, 11_651_346
    ],
}


def _font(size=10, bold=False, color=INK, italic=False, light=False, family=None):
    name = family or FAMILY
    if light:
        name = "Segoe UI Light"
    elif bold:
        name = "Segoe UI Semibold"
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def _fill(c): return PatternFill("solid", fgColor=c)
def _align(h="left", v="center", indent=0, wrap=False):
    return Alignment(horizontal=h, vertical=v, indent=indent, wrap_text=wrap)
def _rule_bottom(color=HAIRLINE):
    return Border(bottom=Side(style="thin", color=color))
def _rule_top(color=HAIRLINE):
    return Border(top=Side(style="thin", color=color))


def build_oversikt(ws):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    # ── Layout: 14 kolumner i landscape ──────────────────────────────────
    # A=marginal, B-D=zone1, E=sep, F-H=zone2, I=sep, J-L=zone3, M=marginal
    widths = {
        "A": 2,
        "B": 16, "C": 14, "D": 8,
        "E": 3,
        "F": 16, "G": 14, "H": 8,
        "I": 3,
        "J": 18, "K": 12, "L": 8,
        "M": 2,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Hero-bild (rad 1-14) ─────────────────────────────────────────────
    hero_path = ROOT / "assets" / "hero.png"
    if hero_path.exists():
        img = XLImage(str(hero_path))
        # Bredd ~ summa av kolumner * 7px = 105*7 ≈ 735 → 1200 för premium
        img.width = 1200
        img.height = 380
        ws.add_image(img, "A1")

    # Reservera rad-höjd för bilden (rader 1-14, ~380pt totalt)
    for r in range(1, 15):
        ws.row_dimensions[r].height = 20

    # ── Etikett-strip: liten varumärkesetikett UNDER bilden ──────────────
    ws.row_dimensions[15].height = 8

    ws.merge_cells("B16:D16")
    ws["B16"] = "LEJONFASTIGHETER · INVESTERINGSANALYS"
    ws["B16"].font = _font(size=8, bold=True, color=ACCENT)
    ws["B16"].alignment = _align(h="left", v="center")
    ws.row_dimensions[16].height = 16

    # ── Titel UNDER hero — stor Light typografi ──────────────────────────
    ws.merge_cells("B17:L17")
    ws["B17"] = "Investeringskalkyl"
    ws["B17"].font = Font(name="Segoe UI Light", size=42, color=INK)
    ws["B17"].alignment = Alignment(horizontal="left", vertical="bottom")
    ws.row_dimensions[17].height = 52

    ws.merge_cells("B18:L18")
    ws["B18"] = f"{DATA['project']} · {DATA['fastighet']} · Kalkylstart {DATA['kalkylstart']}"
    ws["B18"].font = _font(size=12, color=MUTED, italic=False)
    ws["B18"].alignment = Alignment(horizontal="left", vertical="top")
    ws.row_dimensions[18].height = 22

    # Tunn separator
    for col in "BCDEFGHIJKL":
        ws[f"{col}19"].border = _rule_bottom()
    ws.row_dimensions[19].height = 14

    # ── Story-rad ────────────────────────────────────────────────────────
    ws.merge_cells("B20:L20")
    ws["B20"] = DATA["story"]
    ws["B20"].font = _font(size=11, color=INK)
    ws["B20"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[20].height = 22

    # ── KPI-band: 3 stora KPI:er i rad ───────────────────────────────────
    def kpi(start_col, end_col, label, value, fmt, unit, color=INK):
        """Stor KPI: liten label överst, stort tal, liten enhet under."""
        ws.merge_cells(f"{start_col}21:{end_col}21")
        ws[f"{start_col}21"] = label
        ws[f"{start_col}21"].font = _font(size=9, bold=True, color=MUTED)
        ws[f"{start_col}21"].alignment = _align(h="left", v="bottom", indent=1)

        ws.merge_cells(f"{start_col}22:{end_col}22")
        ws[f"{start_col}22"] = value
        ws[f"{start_col}22"].font = Font(name="Segoe UI Light", size=44, color=color)
        ws[f"{start_col}22"].number_format = fmt
        ws[f"{start_col}22"].alignment = _align(h="left", v="center", indent=1)

        ws.merge_cells(f"{start_col}23:{end_col}23")
        ws[f"{start_col}23"] = unit
        ws[f"{start_col}23"].font = _font(size=10, color=MUTED)
        ws[f"{start_col}23"].alignment = _align(h="left", v="top", indent=1)

    kpi("B", "D", "BINDANDE KRAVHYRA", DATA["kravhyra"], "#,##0", "kr/år")
    kpi("F", "H", "PER KVM/ÅR", DATA["kravhyra_per_kvm"], "#,##0", "kr/m²")
    kpi("J", "L", "FAKTISK IRR EK", DATA["irr_faktisk"], "0.00%",
        f"Marginal +{DATA['marginal']*100:.2f} pp · Krav {DATA['irr_krav']*100:.1f}%",
        color=POSITIVE)

    ws.row_dimensions[21].height = 18
    ws.row_dimensions[22].height = 60
    ws.row_dimensions[23].height = 22

    ws.row_dimensions[24].height = 28  # luft

    # ── Driftnetto-tidslinje: typografisk visualisering ──────────────────
    ws.merge_cells("F25:L25")
    ws["F25"] = "DRIFTNETTO ÖVER KALKYLPERIODEN"
    ws["F25"].font = _font(size=9, bold=True, color=MUTED)
    ws["F25"].alignment = _align(h="left", v="bottom", indent=1)
    ws.row_dimensions[25].height = 22

    driftnetto = DATA["driftnetto"]
    # Tre ankarvärden: start / mitten / slut
    start_val = driftnetto[0]
    mid_val = driftnetto[9]   # år 10
    end_val = driftnetto[-1]

    # Tre stora tal i Light typografi, jämnt fördelade
    # F-G för start, I-J för mitten, L för slut
    ws.merge_cells("F27:G27")
    ws["F27"] = "ÅR 1 · 2026"
    ws["F27"].font = _font(size=8, bold=True, color=MUTED)
    ws["F27"].alignment = _align(h="left", v="bottom", indent=1)

    ws.merge_cells("F28:G28")
    ws["F28"] = start_val
    ws["F28"].font = Font(name="Segoe UI Light", size=22, color=MUTED)
    ws["F28"].number_format = "#,##0"
    ws["F28"].alignment = _align(h="left", v="center", indent=1)

    ws.merge_cells("F29:G29")
    ws["F29"] = "kr"
    ws["F29"].font = _font(size=9, color=MUTED)
    ws["F29"].alignment = _align(h="left", v="top", indent=1)

    # Mitten — bara en visuell "→"-pil med tunt streck
    ws.merge_cells("H28:I28")
    ws["H28"] = "─────→"
    ws["H28"].font = _font(size=14, color=MUTED)
    ws["H28"].alignment = _align(h="center", v="center")

    # Slut
    ws.merge_cells("J27:L27")
    ws["J27"] = "ÅR 20 · 2045"
    ws["J27"].font = _font(size=8, bold=True, color=INK)
    ws["J27"].alignment = _align(h="left", v="bottom", indent=1)

    ws.merge_cells("J28:L28")
    ws["J28"] = end_val
    ws["J28"].font = Font(name="Segoe UI Light", size=22, color=INK)
    ws["J28"].number_format = "#,##0"
    ws["J28"].alignment = _align(h="left", v="center", indent=1)

    ws.merge_cells("J29:L29")
    ws["J29"] = f"+{(end_val/start_val-1)*100:.1f}% tillväxt"
    ws["J29"].font = _font(size=9, bold=True, color=POSITIVE)
    ws["J29"].alignment = _align(h="left", v="top", indent=1)

    for r in [27, 28, 29]:
        ws.row_dimensions[r].height = 22 if r != 28 else 40

    # Horisontell tidslinje (hairline-rule) som visualiserar progression
    for col in "FGHIJKL":
        ws[f"{col}31"].border = _rule_top(color=INK)
    ws.row_dimensions[31].height = 4

    # Reservera rader för layout-balans
    for r in range(32, 42):
        ws.row_dimensions[r].height = 14

    # ── Vänster zon: investeringssammanfattning som "panel" ──────────────
    ws.merge_cells("B25:D25")
    ws["B25"] = "INVESTERING"
    ws["B25"].font = _font(size=9, bold=True, color=MUTED)
    ws["B25"].alignment = _align(h="left", v="bottom", indent=1)

    panel_rows = [
        ("Total investering", DATA["investering"], '#,##0 "kr"'),
        ("Investering / m²",  DATA["investering"] // DATA["area"], '#,##0 "kr"'),
        ("Verksamhetsyta",    DATA["area"], '#,##0 "m²"'),
        ("Kalkylperiod",      DATA["period"], '0 "år"'),
        ("Belåningsgrad",     0.37, "0.0%"),
        ("Avkastningskrav EK", 0.063, "0.0%"),
    ]
    for i, (label, val, fmt) in enumerate(panel_rows):
        r = 27 + i
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = _font(size=10, color=MUTED)
        ws[f"B{r}"].alignment = _align(h="left", v="center", indent=1)
        ws.merge_cells(f"C{r}:D{r}")
        ws[f"C{r}"] = val
        ws[f"C{r}"].font = _font(size=11, color=INK)
        ws[f"C{r}"].number_format = fmt
        ws[f"C{r}"].alignment = _align(h="right", v="center", indent=1)
        ws.row_dimensions[r].height = 22

        # Subtil hairline mellan rader
        if i < len(panel_rows) - 1:
            for col in "BCD":
                ws[f"{col}{r}"].border = _rule_bottom()

    # ── Hyresspann-band: visuell horisontell bar ─────────────────────────
    # Rad 44-46: band-visualisering
    spann_row = 44
    ws.merge_cells(f"B{spann_row}:L{spann_row}")
    ws[f"B{spann_row}"] = "HYRESSPANN VID INVESTERINGSUTFALL"
    ws[f"B{spann_row}"].font = _font(size=9, bold=True, color=MUTED)
    ws[f"B{spann_row}"].alignment = _align(h="left", v="bottom", indent=1)
    ws.row_dimensions[spann_row].height = 22

    # Bar via cells: hela bredden = lägsta till högsta, mål markerad
    # Vi använder 11 celler (B-L) som "bar"
    bar_row = 45
    ws.row_dimensions[bar_row].height = 18

    # Underlag: ljusgrå genomgående
    for col in "BCDEFGHIJKL":
        ws[f"{col}{bar_row}"].fill = _fill(SURFACE)
        ws[f"{col}{bar_row}"].border = Border()

    # Mål-cell: gul accent (mitten = col G)
    ws.merge_cells(f"F{bar_row}:H{bar_row}")
    ws[f"F{bar_row}"].fill = _fill(ACCENT)

    # Etikett-rad under
    ws[f"B46"] = f"{DATA['lagsta']:,}".replace(",", " ")
    ws[f"B46"].font = _font(size=10, color=MUTED)
    ws[f"B46"].alignment = _align(h="left", indent=1)

    ws.merge_cells("F46:H46")
    ws[f"F46"] = f"{DATA['kravhyra']:,}".replace(",", " ") + " kr/år"
    ws[f"F46"].font = Font(name="Segoe UI Semibold", size=11, color=INK, bold=True)
    ws[f"F46"].alignment = _align(h="center")

    ws.merge_cells("J46:L46")
    ws[f"J46"] = f"{DATA['hogsta']:,}".replace(",", " ")
    ws[f"J46"].font = _font(size=10, color=MUTED)
    ws[f"J46"].alignment = _align(h="right")

    ws.row_dimensions[46].height = 22

    ws["B47"] = "Lägsta utfall (−10%)"
    ws["B47"].font = _font(size=9, color=MUTED, italic=True)
    ws["B47"].alignment = _align(h="left", indent=1)
    ws.merge_cells("F47:H47")
    ws["F47"] = "Mål-utfall (bindande)"
    ws["F47"].font = _font(size=9, color=MUTED, italic=True)
    ws["F47"].alignment = _align(h="center")
    ws.merge_cells("J47:L47")
    ws["J47"] = "Högsta utfall (+10%)"
    ws["J47"].font = _font(size=9, color=MUTED, italic=True)
    ws["J47"].alignment = _align(h="right")
    ws.row_dimensions[47].height = 18

    # Footer
    ws.row_dimensions[48].height = 28
    ws.merge_cells("B49:L49")
    ws["B49"] = "Lejonfastigheter AB · Investeringskalkyl · 2026"
    ws["B49"].font = _font(size=8, color=MUTED, italic=True)
    ws["B49"].alignment = _align(h="left", indent=1)


def main():
    out = ROOT / "build" / "concept.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Översikt")
    build_oversikt(ws)

    # Page setup för landscape A4 — naturlig scale (inte fit-to-page)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 0
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_setup.scale = 70
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3

    wb.save(out)
    print(f"Sparat: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
