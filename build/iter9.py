"""Iter 9 — bygger ovanpå iter 8 med §10-uppgifter.

Varje funktion = en §10-uppgift. main() kör i sekvens. Idempotent: kan
köras om utan att stega framåt mer än en gång (text-replaces matchar inte
redan-patchade celler).

Genomförda uppgifter (commit-historik förklarar):
- round E: 'år N' / 'år N+1' → 'år 20' / 'år 21' (explicit horisont)
- round B: pedagogisk omskrivning av Beräkningslogik-text (användarcentrerad)
- round C: Indata-fält renamning till branschterminologi
- round F: designcleanup Lönsamhetskontroll (outline grouping av hjälprader)
- round G: Översikt om till beslutsdokument (projektinfo + kalkylantaganden)
"""
from __future__ import annotations
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from tools.patches import load_iter, save_iter, patch_summary_below, ITER8
from tools.recalc import recalc
from tools.styling import (
    style_oversikt, style_indata, style_resultat, section_header,
    accent_value, PRIMARY, ACCENT_BG, LIGHT, RULE_CLR, MUTED, FONT_FAM, FONT_FAM_BOLD,
    NAVY_TEXT, SECONDARY, POSITIVE,
)
from openpyxl.styles import Font, Alignment, Border, Side
from tests.regression import check_baseline

NO_FILL = PatternFill(fill_type=None)


def round_e_year_n(wb: Workbook) -> int:
    """§10 punkt 4: 'år N' → 'år 20', 'år N+1' → 'år 21'.

    Kalkylperioden är 20 år (Indata!C16) och refereras med variabeln N i text.
    Round E gör horisonten explicit för läsaren.

    OBS: 'år N+1' måste replaceas FÖRE 'år N' (annars matchar 'år N' delvis).
    Lämnar formelvariabeln 'N × avskrivnings-...' i Dokumentation!B31 orörd —
    den är en algebraisk variabel, inte en horisontetikett.
    """
    replacements = [
        ("år N+1", "år 21"),
        ("år N",   "år 20"),
        ("År N",   "år 20"),
    ]
    n_changes = 0
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for c in row:
                if not isinstance(c.value, str):
                    continue
                original = c.value
                new_val = original
                for old, new in replacements:
                    new_val = new_val.replace(old, new)
                if new_val != original:
                    c.value = new_val
                    n_changes += 1
    return n_changes


def round_c_indata_rename(wb: Workbook) -> int:
    """§10: Indata-fält renamning till branschterminologi.

    Tre celler byter etikett:
    - Indata!B18: 'Avskrivnings-% per år' → 'Avskrivningstakt'
    - Indata!B62: 'Långsiktig ränta' → 'Räntenivå' (matchar Finansiering!B14)
    - Lönsamhetskontroll!B26: 'Avskrivnings-%' → 'Avskrivningstakt'
    """
    renames = [
        ("Indata", "B18", "Avskrivnings-% per år", "Avskrivningstakt"),
        ("Indata", "B62", "Långsiktig ränta", "Räntenivå"),
        ("Lönsamhetskontroll", "B26", "Avskrivnings-%", "Avskrivningstakt"),
    ]
    n = 0
    for sheet, coord, old, new in renames:
        cell = wb[sheet][coord]
        if cell.value == old:
            cell.value = new
            n += 1
    return n


def round_b_pedagogisk(wb: Workbook) -> int:
    """§10 punkt 1: pedagogisk omskrivning av Beräkningslogik-text (round B).

    Inriktning: användarcentrerat — vad mallen ger användaren, inte bara
    metodbeskrivning. Behåller B33-44 (Goal Seek-jämförelse + fördelar) och
    block-rubriker B59/B121/B164/B179/B197 oförändrade.

    Körs efter round_e så 'år 20' i ny text inte stör replace-logiken.
    """
    new_text = {
        "B3": (
            "Du fyller i projektets förutsättningar i Indata. Mallen räknar baklänges "
            "till den årshyra som precis uppfyller respektive lönsamhetskrav: NPV ≥ 0, "
            "IRR EK ≥ avkastningskrav, samt marknadsvärde ≥ bokfört värde år 20. "
            "Den högsta av de tre kraven blir bindande kravhyra — golvet under vilket "
            "projektet inte är lönsamt. Lösningen är analytisk (ett steg, exakt) — "
            "beskrivet nedan."
        ),
        "B6":  "STEG 1 — Räkna NPV vid två testhyror (0 kr och 1 Mkr/år)",
        "B10": "STEG 2 — Härled hur mycket NPV ändras per krona hyra (lutningen b)",
        "B14": "STEG 3 — Lös linjärt: kravhyran är där NPV korsar noll, dvs −NPV(0) / b",
        "B47": (
            "Här under räknas de tre kravhyrorna ut. Block A ger NPV vid 0 kr hyra, "
            "block B vid 1 Mkr/år — differensen ger lutningen b för NPV-kravet. "
            "Block D-E gör samma för IRR EK. Block F gör marknadsvärdes-kravet. "
            "Detta är inte projektets faktiska kassaflöde — det visas på Kassaflöde-fliken "
            "vid bindande kravhyra."
        ),
    }
    ws = wb["Beräkningslogik"]
    for coord, text in new_text.items():
        ws[coord] = text
    return len(new_text)


def round_f_lonsamhetskontroll_cleanup(wb: Workbook) -> int:
    """§10 round F: Designcleanup av Lönsamhetskontroll.

    EK-cashflöde hjälprader (rad 43, 64-70) görs åtkomliga men visuellt
    separerade via Excel outline grouping (toggle [+] i radmarginalen).
    summaryBelow=False XML-patch körs av main() efter save_iter() så
    toggle hamnar OVANFÖR gruppen (rad 42 resp. rad 63).
    """
    ws = wb["Lönsamhetskontroll"]
    n = 0

    # Rad 43: etikett (synlig när expanderad), toggle hamnar på rad 42
    if ws["B43"].value is None:
        ws["B43"] = "EK-cashflöde år 0–20 (beräkningsunderlag för IRR ovan)"
        n += 1
    ws.row_dimensions[43].outlineLevel = 1
    ws.row_dimensions[43].hidden = True
    n += 1

    # Rad 63: synlig rubrik — indikator för gruppen nedanför
    if ws["B63"].value is None:
        ws["B63"] = "EK-cashflöde per scenario — klicka [+] för att expandera"
        n += 1

    # Rader 64-70: gruppera och kollapsa (toggle på rad 63)
    for row in range(64, 71):
        ws.row_dimensions[row].outlineLevel = 1
        ws.row_dimensions[row].hidden = True
    n += 7

    # Rader 35-41: tomma — dölj för att komprimera gap
    for row in range(35, 42):
        ws.row_dimensions[row].hidden = True
    n += 7

    return n


def round_g_oversikt_redesign(wb: Workbook) -> int:
    """§10 round G: Översikt om till beslutsdokument.

    Lägger till PROJEKTINFORMATION (redigerbara Fyll i-celler) och
    KALKYLANTAGANDEN (formler mot Indata) ovanför befintliga resultatblock.
    Innehållsförteckning skrivs om med korrekta hyperlänkar och förskjuts ner.
    Alla befintliga formler (Resultat, Lönsamhetskontroll) bevaras.
    """
    ws = wb["Översikt"]

    # Rensa rader 2-50 — unmerga först, sedan nolla värden
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    for r in range(2, 51):
        for c in range(1, 11):
            cell = ws.cell(r, c)
            cell.value = None
            cell.hyperlink = None

    def set_val(coord, val):
        ws[coord] = val

    # ── Titel ────────────────────────────────────────────────────────────────
    set_val("B2", "INVESTERINGSKALKYL — LEJONFASTIGHETER AB")
    set_val("B3", '=IF(Indata!B26="","(projekt ej namngivet)","Projekt: "&Indata!B26)')

    # ── A. PROJEKTINFORMATION ────────────────────────────────────────────────
    set_val("B5", "PROJEKTINFORMATION")
    set_val("B6", "Fastighet")
    set_val("C6", "Fyll i")
    set_val("E6", "Utförd av")
    set_val("F6", "Fyll i")
    set_val("B7", "Kalkylstart")
    set_val("C7", "=Indata!C5")
    set_val("E7", "Kalkylperiod")
    set_val("F7", '=Indata!C16&" år"')

    # ── B. KALKYLANTAGANDEN ──────────────────────────────────────────────────
    set_val("B9", "KALKYLANTAGANDEN")
    set_val("B10", "Kalkylränta driftnetto")
    set_val("C10", "=Indata!C7")
    set_val("E10", "IRR-krav EK")
    set_val("F10", "=Indata!C65")
    set_val("B11", "Direktavkastning marknad")
    set_val("C11", "=Indata!C6")
    set_val("E11", "Belåningsgrad")
    set_val("F11", "=Indata!C64")
    set_val("B12", "Inflation")
    set_val("C12", "=Indata!C8")

    # ── C. BINDANDE KRAVHYRA ─────────────────────────────────────────────────
    set_val("B14", "BINDANDE KRAVHYRA")
    set_val("B15", "Årshyra (mål-utfall)")
    set_val("C15", "=Resultat!D14")
    set_val("E15", "Bindande krav")
    set_val("F15", "=Resultat!D15")
    set_val("B16", "Per kvm/år")
    set_val("C16", "=Resultat!D16")
    set_val("E16", "Total area")
    set_val("F16", '=Indata!F31&" m²"')
    set_val("B17", "Total investering")
    set_val("C17", "=Indata!R31")
    set_val("E17", "Investering per kvm")
    set_val("F17", "=IFERROR(Indata!R31/Indata!F31,0)")

    # ── D. HYRESSPANN ────────────────────────────────────────────────────────
    set_val("B19", "HYRESSPANN VID INVESTERINGSUTFALL")
    set_val("B20", "Lägsta utfall (-X%)")
    set_val("C20", "=Resultat!C14")
    set_val("E20", "=Resultat!C7")
    set_val("B21", "Mål-utfall (bindande)")
    set_val("C21", "=Resultat!D14")
    set_val("E21", "=Resultat!D7")
    set_val("B22", "Högsta utfall (+X%)")
    set_val("C22", "=Resultat!E14")
    set_val("E22", "=Resultat!E7")

    # ── E. LÖNSAMHETSKRAV ────────────────────────────────────────────────────
    set_val("B24", "LÖNSAMHETSKRAV — STATUS VID BINDANDE KRAVHYRA")
    set_val("B25", "1. NPV ≥ 0")
    set_val("C25", "=Lönsamhetskontroll!C9")
    set_val("E25", "=Lönsamhetskontroll!D9")
    set_val("B26", "2. IRR EK ≥ avkastningskrav")
    set_val("C26", "=Lönsamhetskontroll!C18")
    set_val("E26", "=Lönsamhetskontroll!D18")
    set_val("B27", "3. MV år 20 ≥ Bokfört värde år 20")
    set_val("C27", "=Lönsamhetskontroll!C30")
    set_val("E27", "=Lönsamhetskontroll!D30")
    set_val("B28", "Faktisk IRR EK")
    set_val("C28", "=Lönsamhetskontroll!C45")
    set_val("E28", "Marginal mot krav")
    set_val("F28", "=Lönsamhetskontroll!C47")

    # ── F. INNEHÅLLSFÖRTECKNING ──────────────────────────────────────────────
    set_val("B30", "INNEHÅLLSFÖRTECKNING")

    toc = [
        (32, "Indata",              "Inmatning",            "Indata",
         "Marknadsförutsättningar, investering, hyresobjekt, drift, finansiering."),
        (34, "Resultat",            "Vad blir hyran?",      "Resultat",
         "Kravhyra för de tre lönsamhetskraven, bindande hyra och hyresspann."),
        (36, "Kassaflöde",          "Driftnetto år för år", "Kassaflöde",
         "Driftnetto-projektion vid bindande hyra (bruttohyra → driftnetto)."),
        (38, "Finansiering",        "Lån och avskrivning",  "Finansiering",
         "Lånebild: lån, amortering, räntekostnad, eget kapital, avskrivning."),
        (40, "Lönsamhetskontroll",  "Är kraven uppfyllda?", "Lönsamhetskontroll",
         "Verifiering av tre lönsamhetskrav, faktisk IRR EK och känslighetsanalys."),
        (42, "Beräkningslogik",     "Hur räknar mallen?",   "Beräkningslogik",
         "Delta-metoden + tekniska beräkningsblock A–F bakom kravhyran."),
        (44, "Dokumentation",       "Varför så här?",       "Dokumentation",
         "Designprinciper, val och paritetstest mot LM 371."),
    ]

    for row, name, tagline, target_sheet, desc in toc:
        cell = ws.cell(row, 2)
        cell.value = name
        cell.hyperlink = Hyperlink(
            ref=cell.coordinate,
            location=f"'{target_sheet}'!A1",
            display=name,
        )
        ws.cell(row, 3).value = tagline
        ws.cell(row, 5).value = desc

    return 1


def round_o_global_theme(wb: Workbook) -> int:
    """Round O: Globalt LF-tema över alla flikar.

    - Stäng av gridlines (showGridLines=False) — bort med 'Excel-känslan'
    - Stäng av rubrikrader (showRowColHeaders=False) på publika flikar
    - Sätt default tab-color till LF-petrol så fliklisten matchar temat
    """
    publika = ["Översikt", "Indata", "Kassaflöde", "Finansiering",
               "Resultat", "Lönsamhetskontroll", "Dokumentation"]
    intern = ["Beräkningslogik"]  # tekniska beräkningar — behåll headers för felsökning
    n = 0
    for sheet in publika + intern:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        ws.sheet_view.showGridLines = False
        if sheet in publika:
            ws.sheet_view.showRowColHeaders = False
        ws.sheet_properties.tabColor = PRIMARY
        n += 1
    return n


def round_p_zebra_oversikt(wb: Workbook) -> int:
    """Round P: Subtil zebra-bg + bottom-rules på Översikt + Resultat.

    Ger struktur utan border-rutnät. Datarader får mjuk LIGHT-bakgrund;
    sektionsavslut får tunn DEE2E6-linje.
    """
    n = 0

    # Översikt — zebra på datablocks i kolumn B:F
    ws = wb["Översikt"]
    from openpyxl.styles import PatternFill as PF
    from tools.theme import SURFACE, RULE
    light_fill = PF("solid", fgColor=SURFACE)
    rule_side = Side(style="thin", color=RULE)
    bottom_rule = Border(bottom=rule_side)

    # Block-definitioner: (start_row, end_row, last_col)
    blocks = [
        (6, 7, "F"),    # PROJEKTINFORMATION
        (10, 12, "F"),  # KALKYLANTAGANDEN
        (15, 17, "F"),  # BINDANDE KRAVHYRA
        (20, 22, "F"),  # HYRESSPANN
        (25, 28, "F"),  # LÖNSAMHETSKRAV
    ]
    for start, end, last in blocks:
        # zebra varannan rad i blocket
        for i, r in enumerate(range(start, end + 1)):
            if i % 2 == 1:  # varannan rad
                for col in "BCDEF":
                    cell = ws[f"{col}{r}"]
                    if cell.fill.fill_type is None:
                        cell.fill = light_fill
                        n += 1
        # bottom-rule på sista raden
        for col in "BCDEF":
            ws[f"{col}{end}"].border = bottom_rule
            n += 1

    # Resultat — bottom-rule på sista raden i Kravhyra-tabellen
    ws = wb["Resultat"]
    for col in "BCDE":
        cell = ws[f"{col}14"]
        cell.border = bottom_rule
        n += 1

    return n


def round_q_hero_block(wb: Workbook) -> int:
    """Round Q: Faktisk IRR EK som accent + Resultat D14 hero.

    Översikt C15 har redan hero via style_oversikt.
    """
    n = 0
    ws = wb["Översikt"]
    from tools.theme import apply, FMT_PCT2
    apply(ws["C28"], "hero")
    ws["C28"].number_format = FMT_PCT2
    n += 1

    ws = wb["Resultat"]
    if ws["D14"].value is not None:
        apply(ws["D14"], "hero")
        ws.row_dimensions[14].height = 32
        n += 1
    return n


def round_r_status_pills(wb: Workbook) -> int:
    """Round R: deprecated — styling sker via style_oversikt → apply(cell,'status')."""
    return 0


def round_s_footer(wb: Workbook) -> int:
    """Round S: Sidfot 'Lejonfastigheter · Investeringskalkyl v9 · 2026' på varje flik."""
    from tools.theme import MUTED as _MUTED
    n = 0
    footer_text = "Lejonfastigheter · Investeringskalkyl v9 · 2026"
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.oddFooter.left.text = footer_text
        ws.oddFooter.left.size = 9
        ws.oddFooter.left.color = _MUTED
        ws.oddFooter.right.text = "Sida &P / &N"
        ws.oddFooter.right.size = 8
        ws.oddFooter.right.color = MUTED
        n += 1
    return n


def round_i_page_setup(wb: Workbook) -> int:
    """Round I: Print area och page setup per flik.

    Drabbade flikar (utan rätt setup blir PDF-export 30% bred):
    Indata, Beräkningslogik, Dokumentation, Lönsamhetskontroll.
    Sätter print_area till faktiskt använt range, fit-to-width=1,
    landscape på breda flikar.
    """
    # (orientation, max_col, fit_to_width)
    # fit_to_width=2 → tillåt 2 sidor i bredd för 25-årsflikar (annars oläsligt)
    setup = {
        "Översikt":           ("portrait",  6,  1),
        "Indata":             ("landscape", 18, 1),
        "Kassaflöde":         ("landscape", 28, 2),  # 25 årskolumner → 2 sidor
        "Finansiering":       ("landscape", 28, 2),
        "Resultat":           ("portrait",  6,  1),
        "Lönsamhetskontroll": ("portrait",  6,  1),
        "Beräkningslogik":    ("landscape", 30, 2),  # 25 årskolumner + block → 2 sidor
        "Dokumentation":      ("portrait",  4,  1),
    }
    n = 0
    for sheet, (orient, max_col, fit_width) in setup.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        last_row = ws.max_row
        last_col = min(max_col, ws.max_column)
        col_letter = get_column_letter(last_col)
        ws.print_area = f"A1:{col_letter}{last_row}"
        ws.page_setup.orientation = orient
        ws.page_setup.fitToWidth = fit_width
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        n += 1
    return n


def round_j_oversikt_fixes(wb: Workbook) -> int:
    """Round J: Översikt header-fixes.

    - Merge title B2:F2 så hela 'INVESTERINGSKALKYL — LEJONFASTIGHETER AB' har bg
    - Merge alla sektion-rubriker B:F så bg-fyllning täcker hela radens bredd
    - Rensa residual mörkblå fyllning på Kalkylränta/Årshyra-rader (B10, B15)
    - Höj TOC-radhöjder så wrappad beskrivning inte klipps
    """
    ws = wb["Översikt"]
    n = 0

    # Merge title och underrad (subtitle) — styling redan satt via style_oversikt
    for merge_range in ["B2:F2", "B3:F3"]:
        try:
            ws.merge_cells(merge_range)
            n += 1
        except Exception:
            pass

    # Merge alla sektion-headers (full bredd B:F)
    section_rows = [5, 9, 14, 19, 24, 30]
    for r in section_rows:
        try:
            ws.merge_cells(f"B{r}:F{r}")
            n += 1
        except Exception:
            pass

    # Rensa residual iter8-fyllning på dataceller som av misstag fått mörk bg
    for coord in ["B10", "C10", "B15", "C15"]:
        ws[coord].fill = NO_FILL
        n += 1

    # TOC-rader: höj höjd så wrap-text inte klipps
    for r in [32, 34, 36, 38, 40, 42, 44]:
        ws.row_dimensions[r].height = 32
        n += 1

    return n


def round_k_format_konsistens(wb: Workbook) -> int:
    """Round K: Talformat — parenteser → minustecken för negativa tal.

    iter8 använde '(#,##0)' för negativa belopp; svensk konvention är
    minustecken. Walkar alla celler i Kassaflöde + Finansiering och
    byter format.
    """
    sheets = ["Kassaflöde", "Finansiering", "Beräkningslogik", "Lönsamhetskontroll"]
    n = 0
    for sheet in sheets:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                fmt = cell.number_format
                if fmt and "(" in fmt and ")" in fmt:
                    new_fmt = fmt.replace("\\(", "-").replace("\\)", "")
                    new_fmt = new_fmt.replace("(", "-").replace(")", "")
                    if new_fmt != fmt:
                        cell.number_format = new_fmt
                        n += 1
    return n


def round_l_lonsamhetskontroll_headers(wb: Workbook) -> int:
    """Round L: Synka sektion-headers på Lönsamhetskontroll till samma bredd.

    iter8 hade headers på olika spann (B:E vs B:G). Merge alla till B:G.
    """
    ws = wb["Lönsamhetskontroll"]
    header_rows = [4, 12, 21, 42, 50]

    # Unmerga eventuella befintliga merges på dessa rader först
    existing = list(ws.merged_cells.ranges)
    for mr in existing:
        if mr.min_row in header_rows and mr.min_row == mr.max_row:
            ws.unmerge_cells(str(mr))

    n = 0
    for r in header_rows:
        try:
            ws.merge_cells(f"B{r}:G{r}")
            n += 1
        except Exception:
            pass

    # Centrera texten i headers
    for r in header_rows:
        ws[f"B{r}"].alignment = ws[f"B{r}"].alignment.copy(horizontal="center")

    return n


def round_m_dokumentation(wb: Workbook) -> int:
    """Round M: Konsekvent sektion-header-stil på Dokumentation.

    Identifierar numrerade huvudsektioner ('1. Syfte', '2. Grundprincip' …)
    och applicerar enhetlig stil (mörkblå bg, vit fet text).
    """
    from tools.theme import apply, ROW_SECTION
    ws = wb["Dokumentation"]
    import re
    main_section_re = re.compile(r"^\d+\.\s+\S")

    n = 0
    for row in ws.iter_rows(min_col=2, max_col=2):
        cell = row[0]
        if isinstance(cell.value, str) and main_section_re.match(cell.value):
            apply(cell, "section")
            try:
                ws.merge_cells(f"B{cell.row}:E{cell.row}")
            except Exception:
                pass
            ws.row_dimensions[cell.row].height = ROW_SECTION
            n += 1

    return n


def round_t_forsattsblad(wb: Workbook) -> int:
    """Round T: NY flik 'Försättsblad' — komplett investeringsmemo enligt LM 371.

    Editorial-stil. Refererar Indata/Resultat/Lönsamhetskontroll via formler.
    Layout:
      Header → Projektinformation → Projektbeskrivning → Bild+Karta →
      Marknadssituation → Kalkylantaganden → Hyresantaganden →
      Finansieringsantaganden → Resultat & avkastning → Kontroll restvärde →
      Investering rekommenderas (signatur) → Bilagor
    """
    from tools.theme import (apply, bottom_rule, INK, MUTED, PAPER, SURFACE,
                              ACCENT, POSITIVE, RULE, FAMILY, FMT_KR,
                              FMT_PCT1, FMT_PCT2, FMT_AR, FMT_M2, FMT_INT)
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.hyperlink import Hyperlink

    # Skapa flik direkt efter Översikt
    if "Försättsblad" in wb.sheetnames:
        del wb["Försättsblad"]
    ws = wb.create_sheet("Försättsblad", index=1)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    # Kolumner: A=margin, B=label, C=value-1, D=value-2, E=value-3, F=spacer,
    #           G=label, H=value-1, I=value-2, J=margin
    widths = {"A": 2, "B": 28, "C": 14, "D": 14, "E": 14,
              "F": 3, "G": 22, "H": 12, "I": 12, "J": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    hairline = Side(style="thin", color=RULE)
    accent_side = Side(style="medium", color=ACCENT)

    def _title(cell_ref, text, size=26, color=INK):
        ws[cell_ref] = text
        ws[cell_ref].font = Font(name="Segoe UI Light", size=size, color=color)
        ws[cell_ref].alignment = Alignment(horizontal="left", vertical="center")

    def _section(row, text):
        """Sektion-rubrik som ALL CAPS muted med tunn rule under."""
        ws[f"B{row}"] = text.upper()
        ws[f"B{row}"].font = Font(name="Segoe UI Semibold", size=10, color=INK, bold=True)
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="bottom", indent=0)
        for col in "BCDEFGHI":
            ws[f"{col}{row}"].border = Border(bottom=Side(style="thin", color=INK))
        ws.row_dimensions[row].height = 22

    def _label(cell, text):
        ws[cell] = text
        ws[cell].font = Font(name=FAMILY, size=10, color=MUTED)
        ws[cell].alignment = Alignment(horizontal="left", vertical="center")

    def _value(cell, value, fmt=None, bold=False, color=INK, italic=False):
        ws[cell] = value
        weight_name = "Segoe UI Semibold" if bold else FAMILY
        ws[cell].font = Font(name=weight_name, size=10, color=color, bold=bold, italic=italic)
        ws[cell].alignment = Alignment(horizontal="right", vertical="center")
        if fmt:
            ws[cell].number_format = fmt

    def _input_marker(cell):
        """Sub-textsmärke 'Fyll i' för fält där värde inte hämtas från Indata."""
        ws[cell] = "Fyll i"
        ws[cell].font = Font(name=FAMILY, size=10, color="0B5F7A", italic=True)
        ws[cell].fill = PatternFill("solid", fgColor=SURFACE)
        ws[cell].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ────────────────────────────────────────────────────────────────────
    # HEADER — luftig hero (samma footprint, mer typografi)
    # ────────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 12

    ws.merge_cells("B2:I2")
    ws["B2"] = "LEJONFASTIGHETER · INVESTERINGSMEMO"
    ws["B2"].font = Font(name="Segoe UI Semibold", size=8, color=ACCENT, bold=True)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Tunn hairline under super-header
    for col in "BCDEFGHI":
        ws[f"{col}3"].border = Border(bottom=Side(style="thin", color=RULE))
    ws.row_dimensions[3].height = 4

    ws.merge_cells("B4:I4")
    _title("B4", "Investeringskalkyl", size=42)
    ws.row_dimensions[4].height = 62

    ws.merge_cells("B5:I5")
    ws["B5"] = '=IF(Indata!B26="","(projekt ej namngivet)",Indata!B26&" · "&IF(Indata!C26="Nyb","Nybyggnad",IF(Indata!C26="Omb","Ombyggnad","Befintligt")))'
    ws["B5"].font = Font(name=FAMILY, size=14, color=MUTED, italic=False)
    ws["B5"].alignment = Alignment(horizontal="left", vertical="top")
    ws.row_dimensions[5].height = 26
    ws.row_dimensions[6].height = 16

    # ────────────────────────────────────────────────────────────────────
    # PROJEKTINFORMATION (rad 7-13)
    # ────────────────────────────────────────────────────────────────────
    _section(7, "Projektinformation")

    info_rows = [
        ("Projektnamn",        "=Indata!B26",         "G", "Kalkylstart",      "=Indata!C5",         FMT_AR),
        ("Fastighet",          None,                  "G", "Kalkylperiod",     "=Indata!C16",        FMT_AR),
        ("Objekt",             None,                  "G", "Total investering","=Indata!R31",        FMT_KR),
        ("Projektnummer",      None,                  "G", "Verksamhetsyta",   "=Indata!F31",        FMT_M2),
        ("Kalkyl utförd av",   None,                  "G", "Typ av projekt",   '=IF(Indata!C26="Nyb","Nybyggnad",IF(Indata!C26="Omb","Ombyggnad","Befintligt"))', None),
        ("Momsregistrering före","=Indata!C12",       "G", "Momsregistrering efter", "=Indata!C13", FMT_PCT1),
    ]
    for i, row in enumerate(info_rows):
        r = 8 + i
        label_l, val_l, _, label_r, val_r, fmt_r = row
        _label(f"B{r}", label_l)
        if val_l is None:
            _input_marker(f"C{r}")
            ws.merge_cells(f"C{r}:E{r}")
        elif isinstance(val_l, str) and val_l.startswith("="):
            _value(f"C{r}", val_l, fmt=FMT_PCT1 if label_l.startswith("Moms") else None)
            ws.merge_cells(f"C{r}:E{r}")
        else:
            _value(f"C{r}", val_l)
            ws.merge_cells(f"C{r}:E{r}")
        _label(f"G{r}", label_r)
        ws.merge_cells(f"H{r}:I{r}")
        _value(f"H{r}", val_r, fmt=fmt_r)
        ws.row_dimensions[r].height = 20

    # ────────────────────────────────────────────────────────────────────
    # PROJEKTBESKRIVNING — tabell (rad 15-25)
    # ────────────────────────────────────────────────────────────────────
    _section(15, "Projektbeskrivning")

    # Tabell-header
    headers = ["Åtgärd", "kvm (BRA)", "Inv belopp", "kr/kvm"]
    for col, h in zip("BCDE", headers):
        ws[f"{col}16"] = h
        ws[f"{col}16"].font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
        ws[f"{col}16"].alignment = Alignment(
            horizontal="left" if col == "B" else "right", vertical="bottom")
        ws[f"{col}16"].border = Border(bottom=Side(style="thin", color=RULE))
    ws.row_dimensions[16].height = 20

    # Tabellrader — refererar Indata!F31 (BRA), Indata!R31 (Inv) för nybyggnad
    project_rows = [
        ("Befintligt",      None, None, None),
        ("Ombyggnad",       None, None, None),
        ("Nybyggnad",       "=Indata!F31", "=Indata!R31", "=IFERROR(Indata!R31/Indata!F31,0)"),
        ("Installationer",  None, None, None),
        ("Mark",            "=Indata!C20", "=Indata!C20*Indata!C21", "=Indata!C21"),
    ]
    for i, (label, area, inv, kvm) in enumerate(project_rows):
        r = 17 + i
        _label(f"B{r}", label)
        if area is not None:
            _value(f"C{r}", area, fmt=FMT_INT)
            _value(f"D{r}", inv, fmt=FMT_KR)
            _value(f"E{r}", kvm, fmt='#,##0 "kr/kvm"')
        else:
            _input_marker(f"C{r}")
            _input_marker(f"D{r}")
            _input_marker(f"E{r}")
        ws.row_dimensions[r].height = 20

    # Summa-rad — Semibold + tunn rule över (LF-stil)
    r_sum = 22
    _label(f"B{r_sum}", "Summa investering")
    ws[f"B{r_sum}"].font = Font(name="Segoe UI Semibold", size=10, color=INK, bold=True)
    ws[f"C{r_sum}"] = "=SUM(C17:C21)"
    ws[f"C{r_sum}"].font = Font(name="Segoe UI Semibold", size=10, color=INK, bold=True)
    ws[f"C{r_sum}"].alignment = Alignment(horizontal="right")
    ws[f"C{r_sum}"].number_format = FMT_INT
    ws[f"D{r_sum}"] = "=Indata!R31"
    ws[f"D{r_sum}"].font = Font(name="Segoe UI Semibold", size=11, color=INK, bold=True)
    ws[f"D{r_sum}"].alignment = Alignment(horizontal="right")
    ws[f"D{r_sum}"].number_format = FMT_KR
    # Tunn rule över summan (hairline, inte tjock)
    for col in "BCDE":
        ws[f"{col}{r_sum}"].border = Border(top=Side(style="thin", color=INK))
    ws.row_dimensions[r_sum].height = 26

    # Fritext-block (projektbeskrivning)
    ws.row_dimensions[23].height = 16
    ws.merge_cells("B24:E24")
    _label("B24", "Projektbeskrivning")
    ws["B24"].font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
    ws.row_dimensions[24].height = 18

    ws.merge_cells("B25:I27")
    ws["B25"] = "[Fyll i en kort beskrivning av projektet — bakgrund, vision, kvalitetsambition, tidplan.]"
    ws["B25"].font = Font(name=FAMILY, size=10, color=MUTED, italic=True)
    ws["B25"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    ws["B25"].fill = PatternFill("solid", fgColor=SURFACE)
    for r in [25, 26, 27]:
        ws.row_dimensions[r].height = 22

    # ────────────────────────────────────────────────────────────────────
    # BILD + KARTA (rad 29-37)
    # ────────────────────────────────────────────────────────────────────
    ws.row_dimensions[28].height = 16
    _section(29, "Visualisering")

    # Bild-placeholder
    ws.merge_cells("B30:E37")
    ws["B30"] = "[Projektunik bild]"
    ws["B30"].font = Font(name=FAMILY, size=10, color=MUTED, italic=True)
    ws["B30"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B30"].fill = PatternFill("solid", fgColor=SURFACE)
    for col in "BCDE":
        for r in range(30, 38):
            ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=SURFACE)

    # Karta-placeholder
    ws.merge_cells("G30:I37")
    ws["G30"] = "[Karta över fastigheten]"
    ws["G30"].font = Font(name=FAMILY, size=10, color=MUTED, italic=True)
    ws["G30"].alignment = Alignment(horizontal="center", vertical="center")
    ws["G30"].fill = PatternFill("solid", fgColor=SURFACE)
    for col in "GHI":
        for r in range(30, 38):
            ws[f"{col}{r}"].fill = PatternFill("solid", fgColor=SURFACE)

    for r in range(30, 38):
        ws.row_dimensions[r].height = 20

    # ────────────────────────────────────────────────────────────────────
    # MARKNADSSITUATION (rad 39-44)
    # ────────────────────────────────────────────────────────────────────
    ws.row_dimensions[38].height = 16
    _section(39, "Marknadssituation")

    _label("B40", "Marknadsvärde, senaste värdering")
    _input_marker("C40")
    ws.merge_cells("C40:E40")
    ws.row_dimensions[40].height = 20

    _label("G40", "Värderingsdatum")
    _input_marker("H40")
    ws.merge_cells("H40:I40")

    # Byggrätter — kompakt tabell
    _label("B42", "Byggrätter")
    ws["B42"].font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
    ws["C42"] = "Före projekt"
    ws["D42"] = "Efter projekt"
    for col in "CD":
        ws[f"{col}42"].font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
        ws[f"{col}42"].alignment = Alignment(horizontal="right")
        ws[f"{col}42"].border = Border(bottom=Side(style="thin", color=RULE))
    ws.row_dimensions[42].height = 18

    _label("B43", "Utnyttjade byggrätter")
    _input_marker("C43"); _input_marker("D43")
    _label("B44", "Outnyttjade byggrätter")
    _input_marker("C44"); _input_marker("D44")
    for r in [43, 44]:
        ws.row_dimensions[r].height = 20

    # ────────────────────────────────────────────────────────────────────
    # KALKYLANTAGANDEN (rad 46-52)
    # ────────────────────────────────────────────────────────────────────
    ws.row_dimensions[45].height = 16
    _section(46, "Kalkylantaganden")

    kalkyl_rows = [
        ("Kalkylperiod",                "=Indata!C16",  FMT_AR,   "Avskrivningstid",      "=Indata!C19", FMT_AR),
        ("Direktavkastningskrav, marknad", "=Indata!C6", FMT_PCT1, "Kalkylränta driftnetto", "=Indata!C7", FMT_PCT1),
        ("Kalkylränta restvärde",       "=Indata!C9",   FMT_PCT1, "Inflation",            "=Indata!C8",  FMT_PCT1),
    ]
    for i, (l1, v1, f1, l2, v2, f2) in enumerate(kalkyl_rows):
        r = 47 + i
        _label(f"B{r}", l1)
        ws.merge_cells(f"C{r}:E{r}")
        _value(f"C{r}", v1, fmt=f1)
        _label(f"G{r}", l2)
        ws.merge_cells(f"H{r}:I{r}")
        _value(f"H{r}", v2, fmt=f2)
        ws.row_dimensions[r].height = 20

    # ────────────────────────────────────────────────────────────────────
    # HYRESANTAGANDEN (rad 51-57)
    # ────────────────────────────────────────────────────────────────────
    ws.row_dimensions[51].height = 16
    _section(52, "Hyresantaganden")

    hyres_rows = [
        ("Bindande kravhyra",      "=Resultat!D14", '#,##0 "kr/år"', "Per kvm/år", "=Resultat!D16", '#,##0 "kr/kvm"'),
        ("Hyresspann lägsta (−)",  "=Resultat!C14", '#,##0 "kr/år"', "Lägsta per kvm","=Resultat!C16", '#,##0 "kr/kvm"'),
        ("Hyresspann högsta (+)",  "=Resultat!E14", '#,##0 "kr/år"', "Högsta per kvm","=Resultat!E16", '#,##0 "kr/kvm"'),
        ("Långsiktig vakansrisk",  "=Indata!C11",   FMT_PCT1,        "Drift & underhåll","=Indata!C41", '#,##0 "kr/kvm"'),
    ]
    for i, (l1, v1, f1, l2, v2, f2) in enumerate(hyres_rows):
        r = 53 + i
        _label(f"B{r}", l1)
        ws.merge_cells(f"C{r}:E{r}")
        _value(f"C{r}", v1, fmt=f1, bold=(i == 0))
        _label(f"G{r}", l2)
        ws.merge_cells(f"H{r}:I{r}")
        _value(f"H{r}", v2, fmt=f2, bold=(i == 0))
        ws.row_dimensions[r].height = 20 if i > 0 else 24

    # ────────────────────────────────────────────────────────────────────
    # FINANSIERINGSANTAGANDEN (rad 58-61)
    # ────────────────────────────────────────────────────────────────────
    ws.row_dimensions[57].height = 16
    _section(58, "Finansieringsantaganden")

    fin_rows = [
        ("Räntenivå (långsiktig)", "=Indata!C62", FMT_PCT2, "Belåningsgrad", "=Indata!C64", FMT_PCT1),
        ("Amortering",             "=Indata!C63", FMT_PCT2, "Avkastningskrav eget kapital", "=Indata!C65", FMT_PCT2),
    ]
    for i, (l1, v1, f1, l2, v2, f2) in enumerate(fin_rows):
        r = 59 + i
        _label(f"B{r}", l1)
        ws.merge_cells(f"C{r}:E{r}")
        _value(f"C{r}", v1, fmt=f1)
        _label(f"G{r}", l2)
        ws.merge_cells(f"H{r}:I{r}")
        _value(f"H{r}", v2, fmt=f2)
        ws.row_dimensions[r].height = 20

    # ────────────────────────────────────────────────────────────────────
    # RESULTAT & AVKASTNING (rad 62-69) — KÄRNAN
    # ────────────────────────────────────────────────────────────────────
    ws.row_dimensions[61].height = 16
    _section(62, "Resultat & avkastning")

    # Stor hero-rad: faktisk IRR vs krav (label muted, värde bär vikt)
    _label("B63", "Faktisk IRR (eget kapital)")
    ws.merge_cells("C63:D63")
    _value("C63", "=Lönsamhetskontroll!C45", fmt=FMT_PCT2, bold=True)
    ws["C63"].font = Font(name="Segoe UI Semibold", size=14, color=INK, bold=True)
    ws["E63"] = '="Krav "&TEXT(Indata!C65,"0.0%")'
    ws["E63"].font = Font(name=FAMILY, size=9, color=MUTED, italic=True)
    ws["E63"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    _label("G63", "Marginal mot krav")
    ws.merge_cells("H63:I63")
    _value("H63", "=Lönsamhetskontroll!C47", fmt=FMT_PCT2, bold=True, color=POSITIVE)
    ws["H63"].font = Font(name="Segoe UI Semibold", size=14, color=POSITIVE, bold=True)
    ws.row_dimensions[63].height = 30

    res_rows = [
        ("NPV (mål-utfall)",              "=Lönsamhetskontroll!C9", FMT_KR, "Krav", ">  0",  None),
        ("PV jmf bokfört värde år 20",    "=Lönsamhetskontroll!C30", FMT_KR, "Krav", "> 0",   None),
        ("Marknadsvärde år 20",           "=Lönsamhetskontroll!C29", FMT_KR, "Bokfört värde år 20", "=Lönsamhetskontroll!C27", FMT_KR),
        ("Bindande kravhyra",             "=Resultat!D14", '#,##0 "kr/år"', "Per kvm/år", "=Resultat!D16", '#,##0 "kr/kvm"'),
    ]
    for i, (l1, v1, f1, l2, v2, f2) in enumerate(res_rows):
        r = 64 + i
        _label(f"B{r}", l1)
        ws.merge_cells(f"C{r}:E{r}")
        _value(f"C{r}", v1, fmt=f1)
        _label(f"G{r}", l2)
        ws.merge_cells(f"H{r}:I{r}")
        _value(f"H{r}", v2, fmt=f2)
        ws.row_dimensions[r].height = 22

    # ────────────────────────────────────────────────────────────────────
    # KONTROLL RESTVÄRDE (rad 69-73)
    # ────────────────────────────────────────────────────────────────────
    ws.row_dimensions[68].height = 16
    _section(69, "Kontroll restvärde")

    kontroll_rows = [
        ("Bokfört värde år 20", "=Lönsamhetskontroll!C27", FMT_KR),
        ("Marknadsvärde år 20", "=Lönsamhetskontroll!C29", FMT_KR),
        ("Differens (MV − BV)", "=Lönsamhetskontroll!C30", FMT_KR),
    ]
    for i, (lbl, val, fmt) in enumerate(kontroll_rows):
        r = 70 + i
        _label(f"B{r}", lbl)
        ws.merge_cells(f"C{r}:E{r}")
        _value(f"C{r}", val, fmt=fmt, bold=(i == 2))
        ws.row_dimensions[r].height = 20

    # ────────────────────────────────────────────────────────────────────
    # INVESTERING REKOMMENDERAS (rad 75-82) — signatur
    # ────────────────────────────────────────────────────────────────────
    ws.row_dimensions[73].height = 16
    ws.row_dimensions[74].height = 8
    _section(75, "Investering rekommenderas")

    # Två signatur-block bredvid varandra (LF-stil — kontraktsformat)
    sig_blocks = [
        ("B", 76, ["C", "D", "E"]),  # Vänster: label B, value-cols C-E
        ("G", 76, ["H", "I"]),       # Höger: label G, value-cols H-I
    ]
    for start_col, start_row, value_cols in sig_blocks:
        first_v, last_v = value_cols[0], value_cols[-1]

        # Namn — input-fält med tunn linje under
        _label(f"{start_col}{start_row}", "Namn")
        ws.merge_cells(f"{first_v}{start_row}:{last_v}{start_row}")
        ws[f"{first_v}{start_row}"].font = Font(name=FAMILY, size=10, color=INK)
        ws[f"{first_v}{start_row}"].alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
        for c in value_cols:
            ws[f"{c}{start_row}"].border = Border(bottom=Side(style="thin", color=INK))
        ws.row_dimensions[start_row].height = 26

        # Datum
        _label(f"{start_col}{start_row+1}", "Datum")
        ws.merge_cells(f"{first_v}{start_row+1}:{last_v}{start_row+1}")
        ws[f"{first_v}{start_row+1}"].font = Font(name=FAMILY, size=10, color=INK)
        ws[f"{first_v}{start_row+1}"].alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
        for c in value_cols:
            ws[f"{c}{start_row+1}"].border = Border(bottom=Side(style="thin", color=INK))
        ws.row_dimensions[start_row+1].height = 26

        # Underskrift
        _label(f"{start_col}{start_row+2}", "Underskrift")
        ws.merge_cells(f"{first_v}{start_row+2}:{last_v}{start_row+2}")
        for c in value_cols:
            ws[f"{c}{start_row+2}"].border = Border(bottom=Side(style="thin", color=INK))
        ws.row_dimensions[start_row+2].height = 32

    # ────────────────────────────────────────────────────────────────────
    # BILAGOR (rad 83-89)
    # ────────────────────────────────────────────────────────────────────
    ws.row_dimensions[81].height = 16
    _section(82, "Bilagor")

    # Tabell-header
    ws["B83"] = "Bilaga"
    ws["B83"].font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
    ws["B83"].alignment = Alignment(horizontal="left", vertical="bottom")
    ws["H83"] = "Flik"
    ws["H83"].font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
    ws["H83"].alignment = Alignment(horizontal="right", vertical="bottom")
    ws.merge_cells("H83:I83")
    for col in "BCDEFGHI":
        ws[f"{col}83"].border = Border(bottom=Side(style="thin", color=RULE))
    ws.row_dimensions[83].height = 20

    bilagor = [
        ("Projektförutsättningar och antaganden", "Indata"),
        ("Kravhyra och hyresspann", "Resultat"),
        ("Driftnetto-projektion 20 år", "Kassaflöde"),
        ("Lån, ränta, amortering, avskrivning", "Finansiering"),
        ("IRR, NPV, marknadsvärde — kontroller", "Lönsamhetskontroll"),
        ("Beräkningsblock A–F — tekniska detaljer", "Beräkningslogik"),
        ("Designprinciper och metod", "Dokumentation"),
    ]
    for i, (text, target) in enumerate(bilagor):
        r = 84 + i
        ws[f"B{r}"] = text
        ws[f"B{r}"].font = Font(name=FAMILY, size=10, color=INK)
        ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"H{r}:I{r}")
        ws[f"H{r}"] = target
        ws[f"H{r}"].font = Font(name=FAMILY, size=10, color=INK, underline="single")
        ws[f"H{r}"].hyperlink = Hyperlink(
            ref=f"H{r}", location=f"'{target}'!A1", display=target)
        ws[f"H{r}"].alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 22

    # Footer
    ws.row_dimensions[91].height = 28
    ws.merge_cells("B92:I92")
    ws["B92"] = "Lejonfastigheter AB · Investeringskalkyl · 2026"
    ws["B92"].font = Font(name=FAMILY, size=8, color=MUTED, italic=True)
    ws["B92"].alignment = Alignment(horizontal="left", vertical="center")

    # Page setup
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.sheet_properties.tabColor = "1B3A6B"

    # Sidfot
    ws.oddFooter.left.text = "Lejonfastigheter · Investeringskalkyl · 2026"
    ws.oddFooter.left.size = 9
    ws.oddFooter.left.color = MUTED
    ws.oddFooter.right.text = "Sida &P / &N"
    ws.oddFooter.right.size = 9
    ws.oddFooter.right.color = MUTED

    return 1


def round_v_indata_polish(wb: Workbook) -> int:
    """Round V: Indata editorial-polish.

    - Sektion-rubriker enhetligt ALL CAPS Semibold + hairline-rule under
    - Ta bort blandning av blå-toner (HYRESOBJEKT-tabellen har mörk turkos bg)
    - Input-celler enhetligt subtil Surface-bg + INK text
    - Tabellrubriker i Muted Semibold med hairline under
    """
    from tools.theme import INK, MUTED, SURFACE, RULE, FAMILY
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb["Indata"]
    n = 0

    # Sektion-headers: 4 (Marknads), 15 (Investering), 24 (Hyresobjekt),
    # 34 (Drift), 43 (Re-investering), 52 (Övriga), 60 (Finansiering), 67 (Hyresspann)
    section_rows = [4, 15, 24, 34, 43, 52, 60, 67]
    for r in section_rows:
        cell = ws.cell(r, 2)
        if cell.value:
            cell.font = Font(name="Segoe UI Semibold", size=10, color=INK, bold=True)
            cell.fill = PatternFill(fill_type=None)
            cell.alignment = Alignment(horizontal="left", vertical="bottom")
            # Hairline under hela raden (B-R range)
            for c in range(2, 19):
                ws.cell(r, c).border = Border(bottom=Side(style="thin", color=INK))
            ws.row_dimensions[r].height = 24
            n += 1

    # HYRESOBJEKT-tabellen: rensa mörka turkos bg och ge stilren editorial
    # Rader 25 (header), 26-30 (data), 31 (summa)
    for r in range(25, 32):
        for c in range(2, 19):
            cell = ws.cell(r, c)
            if cell.fill.fill_type == "solid":
                # Behåll bara Surface-bg på input-celler
                cell.fill = PatternFill(fill_type=None)
                n += 1

    # Tabell-header rad 25: Muted Semibold + hairline under
    for c in range(2, 19):
        cell = ws.cell(25, c)
        if cell.value:
            cell.font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
            cell.border = Border(bottom=Side(style="thin", color=RULE))
            cell.alignment = Alignment(
                horizontal="right" if c > 3 else "left",
                vertical="bottom")
    ws.row_dimensions[25].height = 20

    # Sub-headers i Drift-sektionen rad 35
    for c in range(2, 8):
        cell = ws.cell(35, c)
        if cell.value:
            cell.font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border(bottom=Side(style="thin", color=RULE))
            cell.alignment = Alignment(
                horizontal="right" if c > 2 else "left",
                vertical="bottom")
    ws.row_dimensions[35].height = 20

    # Sub-header Re-investering rad 44
    for c in range(2, 8):
        cell = ws.cell(44, c)
        if cell.value:
            cell.font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border(bottom=Side(style="thin", color=RULE))

    # Summa-rader: rad 31 (Hyresobjekt), 41 (Drift)
    for r in [31, 41]:
        for c in range(2, 19):
            cell = ws.cell(r, c)
            if cell.value is not None:
                cell.font = Font(name="Segoe UI Semibold", size=10, color=INK, bold=True)
                cell.border = Border(top=Side(style="thin", color=INK))
                cell.fill = PatternFill(fill_type=None)
        ws.row_dimensions[r].height = 24

    # Input-celler i C-kolumnen: enhetlig Surface-bg
    for r in range(5, 70):
        cell = ws.cell(r, 3)
        if cell.value is not None and not (isinstance(cell.value, str)
                                            and cell.value.startswith("=")):
            cell.fill = PatternFill("solid", fgColor=SURFACE)
            cell.font = Font(name=FAMILY, size=10, color=INK)
            cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    return n


def round_u_oversikt_v2(wb: Workbook) -> int:
    """Round U: Refaktor Översikt med hero-bild, KPI-band och tidslinje.

    Ersätter befintlig Översikt med editorial dashboard-stil från concept.py.
    Hämtar värden via formler från Resultat/Lönsamhetskontroll/Indata.
    """
    from tools.theme import (INK, MUTED, PAPER, SURFACE, ACCENT, POSITIVE,
                              RULE, FAMILY, FMT_KR, FMT_PCT1, FMT_PCT2,
                              FMT_AR, FMT_M2, FMT_INT)
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    ws = wb["Översikt"]
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False

    # Rensa befintligt innehåll
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    for r in range(1, 60):
        for c in range(1, 15):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.hyperlink = None
        ws.row_dimensions[r].height = 15

    # Kolumnbredder — strikt 90 chars total (landscape A4 = ca 95-100 chars wide)
    widths = {"A": 2, "B": 12, "C": 10, "D": 2,
              "E": 2, "F": 10, "G": 8, "H": 2,
              "I": 2, "J": 14, "K": 10, "L": 2, "M": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Hero-bild
    hero_path = ROOT / "assets" / "hero.png"
    if hero_path.exists():
        img = XLImage(str(hero_path))
        img.width = 1200
        img.height = 380
        ws.add_image(img, "A1")
    for r in range(1, 15):
        ws.row_dimensions[r].height = 20

    # Brand-strip + titel
    ws.row_dimensions[15].height = 8
    ws.merge_cells("B16:D16")
    ws["B16"] = "LEJONFASTIGHETER · INVESTERINGSANALYS"
    ws["B16"].font = Font(name="Segoe UI Semibold", size=8, color=ACCENT, bold=True)
    ws["B16"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[16].height = 16

    ws.merge_cells("B17:L17")
    ws["B17"] = "Investeringskalkyl"
    ws["B17"].font = Font(name="Segoe UI Light", size=42, color=INK)
    ws["B17"].alignment = Alignment(horizontal="left", vertical="bottom")
    ws.row_dimensions[17].height = 52

    ws.merge_cells("B18:L18")
    ws["B18"] = '=Indata!B26&" · Kalkylstart "&Indata!C5'
    ws["B18"].font = Font(name=FAMILY, size=12, color=MUTED)
    ws["B18"].alignment = Alignment(horizontal="left", vertical="top")
    ws.row_dimensions[18].height = 22

    # Separator
    for col in "BCDEFGHIJKL":
        ws[f"{col}19"].border = Border(bottom=Side(style="thin", color=RULE))
    ws.row_dimensions[19].height = 14

    # Story-mening (kort prosa)
    ws.merge_cells("B20:L20")
    ws["B20"] = '="Investering "&TEXT(Indata!R31/1000000,"#,##0")&" Mkr · "&Indata!C16&" år · "&TEXT(Indata!F31,"#,##0")&" m² · Belåningsgrad "&TEXT(Indata!C64,"0%")'
    ws["B20"].font = Font(name=FAMILY, size=11, color=INK)
    ws["B20"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[20].height = 22

    # KPI-band
    def _kpi(start_col, end_col, label, formula, fmt, footer, color=INK):
        ws.merge_cells(f"{start_col}21:{end_col}21")
        ws[f"{start_col}21"] = label
        ws[f"{start_col}21"].font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
        ws[f"{start_col}21"].alignment = Alignment(horizontal="left", vertical="bottom", indent=1)

        ws.merge_cells(f"{start_col}22:{end_col}22")
        ws[f"{start_col}22"] = formula
        ws[f"{start_col}22"].font = Font(name="Segoe UI Light", size=26, color=color)
        ws[f"{start_col}22"].number_format = fmt
        ws[f"{start_col}22"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws.merge_cells(f"{start_col}23:{end_col}23")
        ws[f"{start_col}23"] = footer
        ws[f"{start_col}23"].font = Font(name=FAMILY, size=10, color=MUTED)
        ws[f"{start_col}23"].alignment = Alignment(horizontal="left", vertical="top", indent=1)

    _kpi("B", "D", "BINDANDE KRAVHYRA", "=Resultat!D14", "#,##0", "kr/år")
    _kpi("F", "H", "PER KVM/ÅR",        "=Resultat!D16", "#,##0", "kr/m²")
    _kpi("J", "L", "FAKTISK IRR EK",    "=Lönsamhetskontroll!C45", "0.00%",
         '="Marginal +"&TEXT(Lönsamhetskontroll!C47,"0.00")&" pp · Krav "&TEXT(Indata!C65,"0.0%")',
         color=POSITIVE)

    ws.row_dimensions[21].height = 18
    ws.row_dimensions[22].height = 60
    ws.row_dimensions[23].height = 22
    ws.row_dimensions[24].height = 28

    # Investering-panel (vänster)
    ws.merge_cells("B25:D25")
    ws["B25"] = "INVESTERING"
    ws["B25"].font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
    ws["B25"].alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
    ws.row_dimensions[25].height = 22

    panel_rows = [
        ("Total investering", "=Indata!R31", '#,##0 "kr"'),
        ("Investering / m²",  "=IFERROR(Indata!R31/Indata!F31,0)", '#,##0 "kr"'),
        ("Verksamhetsyta",    "=Indata!F31", '#,##0 "m²"'),
        ("Kalkylperiod",      "=Indata!C16", '0 "år"'),
        ("Belåningsgrad",     "=Indata!C64", "0.0%"),
        ("Avkastningskrav EK","=Indata!C65", "0.0%"),
    ]
    for i, (label, val, fmt) in enumerate(panel_rows):
        r = 27 + i
        ws[f"B{r}"] = label
        ws[f"B{r}"].font = Font(name=FAMILY, size=10, color=MUTED)
        ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(f"C{r}:D{r}")
        ws[f"C{r}"] = val
        ws[f"C{r}"].font = Font(name=FAMILY, size=11, color=INK)
        ws[f"C{r}"].number_format = fmt
        ws[f"C{r}"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws.row_dimensions[r].height = 22
        if i < len(panel_rows) - 1:
            for col in "BCD":
                ws[f"{col}{r}"].border = Border(bottom=Side(style="thin", color=RULE))

    # Driftnetto-tidslinje (höger)
    ws.merge_cells("F25:L25")
    ws["F25"] = "DRIFTNETTO ÖVER KALKYLPERIODEN"
    ws["F25"].font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
    ws["F25"].alignment = Alignment(horizontal="left", vertical="bottom", indent=1)

    # Start
    ws.merge_cells("F27:G27")
    ws["F27"] = "ÅR 1 · =Indata!C5"
    ws["F27"] = '="ÅR 1 · "&Indata!C5'
    ws["F27"].font = Font(name="Segoe UI Semibold", size=8, color=MUTED, bold=True)
    ws["F27"].alignment = Alignment(horizontal="left", vertical="bottom", indent=1)

    ws.merge_cells("F28:G28")
    ws["F28"] = "=Kassaflöde!D23"
    ws["F28"].font = Font(name="Segoe UI Light", size=22, color=MUTED)
    ws["F28"].number_format = "#,##0"
    ws["F28"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("F29:G29")
    ws["F29"] = "kr"
    ws["F29"].font = Font(name=FAMILY, size=9, color=MUTED)
    ws["F29"].alignment = Alignment(horizontal="left", vertical="top", indent=1)

    # Pil i mitten
    ws.merge_cells("H28:I28")
    ws["H28"] = "─────→"
    ws["H28"].font = Font(name=FAMILY, size=14, color=MUTED)
    ws["H28"].alignment = Alignment(horizontal="center", vertical="center")

    # Slut
    ws.merge_cells("J27:L27")
    ws["J27"] = '="ÅR "&Indata!C16&" · "&(Indata!C5+Indata!C16-1)'
    ws["J27"].font = Font(name="Segoe UI Semibold", size=8, color=INK, bold=True)
    ws["J27"].alignment = Alignment(horizontal="left", vertical="bottom", indent=1)

    ws.merge_cells("J28:L28")
    ws["J28"] = "=Kassaflöde!W23"
    ws["J28"].font = Font(name="Segoe UI Light", size=22, color=INK)
    ws["J28"].number_format = "#,##0"
    ws["J28"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("J29:L29")
    ws["J29"] = '="+"&TEXT(Kassaflöde!W23/Kassaflöde!D23-1,"0.0%")&" tillväxt"'
    ws["J29"].font = Font(name="Segoe UI Semibold", size=9, color=POSITIVE, bold=True)
    ws["J29"].alignment = Alignment(horizontal="left", vertical="top", indent=1)

    for r in [27, 28, 29]:
        ws.row_dimensions[r].height = 22 if r != 28 else 40

    # Tidslinje-rule
    for col in "FGHIJKL":
        ws[f"{col}31"].border = Border(top=Side(style="thin", color=INK))
    ws.row_dimensions[31].height = 6

    # Reservera rader
    for r in range(32, 38):
        ws.row_dimensions[r].height = 14

    # Hyresspann-band
    ws.merge_cells("B40:L40")
    ws["B40"] = "HYRESSPANN VID INVESTERINGSUTFALL"
    ws["B40"].font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
    ws["B40"].alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
    ws.row_dimensions[40].height = 22

    # Bar
    for col in "BCDEFGHIJKL":
        ws[f"{col}41"].fill = PatternFill("solid", fgColor=SURFACE)
    ws.merge_cells("F41:H41")
    ws["F41"].fill = PatternFill("solid", fgColor=ACCENT)
    ws.row_dimensions[41].height = 18

    # Etiketter
    ws["B42"] = "=Resultat!C14"
    ws["B42"].font = Font(name=FAMILY, size=10, color=MUTED)
    ws["B42"].number_format = "#,##0"
    ws["B42"].alignment = Alignment(horizontal="left", indent=1)

    ws.merge_cells("F42:H42")
    ws["F42"] = '=TEXT(Resultat!D14,"#,##0")&" kr/år"'
    ws["F42"].font = Font(name="Segoe UI Semibold", size=11, color=INK, bold=True)
    ws["F42"].alignment = Alignment(horizontal="center")

    ws.merge_cells("J42:L42")
    ws["J42"] = "=Resultat!E14"
    ws["J42"].font = Font(name=FAMILY, size=10, color=MUTED)
    ws["J42"].number_format = "#,##0"
    ws["J42"].alignment = Alignment(horizontal="right")
    ws.row_dimensions[42].height = 22

    ws["B43"] = "Lägsta utfall (−10%)"
    ws["B43"].font = Font(name=FAMILY, size=9, color=MUTED, italic=True)
    ws["B43"].alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells("F43:H43")
    ws["F43"] = "Mål-utfall (bindande)"
    ws["F43"].font = Font(name=FAMILY, size=9, color=MUTED, italic=True)
    ws["F43"].alignment = Alignment(horizontal="center")
    ws.merge_cells("J43:L43")
    ws["J43"] = "Högsta utfall (+10%)"
    ws["J43"].font = Font(name=FAMILY, size=9, color=MUTED, italic=True)
    ws["J43"].alignment = Alignment(horizontal="right")
    ws.row_dimensions[43].height = 18

    # Footer-rad
    ws.row_dimensions[44].height = 28
    ws.merge_cells("B45:L45")
    ws["B45"] = "Lejonfastigheter AB · Investeringskalkyl · 2026"
    ws["B45"].font = Font(name=FAMILY, size=8, color=MUTED, italic=True)
    ws["B45"].alignment = Alignment(horizontal="left", indent=1)

    # Page setup — landscape, explicit scale (kringgår render-fit-logic)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 2   # icke-default → render respekterar
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3

    return 1


def round_w_restvardesbedomning(wb: Workbook) -> int:
    """Round W: Restvärdesbedömning — yield-justering ovanpå Gordon-modellen.

    Adderar Indata sektion 9 med:
    - Investeringsgrad-prompt (auto-beräknad, vägleder när justering är meningsfull)
    - 4 dropdowns (Läge / Vakansrisk / Lokalflexibilitet / Byggnadsteknisk standard)
      med 5-gradig skala (±0,25 pp per steg)
    - Motiveringskolumn (fritext)
    - Justerad direktavkastning som referenseras från krav 2+3

    Default = Neutralt på alla fyra → exakt baseline-värden bevaras
    (regressionssäkert mot iter8/iter9).

    Patchar Beräkningslogik C116, C137, C202, C203 + display-celler så
    Gordon-modellen använder justerad yield konsekvent.
    """
    from tools.theme import INK, MUTED, SURFACE, ACCENT, RULE, FAMILY, FMT_PCT2
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb["Indata"]
    n = 0

    # ── Sektion-header rad 70 ───────────────────────────────────────────────
    ws["B70"] = "9. RESTVÄRDESBEDÖMNING"
    ws["B70"].font = Font(name="Segoe UI Semibold", size=10, color=INK, bold=True)
    ws["B70"].fill = PatternFill(fill_type=None)
    ws["B70"].alignment = Alignment(horizontal="left", vertical="bottom")
    for c in range(2, 19):
        ws.cell(70, c).border = Border(bottom=Side(style="thin", color=INK))
    ws.row_dimensions[70].height = 24

    # ── Investeringsgrad-prompt rad 72 ──────────────────────────────────────
    ws["B72"] = "Investeringsgrad"
    ws["B72"].font = Font(name=FAMILY, size=10, color=INK)
    ws["B72"].alignment = Alignment(horizontal="left", vertical="center")

    ws["C72"] = "=IFERROR(R31/(C57+C58+R31),0)"
    ws["C72"].font = Font(name=FAMILY, size=10, color=INK)
    ws["C72"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws["C72"].number_format = "0.0%"
    ws["C72"].fill = PatternFill("solid", fgColor=SURFACE)

    # D72: ingen separat enhet — number_format på C72 visar redan %

    ws["E72"] = (
        '=IF(C72<0.2,"Befintlig fastighet — extern värdering fångar objektet, '
        'justering ej rekommenderad",'
        'IF(C72<0.6,"Större ombyggnad — värderingen är delvis pre-investering, '
        'kalibrering kan vara motiverad",'
        '"Nybyggnation/motsvarande omfattning — kalibrera mot jämförelseobjekt"))'
    )
    ws["E72"].font = Font(name=FAMILY, size=9, color=MUTED, italic=True)
    ws["E72"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells("E72:R72")
    ws.row_dimensions[72].height = 28
    n += 1

    # ── Tabellrubrik rad 74 ─────────────────────────────────────────────────
    headers = [(2, "Parameter"), (3, "Bedömning"), (4, "Effekt"), (5, "Motivering")]
    for col, txt in headers:
        cell = ws.cell(74, col)
        cell.value = txt
        cell.font = Font(name="Segoe UI Semibold", size=9, color=MUTED, bold=True)
        cell.border = Border(bottom=Side(style="thin", color=RULE))
        cell.alignment = Alignment(
            horizontal="right" if col in (3, 4) else "left",
            vertical="bottom",
        )
    ws.merge_cells("E74:R74")
    ws.row_dimensions[74].height = 20

    # ── 4 dropdowns rad 75-78 ───────────────────────────────────────────────
    params = [
        ("Läge (centralitet, infrastruktur)",
         "Närhet till kollektivtrafik, serviceunderlag, demografi"),
        ("Långsiktig vakansrisk",
         "Hyresgästens beroende, alternativa hyresgäster, demografi"),
        ("Lokalflexibilitet (omställbarhet)",
         "Planlösning, bjälklagshöjd, bärande konstruktion"),
        ("Byggnadsteknisk standard",
         "Stomme, klimatskal, installationer, energiprestanda"),
    ]

    dv = DataValidation(
        type="list",
        formula1='"Mycket positivt,Något positivt,Neutralt,Något negativt,Mycket negativt"',
        allow_blank=False,
    )
    dv.error = "Välj från listan"
    dv.errorTitle = "Ogiltigt val"
    dv.prompt = "Bedömning av denna parameter"
    dv.promptTitle = "Restvärdesfaktor"
    ws.add_data_validation(dv)

    for i, (label, hint) in enumerate(params):
        r = 75 + i

        # Parameter-etikett
        ws.cell(r, 2).value = label
        ws.cell(r, 2).font = Font(name=FAMILY, size=10, color=INK)
        ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")

        # Dropdown
        dropdown = ws.cell(r, 3)
        dropdown.value = "Neutralt"
        dropdown.font = Font(name=FAMILY, size=10, color=INK)
        dropdown.alignment = Alignment(horizontal="center", vertical="center")
        dropdown.fill = PatternFill("solid", fgColor=SURFACE)
        dv.add(dropdown)

        # Effekt (formel)
        effekt = ws.cell(r, 4)
        effekt.value = (
            f'=IF(C{r}="Mycket positivt",-0.005,'
            f'IF(C{r}="Något positivt",-0.0025,'
            f'IF(C{r}="Något negativt",0.0025,'
            f'IF(C{r}="Mycket negativt",0.005,0))))'
        )
        effekt.font = Font(name=FAMILY, size=10, color=INK)
        effekt.number_format = '+0.00%;-0.00%;0.00%'
        effekt.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        # Motivering (fritext + hint)
        motiv = ws.cell(r, 5)
        motiv.value = ""
        motiv.font = Font(name=FAMILY, size=9, color=INK)
        motiv.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True, indent=1)
        motiv.fill = PatternFill("solid", fgColor=SURFACE)

        # Hint i kolumn F-R (muted, italic) — visar bedömningsgrunden
        ws.cell(r, 6).value = hint
        ws.cell(r, 6).font = Font(name=FAMILY, size=8, color=MUTED, italic=True)
        ws.cell(r, 6).alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, end_row=r, start_column=5, end_column=5)
        ws.merge_cells(start_row=r, end_row=r, start_column=6, end_column=18)

        ws.row_dimensions[r].height = 22
        n += 1

    # ── Total + yield_adj rad 80-81 ─────────────────────────────────────────
    # Tunn rule ovanför summa
    for col in range(2, 6):
        ws.cell(80, col).border = Border(top=Side(style="thin", color=INK))

    ws["B80"] = "Total justering av yield"
    ws["B80"].font = Font(name="Segoe UI Semibold", size=10, color=INK, bold=True)
    ws["B80"].alignment = Alignment(horizontal="left", vertical="center")

    ws["D80"] = "=SUM(D75:D78)"
    ws["D80"].font = Font(name="Segoe UI Semibold", size=10, color=INK, bold=True)
    ws["D80"].number_format = '+0.00%;-0.00%;0.00%'
    ws["D80"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[80].height = 24

    ws["B81"] = "Justerad direktavkastning"
    ws["B81"].font = Font(name="Segoe UI Semibold", size=10, color=ACCENT, bold=True)
    ws["B81"].alignment = Alignment(horizontal="left", vertical="center")

    ws["D81"] = "=C6+D80"
    ws["D81"].font = Font(name="Segoe UI Semibold", size=11, color=ACCENT, bold=True)
    ws["D81"].number_format = FMT_PCT2
    ws["D81"].alignment = Alignment(horizontal="right", vertical="center", indent=1)

    ws["E81"] = "Används i krav 2 (exit-värde) och krav 3 (MV ≥ BV år 20)"
    ws["E81"].font = Font(name=FAMILY, size=9, color=MUTED, italic=True)
    ws["E81"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("E81:R81")
    ws.row_dimensions[81].height = 26
    n += 1

    # ── Uppdatera print_area så sektion 9 inkluderas (round_i körs före round_w) ─
    ws.print_area = "A1:R81"

    # ── Patcha yield-refs i Beräkningslogik + Lönsamhetskontroll + Försättsblad ─
    yield_ref = "Indata!$D$81"

    patches = [
        ("Beräkningslogik", "C116"),
        ("Beräkningslogik", "C137"),
        ("Beräkningslogik", "C202"),
        ("Beräkningslogik", "C203"),
        ("Lönsamhetskontroll", "C24"),
        ("Försättsblad", "C48"),
    ]
    for sheet, coord in patches:
        if sheet not in wb.sheetnames:
            continue
        cell = wb[sheet][coord]
        if isinstance(cell.value, str) and cell.value.startswith("="):
            new_formula = cell.value.replace("Indata!C6", yield_ref)
            if new_formula != cell.value:
                cell.value = new_formula
                n += 1

    # Block D + E (krav 2 IRR EK): rad 172 (exit-värde 0 hyra) och 187 (exit-värde 1Mkr)
    # Kolumnerna D:AB är år 1-25 — varje kolumn har egen formel med Indata!$C$6
    blk = wb["Beräkningslogik"]
    from openpyxl.utils import get_column_letter
    for r in (172, 187):
        for col_idx in range(4, 29):  # D..AB
            coord = f"{get_column_letter(col_idx)}{r}"
            cell = blk[coord]
            if isinstance(cell.value, str) and "Indata!$C$6" in cell.value:
                cell.value = cell.value.replace("Indata!$C$6", "Indata!$D$81")
                n += 1

    return n


def round_w_guide(wb: Workbook) -> int:
    """Round W (forts): Restvärdesguide i Beräkningslogik.

    Pedagogiskt textavsnitt som förklarar de sex resonemangen för restvärde
    och varför Gordon med yield-kalibrering är valt. Adderas efter sista
    befintliga block. Refererar Indata sektion 9 för kalibreringsmekaniken.
    """
    from tools.theme import INK, MUTED, ACCENT, FAMILY
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    ws = wb["Beräkningslogik"]
    start = ws.max_row + 3  # två tomma rader marginal

    # Header
    ws.cell(start, 2).value = "RESTVÄRDETS LOGIK — RESONEMANG OCH KALIBRERING"
    ws.cell(start, 2).font = Font(name="Segoe UI Semibold", size=11, color=INK, bold=True)
    ws.cell(start, 2).alignment = Alignment(horizontal="left", vertical="bottom")
    for c in range(2, 31):
        ws.cell(start, c).border = Border(bottom=Side(style="thin", color=INK))
    ws.row_dimensions[start].height = 26

    intro = (
        "Restvärdet år 20 är ett antagande, inte ett faktum. Det finns flera "
        "rimliga sätt att resonera om vad fastigheten är värd vid kalkylperiodens slut. "
        "Mallen använder Gordon-modellen (driftnetto år 21 / direktavkastning) med "
        "kalibrering via Indata sektion 9. De andra resonemangen nedan är sanity checks "
        "när Gordon-resultatet ifrågasätts eller behöver kompletteras."
    )
    ws.cell(start + 2, 2).value = intro
    ws.cell(start + 2, 2).font = Font(name=FAMILY, size=10, color=INK)
    ws.cell(start + 2, 2).alignment = Alignment(horizontal="left", vertical="top",
                                                wrap_text=True)
    ws.merge_cells(start_row=start + 2, end_row=start + 2,
                   start_column=2, end_column=18)
    ws.row_dimensions[start + 2].height = 60

    resonemang = [
        ("1. Gordon — kapitaliserad evig driftnetto (modellens default)",
         "Driftnetto år 21 / justerad direktavkastning. Antar att fastigheten kan "
         "drivas vidare evigt och säljas till marknadsmässig avkastning. Den extern "
         "yielden (Indata C6) kommer från värdering (befintlig fastighet) eller "
         "jämförelseobjekt (nybyggnation). Yield-justeringen i sektion 9 kalibrerar "
         "objektets långsiktiga värdebeständighet mot snittet."),
        ("2. Avskrivet bokfört värde",
         "Anskaffning minus ackumulerade avskrivningar. Försiktigt — speglar inte "
         "marknad, bara redovisning. Användbart som golv eller sanity check, inte "
         "som primärt exit-värde. LM 371 viktade 60% mot detta — borttaget iter 7."),
        ("3. Restproduktionsvärde",
         "Vad kostar det att uppföra motsvarande byggnad år 20, med hänsyn till "
         "kvarvarande teknisk livslängd. Frikopplar från driftnetto. Underskattar "
         "ofta läges- och kontraktsvärde."),
        ("4. Mark + nedskriven byggnad",
         "Markvärde plus byggnad till lågt restvärde. Pessimistiskt scenario — "
         "antar att byggnaden närmar sig slutet av sin ekonomiska livslängd. "
         "Lämpligt för specialbyggnader med svag omställbarhet."),
        ("5. Förhandlat övertagandevärde",
         "Kommunen eller offentlig hyresgäst löser ut fastigheten år 20 till "
         "förhandlat pris (ofta bokfört värde). Specialfall där exit till privat "
         "marknad inte är planerad."),
        ("6. Likvidations- eller rivningsvärde",
         "Markvärde minus rivningskostnad. Värsta fall, golv för känslighetsanalys."),
    ]

    row = start + 4
    for title, body in resonemang:
        ws.cell(row, 2).value = title
        ws.cell(row, 2).font = Font(name="Segoe UI Semibold", size=10, color=ACCENT, bold=True)
        ws.cell(row, 2).alignment = Alignment(horizontal="left", vertical="bottom")
        ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=18)
        ws.row_dimensions[row].height = 20

        ws.cell(row + 1, 2).value = body
        ws.cell(row + 1, 2).font = Font(name=FAMILY, size=10, color=INK)
        ws.cell(row + 1, 2).alignment = Alignment(horizontal="left", vertical="top",
                                                  wrap_text=True)
        ws.merge_cells(start_row=row + 1, end_row=row + 1,
                       start_column=2, end_column=18)
        ws.row_dimensions[row + 1].height = 56

        row += 3

    # Uppdatera print_area så guiden inkluderas
    from openpyxl.utils import get_column_letter
    last_col = min(30, ws.max_column)
    ws.print_area = f"A1:{get_column_letter(last_col)}{row}"

    return 1


def round_w_forsattsblad(wb: Workbook) -> int:
    """Round W (forts): Yield-justering på Försättsblad — Kontroll restvärde.

    Lägger en rad i Kontroll restvärde-sektionen (rad 69) som visar
    aktuell yield-justering — gör det tydligt på memo-fliken att Gordon
    kalibrerats (eller är neutral).
    """
    from tools.theme import INK, MUTED, ACCENT, FAMILY, FMT_KR
    from openpyxl.styles import Font, Alignment

    if "Försättsblad" not in wb.sheetnames:
        return 0

    ws = wb["Försättsblad"]

    # Befintlig "Kontroll restvärde"-sektion: rad 69 (header), 70-72 (data).
    # Vi använder rad 73 (var tom buffer) för yield-justeringsraden.
    ws["B73"] = "Yield-justering (restvärde)"
    ws["B73"].font = Font(name=FAMILY, size=10, color=MUTED)
    ws["B73"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("C73:E73")
    ws["C73"] = (
        '=IF(Indata!D80=0,"Neutral (extern yield rakt av)",'
        'TEXT(Indata!D80*100,"+0.00;-0.00")&" pp · justerad yield "'
        '&TEXT(Indata!D81,"0.00%"))'
    )
    ws["C73"].font = Font(name=FAMILY, size=10, color=ACCENT, italic=True)
    ws["C73"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[73].height = 22

    return 1


def round_x_kanslighet(wb: Workbook) -> int:
    """Round X: Känslighetstabell för restvärdesbedömning.

    Tre scenarier på Lönsamhetskontroll (rad 73+) med:
    - Yield-justering (pp)
    - Justerad direktavkastning
    - Bindande kravhyra (analytisk via MAX av tre krav)
    - Bindande krav (NPV/IRR/MV)
    - Marknadsvärde år 20
    - Differens MV − BV

    Faktisk IRR EK visas endast för bedömt scenario (kräver full
    EK-cashflow-rekonstruktion för opt/pess som inte är analytiskt
    görbart i ren formel).

    Yields per scenario:
      Optimistisk: extern yield − 1,00 pp
      Bedömt: justerad yield (Indata!D81, från sektion 9)
      Pessimistisk: extern yield + 1,00 pp
    """
    from tools.theme import INK, MUTED, ACCENT, SURFACE, RULE, FAMILY, FMT_KR, FMT_PCT2
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb["Lönsamhetskontroll"]
    start = 73  # efter befintliga sektioner

    # ── Header rad 73 ───────────────────────────────────────────────────────
    ws.cell(start, 2).value = "KÄNSLIGHETSANALYS — RESTVÄRDESBEDÖMNING"
    ws.cell(start, 2).font = Font(name="Segoe UI Semibold", size=11, color=INK, bold=True)
    ws.cell(start, 2).alignment = Alignment(horizontal="left", vertical="bottom")
    try:
        ws.merge_cells(start_row=start, end_row=start, start_column=2, end_column=7)
    except Exception:
        pass
    for c in range(2, 8):
        ws.cell(start, c).border = Border(bottom=Side(style="thin", color=INK))
    ws.row_dimensions[start].height = 26

    # Intro
    ws.cell(start + 1, 2).value = (
        "Hur ändras bindande kravhyra och marknadsvärde om yielden slår "
        "±1,00 pp jämfört med basbedömningen? Bedömt scenario = justeringen "
        "från Indata sektion 9."
    )
    ws.cell(start + 1, 2).font = Font(name=FAMILY, size=10, color=MUTED, italic=True)
    ws.cell(start + 1, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    try:
        ws.merge_cells(start_row=start + 1, end_row=start + 1, start_column=2, end_column=7)
    except Exception:
        pass
    ws.row_dimensions[start + 1].height = 32

    # ── Scenario-header rad 75 ──────────────────────────────────────────────
    hdr = start + 2
    headers = [(3, "Optimistisk"), (4, "Bedömt"), (5, "Pessimistisk")]
    for col, txt in headers:
        cell = ws.cell(hdr, col)
        cell.value = txt
        cell.font = Font(name="Segoe UI Semibold", size=10, color=INK, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="bottom")
        cell.border = Border(bottom=Side(style="thin", color=INK))
    # Underrubrik
    ws.cell(hdr + 1, 3).value = "yield − 1,00 pp"
    ws.cell(hdr + 1, 4).value = "Indata sektion 9"
    ws.cell(hdr + 1, 5).value = "yield + 1,00 pp"
    for col in (3, 4, 5):
        ws.cell(hdr + 1, col).font = Font(name=FAMILY, size=9, color=MUTED, italic=True)
        ws.cell(hdr + 1, col).alignment = Alignment(horizontal="center", vertical="top")
    ws.row_dimensions[hdr].height = 22
    ws.row_dimensions[hdr + 1].height = 18

    # ── Datarader 77+ ───────────────────────────────────────────────────────
    # Yield-referenser per scenario
    yields = {
        "C": "(Indata!$C$6-0.01)",       # Optimistisk
        "D": "Indata!$D$81",              # Bedömt
        "E": "(Indata!$C$6+0.01)",       # Pessimistisk
    }

    def scenario_formula(template: str, col: str) -> str:
        """Substituera {Y} med yield-referens för scenariot."""
        return template.replace("{Y}", yields[col])

    # Rad-definitioner: (label, template_or_callable, fmt, bold, hint)
    # Konstanter från Beräkningslogik:
    #   C115 = PV driftnetto vid 0 hyra (yield-oberoende)
    #   C118 = PV investering (yield-oberoende)
    #   C136 = PV driftnetto vid 1 Mkr hyra (yield-oberoende)
    #   C177 = NPV EK vid 0 hyra (vid nuvarande D81)
    #   C192 = NPV EK vid 1 Mkr hyra (vid nuvarande D81)
    #   C199 = Driftnetto år 21 vid 0 hyra
    #   C200 = Driftnetto år 21 vid 1 Mkr hyra
    #   C204 = Bokfört värde år 20

    # Krav 1: NPV ≥ 0  →  hyra = -NPV_0(y) / b_NPV(y)
    #   NPV_0(y) = C115 + C199/y/(1+C9)^C16 + C118
    #   NPV_1(y) = C136 + C200/y/(1+C9)^C16 + C118
    krav1_template = (
        "=-(Beräkningslogik!C115+Beräkningslogik!C199/{Y}/(1+Indata!C9)^Indata!C16+Beräkningslogik!C118)"
        "/((Beräkningslogik!C136-Beräkningslogik!C115+(Beräkningslogik!C200-Beräkningslogik!C199)/{Y}"
        "/(1+Indata!C9)^Indata!C16)/1000000)"
    )

    # Krav 3: MV ≥ BV  →  hyra = (BV*y - DN21_0) / ((DN21_1-DN21_0)/1e6)
    krav3_template = (
        "=(Beräkningslogik!C204*{Y}-Beräkningslogik!C199)"
        "*1000000/(Beräkningslogik!C200-Beräkningslogik!C199)"
    )

    # Krav 2: NPV_EK_0(y) = C177 + DN82_21*(1/y - 1/D81)/(1+C65)^C16
    #   där DN82_21 = INDEX(Beräkningslogik!D82:AB82, MATCH(C5+C16, Beräkningslogik!D52:AB52,0))
    #        DN132_21 = INDEX(Beräkningslogik!D132:AB132, MATCH(C5+C16, ...,0))
    #   krav 2 hyra = -NPV_EK_0(y) / b_IRR(y)
    dn82_21 = ("INDEX(Beräkningslogik!D82:AB82,MATCH(Indata!C5+Indata!C16,"
               "Beräkningslogik!D52:AB52,0))")
    dn132_21 = ("INDEX(Beräkningslogik!D132:AB132,MATCH(Indata!C5+Indata!C16,"
                "Beräkningslogik!D52:AB52,0))")
    disc_ek = "(1+Indata!C65)^Indata!C16"
    npv_ek_0 = (f"(Beräkningslogik!C177+{dn82_21}*(1/{{Y}}-1/Indata!$D$81)/{disc_ek})")
    npv_ek_1 = (f"(Beräkningslogik!C192+{dn132_21}*(1/{{Y}}-1/Indata!$D$81)/{disc_ek})")
    krav2_template = f"=-{npv_ek_0}/(({npv_ek_1}-{npv_ek_0})/1000000)"

    # MV år 20 vid bindande kravhyra:
    #   MV = (DN21_0 + b_DN21 * h_bind) / y
    #   där b_DN21 = (C200-C199)/1e6
    #   h_bind = scenariots egna bindande hyra (referensera C{row} där bindande står)

    rows = []

    # Yield-justering (pp)
    rows.append((
        "Yield-justering",
        {"C": "=-0.01", "D": "=Indata!D80", "E": "=0.01"},
        '+0.00"pp";-0.00"pp";0.00"pp"',
        False,
    ))
    # Justerad direktavkastning
    rows.append((
        "Justerad direktavkastning",
        {col: f"={yields[col]}" for col in "CDE"},
        FMT_PCT2,
        False,
    ))
    # Krav 1 — bedömt = Resultat!D10 (exakt), opt/pess = analytisk
    rows.append((
        "Kravhyra NPV ≥ 0",
        {
            "C": scenario_formula(krav1_template, "C"),
            "D": "=Resultat!D10",
            "E": scenario_formula(krav1_template, "E"),
        },
        '#,##0',
        False,
    ))
    # Krav 2 — bedömt = Resultat!D11 (exakt), opt/pess = analytisk
    rows.append((
        "Kravhyra IRR ≥ avk.krav",
        {
            "C": scenario_formula(krav2_template, "C"),
            "D": "=Resultat!D11",
            "E": scenario_formula(krav2_template, "E"),
        },
        '#,##0',
        False,
    ))
    # Krav 3 — bedömt = Resultat!D12 (exakt), opt/pess = analytisk
    rows.append((
        "Kravhyra MV ≥ BV",
        {
            "C": scenario_formula(krav3_template, "C"),
            "D": "=Resultat!D12",
            "E": scenario_formula(krav3_template, "E"),
        },
        '#,##0',
        False,
    ))
    # Bindande kravhyra = MAX
    data_start = start + 4  # första datarad efter header+intro+scenariorad+underrubrik
    bind_row = data_start + 5  # = MAX rad (efter 5 ovanstående)

    def bind_max(col):
        # Refererar de tre föregående raderna i samma kolumn
        return f"=MAX({col}{data_start+2},{col}{data_start+3},{col}{data_start+4})"

    rows.append((
        "Bindande kravhyra",
        {col: bind_max(col) for col in "CDE"},
        '#,##0',
        True,
    ))
    # Bindande krav (NPV/IRR/MV)
    def bind_label(col):
        return (f'=IF({col}{bind_row}={col}{data_start+2},"NPV",'
                f'IF({col}{bind_row}={col}{data_start+3},"IRR","MV"))')
    rows.append((
        "Bindande krav",
        {col: bind_label(col) for col in "CDE"},
        None,
        False,
    ))
    # MV år 20 vid bindande kravhyra
    def mv_formula(col):
        Y = yields[col]
        return (f"=(Beräkningslogik!C199+(Beräkningslogik!C200-Beräkningslogik!C199)"
                f"/1000000*{col}{bind_row})/{Y}")
    rows.append((
        "Marknadsvärde år 20",
        {col: mv_formula(col) for col in "CDE"},
        '#,##0',
        False,
    ))
    # Differens MV − BV
    def diff_formula(col):
        return f"={col}{bind_row + 2}-Beräkningslogik!C204"
    rows.append((
        "Differens MV − BV",
        {col: diff_formula(col) for col in "CDE"},
        '#,##0',
        False,
    ))

    # Skriv ut raderna
    surface_fill = PatternFill("solid", fgColor=SURFACE)
    for i, (label, vals, fmt, bold) in enumerate(rows):
        r = data_start + i

        ws.cell(r, 2).value = label
        ws.cell(r, 2).font = Font(
            name="Segoe UI Semibold" if bold else FAMILY,
            size=10, color=INK, bold=bold)
        ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")

        for col_letter in "CDE":
            col_idx = ord(col_letter) - ord("A") + 1
            cell = ws.cell(r, col_idx)
            cell.value = vals[col_letter]
            cell.font = Font(
                name="Segoe UI Semibold" if bold else FAMILY,
                size=11 if bold else 10,
                color=ACCENT if (bold and col_letter == "D") else INK,
                bold=bold)
            cell.alignment = Alignment(horizontal="center" if fmt is None else "right",
                                       vertical="center", indent=1)
            if fmt is not None:
                cell.number_format = fmt
            if col_letter == "D":  # bedömt = highlight
                cell.fill = surface_fill
        ws.row_dimensions[r].height = 26 if bold else 22

        # Tunn rule ovanför "Bindande kravhyra"
        if label == "Bindande kravhyra":
            for c in range(2, 6):
                ws.cell(r, c).border = Border(top=Side(style="thin", color=INK))

    # Not om faktisk IRR EK
    note_row = data_start + len(rows) + 1
    ws.cell(note_row, 2).value = (
        "Faktisk IRR EK per scenario kräver full EK-cashflow-rekonstruktion "
        "och visas inte här. Se Lönsamhetskontroll C45 för bedömt scenario. "
        "Om krav 2 (IRR) är bindande är IRR EK exakt = avkastningskravet; "
        "om krav 1 eller 3 är bindande är IRR EK strikt högre än kravet."
    )
    ws.cell(note_row, 2).font = Font(name=FAMILY, size=9, color=MUTED, italic=True)
    ws.cell(note_row, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    try:
        ws.merge_cells(start_row=note_row, end_row=note_row, start_column=2, end_column=7)
    except Exception:
        pass
    ws.row_dimensions[note_row].height = 42

    # Print_area utökas
    ws.print_area = f"A1:G{note_row}"

    return len(rows)


def main() -> int:
    out = ROOT / "build" / "iter9.xlsx"
    out.parent.mkdir(exist_ok=True)

    print(f"1. Laddar {ITER8.name}")
    wb = load_iter(ITER8)

    print("2. Round E: 'år N' → 'år 20' / 'år N+1' → 'år 21'")
    n = round_e_year_n(wb)
    print(f"   {n} celler patchade")

    print("3. Round B: pedagogisk omskrivning av Beräkningslogik-text")
    n = round_b_pedagogisk(wb)
    print(f"   {n} celler patchade")

    print("4. Round C: Indata-fält renamning (branschterminologi)")
    n = round_c_indata_rename(wb)
    print(f"   {n} celler patchade")

    print("5. Round F: designcleanup Lönsamhetskontroll")
    n = round_f_lonsamhetskontroll_cleanup(wb)
    print(f"   {n} rader/celler patchade")

    print("6. Round G: Översikt om till beslutsdokument")
    n = round_g_oversikt_redesign(wb)
    print(f"   {n} sektioner skrivna")

    print("7. Round H: finansmodell-styling (färg, format, typografi)")
    style_oversikt(wb["Översikt"])
    style_indata(wb["Indata"])
    style_resultat(wb["Resultat"])
    print("   styling klar")

    print("8a. Round I: print area & page setup")
    n = round_i_page_setup(wb)
    print(f"    {n} flikar konfigurerade")

    print("8b. Round J: Översikt header-fixes")
    n = round_j_oversikt_fixes(wb)
    print(f"    {n} ändringar")

    print("8c. Round K: talformat-konsistens (parens → minus)")
    n = round_k_format_konsistens(wb)
    print(f"    {n} celler reformaterade")

    print("8d. Round L: Lönsamhetskontroll header-bredd")
    n = round_l_lonsamhetskontroll_headers(wb)
    print(f"    {n} headers mergade")

    print("8e. Round M: Dokumentation header-konsistens")
    n = round_m_dokumentation(wb)
    print(f"    {n} sektioner stylade")

    print("8f. Round O: globalt LF-tema (gridlines av, fliktag färgad)")
    n = round_o_global_theme(wb)
    print(f"    {n} flikar uppdaterade")

    print("8g. Round P: zebra & bottom-rules")
    n = round_p_zebra_oversikt(wb)
    print(f"    {n} celler stylade")

    print("8h. Round Q: hero-block för nyckelresultat (gul accent)")
    n = round_q_hero_block(wb)
    print(f"    {n} hero-celler")

    print("8i. Round R: status-pillar")
    n = round_r_status_pills(wb)
    print(f"    {n} pillar")

    print("8j. Round S: sidfot")
    n = round_s_footer(wb)
    print(f"    {n} flikar")

    print("8k. Round U: refaktor Översikt med hero + KPI:er")
    n = round_u_oversikt_v2(wb)
    print(f"    Översikt v2 byggd")

    print("8l. Round T: ny flik Försättsblad (LM 371-spegling)")
    n = round_t_forsattsblad(wb)
    print(f"    Försättsblad byggd")

    print("8m. Round V: Indata styling-polish (enhetlig editorial)")
    n = round_v_indata_polish(wb)
    print(f"    {n} cells uppdaterade")

    print("8n. Round W: Restvärdesbedömning (sektion 9 + yield-patches)")
    n = round_w_restvardesbedomning(wb)
    print(f"    {n} celler uppdaterade")

    print("8o. Round W: Restvärdesguide i Beräkningslogik")
    n = round_w_guide(wb)
    print(f"    guide skriven")

    print("8p. Round W: Yield-justering på Försättsblad")
    n = round_w_forsattsblad(wb)
    print(f"    {n} rad uppdaterad")

    print("8q. Round X: Känslighetstabell — restvärdesbedömning")
    n = round_x_kanslighet(wb)
    print(f"    {n} rader byggda")

    print(f"\n9. Sparar → {out.name}")
    save_iter(wb, out)

    print("\n10. XML-patch summaryBelow=False (Lönsamhetskontroll)")
    patch_summary_below(out, ["sheet6.xml"])
    print("    patch klar")

    print("\n11. Recalc (Excel COM / LibreOffice)")
    recalc(out)

    print("\n12. Regressionstest:")
    res = check_baseline(out)
    if res["rent"] is not None:
        print(f"    Hyra={res['rent']:,.2f}  IRR={res['irr']:.4%}  margin={res['margin']*100:+.2f} pp")
    if res["ok"]:
        print("    ✓ Regression OK")
        return 0
    print("    ✗ FAIL:", res["fails"])
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
